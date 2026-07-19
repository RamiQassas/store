import json
import uuid
from decimal import Decimal
import os
import random
import string
import pyotp
import qrcode
import io
import base64
from datetime import timedelta

from django.contrib import messages
from django.contrib.auth import authenticate, login, logout, update_session_auth_hash
from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator
from django.db import transaction
from django.db.models import Q, Count
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse
from django.utils import timezone
from django.conf import settings
from django.core.signing import TimestampSigner, SignatureExpired, BadSignature

from apps.accounts.models import User, OTPToken, KYCRequest, KYCSettings, ActivityLog
from apps.accounts.services import send_brevo_email
from apps.catalog.models import Category, Product, ProductVariant, ProductSuggestion, ProductKey, ProductImage
from apps.common.models import Currency, SocialMediaLink, SiteAnnouncement
from apps.notifications.models import Notification, NotificationSetting
from apps.notifications.services import notify_bulk, notify_user
from apps.orders.models import Order, OrderLog, Coupon, OrderItem
from apps.payments.models import DepositRequest, PaymentMethod, WithdrawalRequest
from apps.site.forms import (
    LoginForm, RegisterForm, PaymentMethodForm, CurrencyForm, ModerateUserForm, 
    ProductForm, CategoryForm, KYCRequestForm, KYCSettingsForm, ChangePasswordForm, 
    CouponForm, SendNotificationForm, AdminChatForm, SiteAnnouncementForm, 
    ChatCannedReplyForm, SupportSettingsForm, ProductSuggestionForm, NotificationSettingForm,
    TestimonialForm, PlatformStatisticForm
)
from apps.support.models import ChatRoom, ChatMessage, ChatCannedReply, SupportSettings
from apps.wallets.models import Wallet
from apps.wallets.services import (
    get_or_create_wallet, track_pending_deposit, freeze_funds, credit_wallet,
    debit_wallet, finalize_withdrawal, release_funds
)
from apps.common.decorators import staff_required, admin_required, support_required, finance_required, kyc_required

signer = TimestampSigner()

# ==========================================
# --- AUTHENTICATION HELPERS (V3) ---
# ==========================================

def v3_generate_otp(user, purpose):
    OTPToken.objects.filter(user=user, purpose=purpose, is_used=False).update(is_used=True)
    code = ''.join(random.choices(string.digits, k=6))
    expires_at = timezone.now() + timedelta(minutes=10)
    return OTPToken.objects.create(user=user, code=code, purpose=purpose, expires_at=expires_at)

def v3_send_otp_email(user, otp_token):
    from apps.common.tenant_utils import get_current_store
    active_store = getattr(user, 'store', None) or get_current_store()
    store_name = active_store.name if active_store else "رقميات"
    store_brand = active_store.name if active_store else "رقميات | RAQAMIYAT"
    
    subject = f"{otp_token.code} هو رمز التحقق الخاص بك | {store_name}"
    purpose_text = "لتفعيل حسابك" if otp_token.purpose == OTPToken.Purpose.REGISTRATION else \
                   "لتسجيل الدخول" if otp_token.purpose == OTPToken.Purpose.LOGIN else \
                   "لإتمام العملية"
    
    html_content = f"""
    <div dir="rtl" style="font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif; max-width: 500px; margin: 0 auto; padding: 20px; color: #1e293b; background-color: #ffffff; border-radius: 16px; border: 1px solid #e2e8f0;">
        <div style="text-align: center; margin-bottom: 30px;">
            <h2 style="color: #06b6d4; margin: 0; font-size: 24px; font-weight: 900;">{store_brand}</h2>
        </div>
        <div style="background-color: #f8fafc; padding: 30px; border-radius: 12px; text-align: center;">
            <p style="font-size: 16px; margin-bottom: 10px; color: #64748b;">رمز التحقق الخاص بك {purpose_text}:</p>
            <h1 style="font-size: 42px; font-weight: 900; color: #0f172a; margin: 0; letter-spacing: 10px;">{otp_token.code}</h1>
            <p style="font-size: 12px; margin-top: 20px; color: #94a3b8;">هذا الرمز صالح لمدة 10 دقائق فقط. لا تشارك هذا الرمز مع أي شخص.</p>
        </div>
        <div style="margin-top: 30px; padding-top: 20px; border-top: 1px solid #f1f5f9; text-align: center;">
            <p style="font-size: 12px; color: #94a3b8; line-height: 1.6;">إذا لم تطلب هذا الرمز، يمكنك تجاهل هذا البريد الإلكتروني.<br>© 2026 {store_name} - جميع الحقوق محفوظة.</p>
        </div>
    </div>
    """
    import threading
    threading.Thread(
        target=send_brevo_email,
        kwargs={
            "to_email": user.email,
            "to_name": user.get_full_name() or user.email,
            "subject": subject,
            "html_content": html_content,
            "store": active_store
        }
    ).start()
    return True

def v3_verify_otp_logic(user, code, purpose):
    otp = OTPToken.objects.filter(user=user, code=code, purpose=purpose, is_used=False, expires_at__gt=timezone.now()).first()
    if otp:
        otp.is_used = True; otp.save(update_fields=["is_used", "updated_at"]); return True
    return False

def v3_security_redirect(methods):
    if not methods: return redirect("dashboard")
    m = methods[0]
    if m == "APP": return redirect("site_2fa_verify")
    if m == "SP": return redirect("site_sp_verify")
    return redirect("site_verify_otp")

def v3_get_required_methods(user, action_type):
    method_map = {
        "login": user.security_login_method,
        "deposit": user.security_deposit_method,
        "purchase": user.security_purchase_method,
        "withdraw": user.security_withdraw_method,
        "settings": "SP" if user.security_password_enabled else "NONE"
    }
    pref = method_map.get(action_type, "NONE")
    
    # Force Email OTP if never verified (one-time requirement for new accounts)
    if action_type == "login" and not user.email_verified:
        return ["EMAIL"]
        
    if pref == "NONE": return []
    if pref == "EMAIL": return ["EMAIL"]
    if pref == "APP":
        return ["APP"] if user.totp_enabled else ["EMAIL"]
    if pref == "BOTH":
        return ["APP", "EMAIL"] if user.totp_enabled else ["EMAIL"]
    if pref == "SP":
        return ["SP"] if user.security_password else ["EMAIL"]
    return []

def v3_init_verification(request, user, action_type):
    # Clear any old pending action to prevent cross-contamination
    if "v3_pending_action_id" in request.session:
        del request.session["v3_pending_action_id"]
        
    # Check if user has any security methods enabled
    methods = v3_get_required_methods(user, action_type)
    if not methods: 
        return True # No security methods configured, bypass
    
    # Check for 5-minute cooldown
    last_verified = request.session.get("v3_action_verified_at")
    if last_verified:
        try:
            if (timezone.now() - timezone.datetime.fromisoformat(last_verified)).total_seconds() < 300:
                return True # Within 5-minute cooldown, bypass
        except:
            pass
            
    # Proceed with verification
    request.session["v3_auth_uid"] = str(user.id)
    request.session["v3_auth_methods"] = methods
    request.session["v3_auth_purpose"] = action_type
    request.session["v3_auth_next"] = request.get_full_path()
    
    first_method = methods[0]
    if first_method == "EMAIL":
        otp_sent = v3_send_otp_email(user, v3_generate_otp(user, action_type))
        request.session["v3_otp_sent_success"] = otp_sent
    else:
        request.session["v3_otp_sent_success"] = True
        
    return False # Always return False to trigger redirect to verification

def v3_redirect_to_verification(request, methods):
    if not methods: 
        next_url = request.session.get("v3_auth_next", "dashboard")
        return redirect(next_url)
    first = methods[0]
    if first == "EMAIL": return redirect("site_verify_otp")
    if first == "APP": return redirect("site_2fa_verify")
    if first == "SP": return redirect("site_sp_verify")
    next_url = request.session.get("v3_auth_next", "dashboard")
    return redirect(next_url)

def clean_verification_session(request):
    """Cleans up authentication session keys after successful verification."""
    keys = ["v3_auth_uid", "v3_auth_methods", "v3_auth_purpose", "v3_new_email", "v3_pending_action_id", "v3_auth_next", "v3_pending_purchase", "v3_otp_sent_success"]
    for k in keys:
        if k in request.session:
            del request.session[k]

def complete_pending_financial_action(request, user):
    """
    Checks for any pending financial action (deposit or withdrawal) in the session
    and marks it as verified. Triggers notifications and messages.
    """
    pending_action_id = request.session.get("v3_pending_action_id")
    if pending_action_id:
        from apps.payments.models import DepositRequest, WithdrawalRequest
        
        # Check Deposit
        deposit = DepositRequest.objects.filter(id=pending_action_id, user=user).first()
        if deposit:
            deposit.is_verified = True
            deposit.save(update_fields=["is_verified"])
            
            # Notify user
            send_financial_notification(
                user=user,
                title="تم استلام طلب الإيداع",
                body=f"تم استلام طلب الإيداع الخاص بك رقم {deposit.id} بقيمة {deposit.amount} {deposit.currency.code}. سيتم مراجعته من قبل الإدارة قريباً."
            )
            # Notify staff
            from apps.notifications.services import notify_staff
            notify_staff(
                title="طلب إيداع جديد",
                body=f"قام {user.email} بتقديم طلب إيداع بقيمة {deposit.amount} {deposit.currency.code}",
                action_url=f"/control/deposits/{deposit.id}/",
                category='admin_new_deposit'
            )
            
            messages.success(request, "تم التحقق وتقديم طلب الإيداع بنجاح.")
            del request.session["v3_pending_action_id"]
            
            # Determine next URL
            next_url = request.session.get("v3_auth_next", "dashboard_deposits")
            clean_verification_session(request)
            return redirect(next_url)
            
        # Check Withdrawal
        withdrawal = WithdrawalRequest.objects.filter(id=pending_action_id, user=user).first()
        if withdrawal:
            withdrawal.is_verified = True
            withdrawal.save(update_fields=["is_verified"])
            
            # Notify user
            send_financial_notification(
                user=user,
                title="تم استلام طلب السحب",
                body=f"تم استلام طلب السحب الخاص بك رقم {withdrawal.id} بقيمة {withdrawal.amount} {withdrawal.currency.code}. سيتم مراجعته من قبل الإدارة قريباً."
            )
            # Notify staff
            from apps.notifications.services import notify_staff
            notify_staff(
                title="طلب سحب جديد",
                body=f"قام {user.email} بتقديم طلب سحب بقيمة {withdrawal.amount} {withdrawal.currency.code}",
                action_url=f"/control/withdrawals/{withdrawal.id}/",
                category='admin_new_withdrawal'
            )
            
            messages.success(request, "تم التحقق وتقديم طلب السحب بنجاح.")
            del request.session["v3_pending_action_id"]
            
            # Determine next URL
            next_url = request.session.get("v3_auth_next", "dashboard_withdrawals")
            clean_verification_session(request)
            return redirect(next_url)
            
    return None

def complete_pending_purchase(request, user):
    """
    Checks for any pending purchase in the session and completes it using create_order service.
    """
    pending_purchase = request.session.get("v3_pending_purchase")
    if pending_purchase:
        variant_id = pending_purchase.get("variant_id")
        coupon_code = pending_purchase.get("coupon_code")
        metadata = pending_purchase.get("metadata", {})
        
        from apps.orders.services import create_order
        from apps.catalog.models import ProductVariant
        
        variant = get_object_or_404(ProductVariant, id=variant_id)
        coupon = None
        if coupon_code:
            coupon = Coupon.objects.filter(code__iexact=coupon_code).first()
            
        try:
            order = create_order(
                customer=user,
                variant_id=variant_id,
                quantity=1,
                coupon=coupon,
                metadata=metadata,
                shipping_name=pending_purchase.get("shipping_name"),
                shipping_phone=pending_purchase.get("shipping_phone"),
                shipping_address=pending_purchase.get("shipping_address"),
            )
            
            # Success
            messages.success(request, "تم إتمام الشراء بنجاح.")
            
            # Clear pending purchase session keys
            del request.session["v3_pending_purchase"]
            clean_verification_session(request)
            
            # If variant is auto-delivery keys, redirect to order detail so they see the keys immediately!
            if variant.delivery_type == 'keys':
                return redirect(reverse("dashboard_order_detail", kwargs={"pk": order.id}))
            else:
                return redirect("dashboard_orders")
                
        except Exception as e:
            messages.error(request, f"فشل إتمام الشراء: {str(e)}")
            # Clear pending purchase session keys to prevent loop
            if "v3_pending_purchase" in request.session:
                del request.session["v3_pending_purchase"]
            clean_verification_session(request)
            # Redirect back to the product details page
            return redirect(reverse("product_detail", kwargs={"pk": variant.product.id}))
            
    return None

def v3_verify_sp_view(request):
    uid, purpose = request.session.get("v3_auth_uid"), request.session.get("v3_auth_purpose")
    if not uid: return redirect("site_login")
    user = get_object_or_404(User, id=uid)
    methods = request.session.get("v3_auth_methods", ["SP"])
    
    if request.method == "POST":
        password = request.POST.get("password")
        from django.contrib.auth.hashers import check_password
        if user.security_password and check_password(password, user.security_password):
            remaining = [m for m in methods if m != "SP"]
            request.session["v3_auth_methods"] = remaining
            
            if remaining:
                if remaining[0] == "EMAIL":
                    otp_sent = v3_send_otp_email(user, v3_generate_otp(user, purpose))
                    request.session["v3_otp_sent_success"] = otp_sent
                return v3_redirect_to_verification(request, remaining)
            
            # All verified
            if not request.user.is_authenticated:
                user.backend = 'apps.stores.auth_backend.TenantModelBackend'
                login(request, user)
            
            # Set grace period for BOTH keys
            now_iso = timezone.now().isoformat()
            request.session["v3_action_verified_at"] = now_iso
            request.session["v3_sp_verified_at"] = now_iso
            
            # Complete pending purchase or financial action
            purchase_redirect = complete_pending_purchase(request, user)
            if purchase_redirect:
                return purchase_redirect
                
            financial_redirect = complete_pending_financial_action(request, user)
            if financial_redirect:
                return financial_redirect

            next_url = request.session.get("v3_auth_next")
            clean_verification_session(request)
                
            if next_url: return redirect(next_url)
            return redirect("control_dashboard" if (user.is_staff and getattr(request, 'store', None) is None) else "dashboard")
            
        messages.error(request, "كلمة مرور الحماية غير صحيحة.")
        
    return render(request, "site/v3/v3_sp_verify.html", {"purpose": purpose})

# ==========================================
# --- AUTH VIEWS (V3) ---
# ==========================================

def v3_login_view(request):
    active_store = getattr(request, 'store', None)

    if request.user.is_authenticated:
        return redirect("control_dashboard" if (request.user.is_staff and getattr(request, 'store', None) is None) else "dashboard")
    form = LoginForm(request.POST or None)
    if request.method == "POST" and form.is_valid():
        user = authenticate(request, username=form.cleaned_data["email"], password=form.cleaned_data["password"])
        if user:
            if not user.is_active:
                messages.error(request, "الحساب معطل.")
                return render(request, "site/v3/v3_login.html", {"form": form})

            # Multi-Tenant isolation:
            # - On main platform (active_store=None): reject users linked to any store
            # - On store tenant (active_store set): only allow users linked to THIS store
            if active_store is None:
                # Main platform: users with store_id cannot login here (except store owners)
                if user.store_id is not None:
                    from apps.common.tenant_utils import bypass_tenant_filter
                    with bypass_tenant_filter():
                        is_owner = user.owned_stores.exists()
                    if not is_owner:
                        messages.error(request, "هذا الحساب مرتبط بمتجر فرعي ولا يمكنه تسجيل الدخول هنا. يرجى التوجه إلى صفحة تسجيل الدخول الخاصة بمتجرك.")
                        return render(request, "site/v3/v3_login.html", {"form": form})
            else:
                # Store tenant: only allow users who belong to this store (or superusers)
                if not user.is_superuser:
                    from apps.stores.models import StoreEmployee
                    from apps.common.tenant_utils import bypass_tenant_filter
                    with bypass_tenant_filter():
                        is_store_member = (
                            user.store_id == active_store.pk or
                            StoreEmployee.objects.filter(store=active_store, user=user).exists() or
                            active_store.owner_id == user.pk
                        )
                    if not is_store_member:
                        messages.error(request, "هذا الحساب غير مرتبط بهذا المتجر.")
                        return render(request, "site/v3/v3_login.html", {"form": form})

            # Ensure backend is set for session authentication
            user.backend = 'apps.stores.auth_backend.TenantModelBackend'

            # Check security settings only if user exists
            if v3_init_verification(request, user, "login"):
                login(request, user)
                return redirect("control_dashboard" if (user.is_staff and getattr(request, 'store', None) is None) else "dashboard")

            methods = request.session.get("v3_auth_methods", [])
            if not methods:
                login(request, user)
                return redirect("dashboard")

            # If it's a login action, we don't want to redirect back to login page after verification
            if request.session.get("v3_auth_purpose") == "login":
                request.session["v3_auth_next"] = reverse("dashboard")

            return v3_redirect_to_verification(request, methods)
        messages.error(request, "بيانات الدخول غير صحيحة.")
    return render(request, "site/v3/v3_login.html", {"form": form})

def v3_register_view(request):
    if request.user.is_authenticated: return redirect("dashboard")
    form = RegisterForm(request.POST or None)
    if request.method == "POST" and form.is_valid():
        try:
            with transaction.atomic():
                user = User.objects.create_user(
                    email=form.cleaned_data["email"],
                    password=form.cleaned_data["password"],
                    phone=request.POST.get("phone"),
                    first_name=form.cleaned_data["first_name"],
                    last_name=form.cleaned_data["last_name"],
                    store=getattr(request, 'store', None)
                )
                get_or_create_wallet(user)
                otp = v3_generate_otp(user, OTPToken.Purpose.REGISTRATION)
                if v3_send_otp_email(user, otp):
                    request.session["v3_auth_uid"], request.session["v3_auth_purpose"] = str(user.id), OTPToken.Purpose.REGISTRATION
                    return redirect("site_verify_otp")
                else: raise Exception("فشل إرسال البريد الإلكتروني.")
        except Exception as e: form.add_error(None, str(e))
    return render(request, "site/v3/v3_register.html", {"form": form})

def v3_verify_otp_view(request):
    uid, purpose = request.session.get("v3_auth_uid"), request.session.get("v3_auth_purpose")
    if not uid: return redirect("site_login")
    user = get_object_or_404(User, id=uid)
    methods = request.session.get("v3_auth_methods", ["EMAIL"])
    
    # Check if the initial OTP send failed
    otp_sent_success = request.session.get("v3_otp_sent_success", True)
    
    settings_obj = KYCSettings.get_settings()
    current_cooldown_limit = min(settings_obj.otp_base_cooldown * (2 ** user.otp_resend_count), 600)
    last_otp = OTPToken.objects.filter(user=user, purpose=purpose).order_by("-created_at").first()
    remaining_cooldown = 0
    if last_otp and otp_sent_success:
        seconds_passed = (timezone.now() - last_otp.created_at).total_seconds()
        if seconds_passed < current_cooldown_limit: remaining_cooldown = int(current_cooldown_limit - seconds_passed)
    is_locked = False
    if user.otp_lockout_until and user.otp_lockout_until > timezone.now(): is_locked = True
    
    if not otp_sent_success:
        messages.error(request, "فشل إرسال رمز التحقق إلى بريدك الإلكتروني. يرجى المحاولة مرة أخرى.")
        # Clear the flag so the message doesn't keep appearing, and let them retry
        request.session["v3_otp_sent_success"] = True
        
    if request.method == "POST":
        if is_locked: return render(request, "site/v3/v3_otp_verify.html", {"user_email": user.email, "remaining_cooldown": remaining_cooldown, "is_locked": True})
        action = request.POST.get("action")
        if action == "resend":
            if remaining_cooldown > 0:
                messages.error(request, f"يرجى الانتظار {remaining_cooldown} ثانية.")
                return render(request, "site/v3/v3_otp_verify.html", {"user_email": user.email, "remaining_cooldown": remaining_cooldown})
            otp = v3_generate_otp(user, purpose)
            if v3_send_otp_email(user, otp):
                user.otp_resend_count += 1; user.save()
                messages.success(request, "تم إعادة إرسال الرمز.")
                return redirect("site_verify_otp")
        code = request.POST.get("code")
        if v3_verify_otp_logic(user, code, purpose):
            user.otp_failed_attempts = 0; user.otp_lockout_until = None; user.otp_resend_count = 0; user.save()
            if purpose in [OTPToken.Purpose.REGISTRATION, "login"]: user.email_verified = True; user.save()
            if purpose == "email_change":
                new_email = request.session.get("v3_new_email")
                if new_email:
                    user.email = new_email; user.username = new_email; user.save()
                    messages.success(request, "تم تغيير البريد الإلكتروني بنجاح.")
            
            remaining = [m for m in methods if m != "EMAIL"]
            request.session["v3_auth_methods"] = remaining
            
            if remaining:
                return v3_redirect_to_verification(request, remaining)
            
            if not request.user.is_authenticated:
                user.backend = 'apps.stores.auth_backend.TenantModelBackend'
                login(request, user)
            
            now_iso = timezone.now().isoformat()
            request.session["v3_action_verified_at"] = now_iso
            request.session["v3_sp_verified_at"] = now_iso

            # Complete pending purchase or financial action
            purchase_redirect = complete_pending_purchase(request, user)
            if purchase_redirect:
                return purchase_redirect
                
            financial_redirect = complete_pending_financial_action(request, user)
            if financial_redirect:
                return financial_redirect

            next_url = request.session.get("v3_auth_next")
            clean_verification_session(request)
            
            if next_url: return redirect(next_url)
            return redirect("control_dashboard" if (user.is_staff and getattr(request, 'store', None) is None) else "dashboard")
        user.otp_failed_attempts += 1
        if user.otp_failed_attempts >= settings_obj.otp_max_attempts:
            user.otp_lockout_until = timezone.now() + timedelta(minutes=15)
            user.otp_failed_attempts = 0 
        user.save()
        messages.error(request, "رمز التحقق غير صحيح.")
    return render(request, "site/v3/v3_otp_verify.html", {"user_email": user.email, "remaining_cooldown": remaining_cooldown, "is_locked": is_locked})

def v3_2fa_verify_view(request):
    uid, purpose = request.session.get("v3_auth_uid"), request.session.get("v3_auth_purpose")
    if not uid: return redirect("site_login")
    user = get_object_or_404(User, id=uid)
    methods = request.session.get("v3_auth_methods", ["APP"])
    if request.method == "POST":
        code = request.POST.get("code")
        if user.totp_secret and pyotp.TOTP(user.totp_secret).verify(code):
            remaining = [m for m in methods if m != "APP"]
            request.session["v3_auth_methods"] = remaining
            if remaining and remaining[0] == "EMAIL":
                otp_sent = v3_send_otp_email(user, v3_generate_otp(user, purpose))
                request.session["v3_otp_sent_success"] = otp_sent
                return redirect("site_verify_otp")
            
            if not request.user.is_authenticated:
                user.backend = 'apps.stores.auth_backend.TenantModelBackend'
                login(request, user)
                
            request.session["v3_action_verified_at"] = timezone.now().isoformat()
            
            # Complete pending purchase or financial action
            purchase_redirect = complete_pending_purchase(request, user)
            if purchase_redirect:
                return purchase_redirect
                
            financial_redirect = complete_pending_financial_action(request, user)
            if financial_redirect:
                return financial_redirect

            next_url = request.session.get("v3_auth_next")
            clean_verification_session(request)
            
            if next_url: return redirect(next_url)
            return redirect("control_dashboard" if (user.is_staff and getattr(request, 'store', None) is None) else "dashboard")
        messages.error(request, "الرمز غير صحيح.")
    return render(request, "site/v3/v3_2fa_verify.html")

@login_required
def v3_2fa_setup_view(request):
    if not v3_init_verification(request, request.user, "settings"):
        return v3_security_redirect(request.session.get("v3_auth_methods", []))
        
    user = request.user
    if user.totp_enabled:
        if request.method == "POST" and request.POST.get("action") == "disable":
            user.totp_enabled = False; user.totp_secret = None; user.save()
            messages.success(request, "تم تعطيل المصادقة الثنائية."); return redirect("site_2fa_setup")
        return render(request, "site/v3/v3_2fa_setup.html", {"enabled": True})
    if not user.totp_secret: user.totp_secret = pyotp.random_base32(); user.save()
    totp = pyotp.TOTP(user.totp_secret); provisioning_url = totp.provisioning_uri(name=user.email, issuer_name="Raqamiyat")
    qr = qrcode.QRCode(version=1, box_size=10, border=5); qr.add_data(provisioning_url); qr.make(fit=True)
    img = qr.make_image(fill_color="black", back_color="white"); buffered = io.BytesIO(); img.save(buffered, format="PNG")
    qr_base64 = base64.b64encode(buffered.getvalue()).decode()
    if request.method == "POST":
        if totp.verify(request.POST.get("code")):
            user.totp_enabled = True; user.save(); messages.success(request, "تم تفعيل المصادقة الثنائية بنجاح."); return redirect("dashboard")
        messages.error(request, "الرمز غير صحيح.")
    return render(request, "site/v3/v3_2fa_setup.html", {"qr_base64": qr_base64, "secret": user.totp_secret, "enabled": False})

def v3_forgot_password_view(request):
    if request.method == "POST":
        email = request.POST.get("email", "").lower().strip()
        user = User.objects.filter(email=email).first()
        if user:
            from django.contrib.auth.tokens import default_token_generator
            token = default_token_generator.make_token(user)
            
            # Determine store branding dynamically
            active_store = getattr(request, "store", None)
            store_name = active_store.name if active_store else "رقميات"
            store_brand = active_store.name if active_store else "رقميات | RAQAMIYAT"
            
            # Build base URL depending on context
            if active_store:
                base_url = f"{request.scheme}://{request.get_host()}"
            else:
                base_url = settings.SITE_URL
                
            reset_url = urljoin(base_url, reverse('site_reset_password')) + f"?token={token}&uid={user.id}"
            subject = f"رابط استعادة كلمة المرور | {store_name}"
            html_content = f"""
            <div dir="rtl" style="font-family: 'Segoe UI', sans-serif; max-width: 500px; margin: 0 auto; padding: 20px; border: 1px solid #e2e8f0; border-radius: 16px; background-color: #ffffff;">
                <h2 style="color: #06b6d4; text-align: center;">{store_brand}</h2>
                <p>مرحباً،</p>
                <p>لقد طلبت إعادة تعيين كلمة المرور لحسابك. يرجى الضغط على الزر أدناه للمتابعة:</p>
                <div style="text-align: center; margin: 30px 0;">
                    <a href="{reset_url}" style="display: inline-block; padding: 14px 30px; background-color: #06b6d4; color: white; text-decoration: none; border-radius: 12px; font-weight: bold;">إعادة تعيين كلمة المرور</a>
                </div>
                <p style="font-size: 12px; color: #94a3b8; text-align: center;">هذا الرابط صالح لمدة 10 دقائق فقط. إذا لم تطلب هذا، يمكنك تجاهل هذا البريد.</p>
                <hr style="border: 0; border-top: 1px solid #f1f5f9; margin: 20px 0;">
                <p style="font-size: 10px; color: #cbd5e1; text-align: center;">© 2026 {store_name} - جميع الحقوق محفوظة.</p>
            </div>
            """
            if send_brevo_email(user.email, user.get_full_name() or user.email, subject, html_content, store=active_store):
                return render(request, "site/v3/v3_forgot_password.html", {"sent": True})
            else:
                messages.error(request, "فشل إرسال البريد الإلكتروني. يرجى المحاولة لاحقاً.")
        else:
            messages.error(request, "البريد الإلكتروني غير مسجل لدينا.")
    return render(request, "site/v3/v3_forgot_password.html")

def v3_reset_password_view(request):
    token = request.GET.get("token") or request.POST.get("token")
    uid = request.GET.get("uid") or request.POST.get("uid")
    if not token or not uid: return redirect("site_forgot_password")
    user = get_object_or_404(User, id=uid)
    from django.contrib.auth.tokens import default_token_generator
    if not default_token_generator.check_token(user, token):
        messages.error(request, "الرابط غير صالح أو منتهي الصلاحية."); return redirect("site_forgot_password")
    if request.method == "POST":
        p1, p2 = request.POST.get("password"), request.POST.get("confirm_password")
        if p1 and p1 == p2 and len(p1) >= 10:
            user.set_password(p1); user.save(); 
            login(request, user, backend='apps.stores.auth_backend.TenantModelBackend')
            messages.success(request, "تم تغيير كلمة المرور بنجاح. تم تسجيل دخولك تلقائياً.")
            return redirect("dashboard")
        messages.error(request, "كلمات المرور غير متطابقة أو لا تستوفي شروط الطول (10 خانات على الأقل).")
    return render(request, "site/v3/v3_reset_password.html", {"user_email": user.email, "token": token, "uid": uid})

@login_required
def v3_logout_view(request):
    logout(request); return redirect("site_login")

@login_required
def resend_verification(request):
    if request.user.email_verified:
        messages.info(request, "بريدك الإلكتروني موثق بالفعل.")
        return redirect("dashboard")
    
    from apps.accounts.services import send_verification_otp
    try:
        if send_verification_otp(request.user):
            messages.success(request, "تم إعادة إرسال رمز التحقق إلى بريدك الإلكتروني.")
            return redirect("site_verify_otp")
        else:
            messages.error(request, "فشل إرسال الرمز. يرجى المحاولة مرة أخرى لاحقاً.")
    except Exception as e:
        messages.error(request, f"خطأ: {str(e)}")
        
    return redirect("dashboard")
def email_verify(request, uidb64, token): return redirect("site_login")

# ==========================================
# --- USER VIEWS (V3) ---
# ==========================================

@login_required
def dashboard(request):
    print(f"[dashboard VIEW DEBUG] User: {request.user}, is_authenticated: {request.user.is_authenticated}")
    # Multi-Tenant: Order.objects uses TenantManager, so it filters by store automatically.
    # On main site (store=None): returns orders where store=None (platform orders).
    # On store tenant: returns only that store's orders for this customer.
    wallet = get_or_create_wallet(request.user)
    active_store = getattr(request, 'store', None)

    # Handle store management actions on main platform
    if request.method == "POST" and not active_store:
        action = request.POST.get("action")
        store_id = request.POST.get("store_id")
        
        from apps.stores.models import Store, SubscriptionPlan
        from apps.common.tenant_utils import bypass_tenant_filter
        
        with bypass_tenant_filter():
            store = Store.unfiltered.filter(id=store_id, owner=request.user).first()
            
        if store:
            if action == "renew_store":
                success = store.renew_subscription()
                if success:
                    messages.success(request, f"تم تجديد اشتراك متجر '{store.name}' بنجاح.")
                else:
                    messages.error(request, f"فشل تجديد الاشتراك. يرجى التحقق من رصيد محفظتك.")
            
            elif action == "toggle_auto_renew":
                store.auto_renew = not store.auto_renew
                store.save()
                messages.success(request, f"تم {'تفعيل' if store.auto_renew else 'إلغاء تفعيل'} التجديد التلقائي لمتجر '{store.name}' بنجاح.")
                
            elif action == "change_billing_cycle":
                cycle = request.POST.get("billing_cycle")
                if cycle in ["monthly", "yearly"]:
                    store.billing_cycle = cycle
                    store.save()
                    messages.success(request, f"تم تغيير دورة دفع متجر '{store.name}' إلى: {'سنوي' if cycle == 'yearly' else 'شهري'}.")
                else:
                    messages.error(request, "دورة الدفع المحددة غير صالحة.")
                    
            elif action == "change_plan":
                plan_id = request.POST.get("plan_id")
                plan = SubscriptionPlan.objects.filter(id=plan_id, is_active=True).first()
                if plan:
                    store.subscription_plan = plan
                    store.save()
                    messages.success(request, f"تم تغيير باقة متجر '{store.name}' إلى: {plan.name}.")
                else:
                    messages.error(request, "الباقة المحددة غير صالحة.")
                    
            return redirect("dashboard")

    # Orders filtered by current tenant context automatically
    digital_deliveries = Order.objects.filter(
        customer=request.user, status=Order.Status.COMPLETED, is_delivery_read=False
    ).exclude(fulfillment_data={})
    recent_orders = Order.objects.filter(customer=request.user).order_by('-created_at')[:5]

    ctx = {
        "wallet": wallet,
        "digital_deliveries": digital_deliveries,
        "orders": recent_orders,
        "notifications": Notification.objects.filter(user=request.user, is_read=False)[:5],
    }

    if not active_store:
        # Main platform: show deposits, withdrawals, KYC, and owned stores
        from apps.stores.models import Store, SubscriptionPlan
        from apps.common.tenant_utils import bypass_tenant_filter
        ctx["recent_deposits"] = DepositRequest.objects.filter(user=request.user).order_by('-created_at')[:5]
        ctx["recent_withdrawals"] = WithdrawalRequest.objects.filter(user=request.user).order_by('-created_at')[:5]
        ctx["kyc_request"] = KYCRequest.objects.filter(user=request.user).first()
        with bypass_tenant_filter():
            ctx["user_stores"] = list(Store.unfiltered.filter(owner=request.user))
        ctx["deposits"] = ctx["recent_deposits"]
        ctx["withdrawals"] = ctx["recent_withdrawals"]
        ctx["plans"] = list(SubscriptionPlan.objects.filter(is_active=True).order_by("price_monthly"))
    else:
        # Store tenant: no deposits/withdrawals/KYC in store context
        ctx["recent_deposits"] = []
        ctx["recent_withdrawals"] = []
        ctx["kyc_request"] = None
        ctx["user_stores"] = []
        ctx["deposits"] = []
        ctx["withdrawals"] = []

    return render(request, "site/v3/v3_dashboard.html", ctx)

@login_required
def wallet_page(request):
    from django.db.models import Sum
    from apps.payments.models import DepositRequest, WithdrawalRequest
    from apps.wallets.models import RechargeCard
    
    request.user.reset_daily_limits_if_needed()
    wallet = Wallet.objects.filter(user=request.user).select_related("currency").first() or get_or_create_wallet(request.user)
    
    # Calculate totals
    total_deposited = DepositRequest.objects.filter(
        user=request.user, 
        status=DepositRequest.Status.COMPLETED, 
        is_verified=True
    ).aggregate(total=Sum('wallet_amount'))['total'] or Decimal("0.00")
    
    total_withdrawn = WithdrawalRequest.objects.filter(
        user=request.user, 
        status=WithdrawalRequest.Status.COMPLETED, 
        is_verified=True
    ).aggregate(total=Sum('wallet_amount'))['total'] or Decimal("0.00")

    total_recharged = RechargeCard.objects.filter(
        redeemed_by=request.user,
        status=RechargeCard.Status.REDEEMED
    ).aggregate(total=Sum('amount'))['total'] or Decimal("0.00")
    
    show_all = request.GET.get("show_all") == "1"
    ledger_entries = wallet.ledger_entries.all()
    if not show_all:
        ledger_entries = ledger_entries[:20]
        
    return render(request, "site/wallet.html", {
        "wallet": wallet, 
        "ledger_entries": ledger_entries,
        "show_all": show_all,
        "total_deposited": total_deposited,
        "total_withdrawn": total_withdrawn,
        "total_recharged": total_recharged,
        "total_added": total_deposited + total_recharged
    })

@login_required
def orders_list(request):
    # Multi-Tenant: Order.objects filtered by TenantManager automatically.
    # On store tenant: returns only orders placed within that store.
    orders = Order.objects.filter(customer=request.user).prefetch_related('items__variant__product').order_by('-created_at')
    return render(request, "site/orders_list.html", {"orders": orders})

@login_required
def order_detail(request, pk):
    # Multi-Tenant: filtered by customer + TenantManager (store context)
    order = get_object_or_404(
        Order.objects.filter(customer=request.user).prefetch_related('items__variant__product', 'logs'),
        pk=pk
    )
    return render(request, "site/order_detail.html", {"order": order})

from django.contrib.auth.hashers import make_password, check_password as check_password_hash

# ... (rest of imports)

@login_required
def transfer_page(request):
    from apps.wallets.services import execute_p2p_transfer
    from django.core.exceptions import ValidationError
    
    settings_obj = KYCSettings.get_settings()
    if not settings_obj.p2p_transfer_enabled:
        messages.error(request, "ميزة التحويل معطلة حالياً.")
        return redirect("dashboard")
        
    wallet = get_or_create_wallet(request.user)
    
    if request.method == "POST":
        recipient_uid = request.POST.get("recipient_uid")
        amount = Decimal(request.POST.get("amount", "0"))
        note = request.POST.get("note", "")
        
        active_store = getattr(request, 'store', None) or request.user.store
        recipient = User.objects.filter(public_uuid=recipient_uid, store=active_store).first()
        if not recipient:
            messages.error(request, "المستلم غير موجود أو لا ينتمي لهذا المتجر.")
            return redirect("dashboard_transfer")
            
        try:
            transfer = execute_p2p_transfer(
                sender=request.user,
                recipient=recipient,
                amount=amount,
                currency=wallet.currency,
                note=note
            )
            messages.success(request, f"تم تحويل {transfer.net_amount} {transfer.currency.code} إلى {recipient.display_name} بنجاح.")
            return redirect("dashboard_transfer_history")
        except ValidationError as e:
            messages.error(request, str(e.message) if hasattr(e, 'message') else str(e))
            
    # Calculate limits in USD
    limit_usd = request.user.custom_p2p_transfer_limit if request.user.has_custom_limits and request.user.custom_p2p_transfer_limit else (settings_obj.verified_transfer_limit if request.user.is_kyc_verified else settings_obj.unverified_transfer_limit)
    
    # Calculate sent transfers today in USD
    from django.db.models import Sum
    from apps.wallets.models import BalanceTransfer
    from django.utils import timezone
    today_start = timezone.localtime(timezone.now()).replace(hour=0, minute=0, second=0, microsecond=0)
    sent_transfers_today = BalanceTransfer.objects.filter(
        sender=request.user,
        status=BalanceTransfer.Status.COMPLETED,
        created_at__gte=today_start
    )
    daily_usage_usd = Decimal("0.00")
    for tr in sent_transfers_today:
        daily_usage_usd += tr.currency.to_base(tr.amount, "withdraw")
        
    remaining_limit_usd = max(Decimal("0.00"), limit_usd - daily_usage_usd)
    
    # Convert limits/usage to the active wallet currency for UI rendering
    limit_wallet = wallet.currency.from_base(limit_usd, "withdraw")
    remaining_wallet = wallet.currency.from_base(remaining_limit_usd, "withdraw")
    daily_usage_wallet = wallet.currency.from_base(daily_usage_usd, "withdraw")

    return render(request, "site/v3/v3_transfer.html", {
        "wallet": wallet,
        "settings": settings_obj,
        "fee_percent": settings_obj.transfer_fee_percent,
        "preferred_currency": request.user.preferred_currency or wallet.currency,
        "currencies": Currency.objects.filter(is_active=True),
        "limit_usd": limit_usd,
        "remaining_limit_usd": remaining_limit_usd,
        "daily_limit": limit_wallet,
        "remaining_limit": remaining_wallet,
        "daily_usage": daily_usage_wallet,
    })

@login_required
def transfer_history(request):
    from apps.wallets.models import BalanceTransfer
    sent = BalanceTransfer.objects.filter(sender=request.user).order_by('-created_at')
    received = BalanceTransfer.objects.filter(recipient=request.user).order_by('-created_at')
    return render(request, "site/v3/v3_transfer_history.html", {
        "sent_transfers": sent,
        "received_transfers": received
    })

@login_required
def deposits(request):
    if request.method == "POST":
        method_id = request.POST.get("payment_method")
        currency_id = request.POST.get("currency")
        amount_str = request.POST.get("amount", "0")
        proof_image = request.FILES.get("proof_image")
        
        # Validation
        if not method_id or not currency_id:
            messages.error(request, "بيانات وسيلة الدفع أو العملة ناقصة.")
            return redirect("dashboard_deposits")
            
        method = get_object_or_404(PaymentMethod, id=method_id, is_active=True, can_deposit=True)
        
        # KYC Check (Task 9)
        if method.requires_kyc and not request.user.is_kyc_verified:
            messages.error(request, "هذه الوسيلة تتطلب توثيق الحساب (KYC) أولاً.")
            return redirect("dashboard_deposits")
            
        currency = get_object_or_404(Currency, id=currency_id, is_active=True)
        
        # Security check: Does currency belong to method?
        if not method.supported_currencies.filter(id=currency.id).exists():
            messages.error(request, "العملة المختارة غير مدعومة لهذه الوسيلة.")
            return redirect("dashboard_deposits")

        try:
            amount = Decimal(amount_str)
            if amount <= 0: raise ValueError()
        except:
            messages.error(request, "يرجى إدخال مبلغ صحيح.")
            return redirect("dashboard_deposits")

        # --- Task 2 & 3: Limit Checks ---
        amount_in_usd = currency.to_base(amount, "deposit")
        
        # 1. User Daily Limit
        request.user.reset_daily_limits_if_needed()
        if request.user.daily_deposit_usage + amount_in_usd > request.user.daily_deposit_limit:
            messages.error(request, f"لقد تجاوزت حد الإيداع اليومي المسموح لك ({request.user.daily_deposit_limit} USD). المتبقي لك اليوم: {max(0, request.user.daily_deposit_limit - request.user.daily_deposit_usage):,.2f} USD")
            return redirect("dashboard_deposits")
            
        # 2. Method Daily Limit
        method.reset_daily_limits_if_needed()
        if method.daily_deposit_usage + amount_in_usd > method.daily_deposit_limit:
            messages.error(request, "عذراً، هذه الوسيلة وصلت للحد الأقصى للإيداعات اليومية. يرجى المحاولة غداً أو استخدام وسيلة أخرى.")
            return redirect("dashboard_deposits")
            
        # 3. Global Cap (Task 5)
        if method.global_deposit_cap > 0 and method.global_deposit_usage + amount_in_usd > method.global_deposit_cap:
            messages.error(request, "عذراً، هذه الوسيلة غير متوفرة حالياً لتجاوزها الحد الإجمالي المسموح به.")
            return redirect("dashboard_deposits")

        # Extract metadata from custom fields
        metadata = {}
        schema = method.deposit_form_schema
        customer_note = request.POST.get("customer_note") or request.POST.get("note") or request.POST.get("ملاحظة") or ""
        
        for field in schema.get("fields", []) if isinstance(schema, dict) else []:
            field_name = field.get("name") or field.get("id") or field.get("key") or field.get("label")
            
            if field.get("type") == "image":
                val_file = request.FILES.get(f"custom_{field_name}")
                if val_file:
                    from django.core.files.storage import default_storage
                    path = default_storage.save(f"deposit-proofs/metadata/{val_file.name}", val_file)
                    val = f"{settings.MEDIA_URL}{path}"
                else:
                    val = ""
            else:
                val = request.POST.get(f"custom_{field_name}")

            if field.get("required") and not val:
                messages.error(request, f"الحقل {field.get('label')} مطلوب.")
                return redirect("dashboard_deposits")
            metadata[field_name] = val
            
            field_label = field.get("label", "").lower()
            field_name_lower = field_name.lower()
            is_note_field = False
            for note_keyword in ("customer_note", "note", "notes", "ملاحظة", "ملاحظات", "ملاحظة العميل", "حقل ملاحظة", "الرسالة", "message"):
                if note_keyword in field_label or note_keyword in field_name_lower:
                    is_note_field = True
                    break
            if is_note_field and val:
                customer_note = val

        # Create the request (unverified first)
        with transaction.atomic():
            deposit = DepositRequest.objects.create(
                user=request.user,
                payment_method=method,
                currency=currency,
                amount=amount,
                proof_image=proof_image,
                metadata=metadata,
                customer_note=customer_note,
                status=DepositRequest.Status.PENDING,
                is_verified=False
            )
            # Task 2 & 3: Increment Usage
            request.user.add_deposit_usage(amount_in_usd)
            method.add_deposit_usage(amount_in_usd)

            # Log pending activity safely
            try:
                ActivityLog.objects.create(user=request.user, action="deposit_requested", description=f"Requested {amount} {currency.code} via {method.name}")
            except:
                pass

        # AFTER creation: Check if verification is needed
        if v3_init_verification(request, request.user, "deposit"):
            # If no security method enabled, auto-verify
            deposit.is_verified = True
            deposit.save(update_fields=["is_verified"])
            
            # Notify user about the request submission ONLY if verified
            send_financial_notification(
                user=request.user,
                title="تم استلام طلب الإيداع",
                body=f"تم استلام طلب الإيداع الخاص بك رقم {deposit.id} بقيمة {deposit.amount} {deposit.currency.code}. سيتم مراجعته من قبل الإدارة قريباً."
            )
            from apps.notifications.services import notify_staff
            notify_staff(
                title="طلب إيداع جديد",
                body=f"قام {request.user.email} بتقديم طلب إيداع بقيمة {deposit.amount} {deposit.currency.code}",
                action_url=f"/control/deposits/{deposit.id}/",
                category='admin_new_deposit'
            )
            messages.success(request, "تم تقديم طلب الإيداع بنجاح.")

            return redirect("dashboard_deposits")
        else:
            # Save request ID in session for verification callback
            request.session["v3_pending_action_id"] = str(deposit.id)
            messages.info(request, "يرجى التحقق لإكمال الطلب.")
            methods = request.session.get("v3_auth_methods", [])
            return v3_redirect_to_verification(request, methods)

    # Task 9: Show all methods, KYC check handled in template/POST
    methods = PaymentMethod.objects.filter(is_active=True, can_deposit=True)

    return render(request, "site/v3/v3_deposits.html", {
        "payment_methods": methods, 
        "requests": DepositRequest.objects.filter(user=request.user, is_verified=True).order_by('-created_at'),
        "daily_limit": request.user.daily_deposit_limit,
        "remaining_limit": request.user.remaining_deposit_limit,
        "kyc_request": KYCRequest.objects.filter(user=request.user).order_by('-created_at').first(),
    })

@login_required
def withdrawals(request):
    # (Keep existing withdrawals logic, I will insert the new views below it)
    pass # Replaced temporarily for structure matching if needed, wait, I shouldn't replace `withdrawals` body. I'll just append it after `withdrawals`.

    if request.method == "POST":
        method_id = request.POST.get("payment_method")
        currency_id = request.POST.get("currency")
        amount_str = request.POST.get("amount", "0")
        
        if not method_id or not currency_id:
            messages.error(request, "بيانات ناقصة.")
            return redirect("dashboard_withdrawals")

        method = get_object_or_404(PaymentMethod, id=method_id, is_active=True, can_withdraw=True)
        
        # Task 9: KYC Check
        if method.requires_kyc and not request.user.is_kyc_verified:
            messages.error(request, "هذه الوسيلة تتطلب توثيق الحساب (KYC) أولاً.")
            return redirect("dashboard_withdrawals")

        currency = get_object_or_404(Currency, id=currency_id, is_active=True)
        
        if not method.supported_currencies.filter(id=currency.id).exists():
            messages.error(request, "العملة المختارة غير مدعومة.")
            return redirect("dashboard_withdrawals")

        try:
            amount = Decimal(amount_str)
            if amount <= 0: raise ValueError()
        except:
            messages.error(request, "مبلغ غير صحيح.")
            return redirect("dashboard_withdrawals")

        # Limit checks (pre-request)
        amount_in_usd = currency.to_base(amount, "withdraw")

        # 1. Method-specific transaction limits
        if amount_in_usd < method.withdrawal_min_amount:
            messages.error(request, f"الحد الأدنى للسحب عبر هذه الوسيلة هو {method.withdrawal_min_amount:,.2f} USD")
            return redirect("dashboard_withdrawals")
        
        if amount_in_usd > method.withdrawal_max_amount:
            messages.error(request, f"الحد الأقصى للسحب في العملية الواحدة عبر هذه الوسيلة هو {method.withdrawal_max_amount:,.2f} USD")
            return redirect("dashboard_withdrawals")

        # --- Task 2 & 3: Daily Limits ---
        # A) User Daily Limit
        request.user.reset_daily_limits_if_needed()
        if request.user.daily_withdrawal_usage + amount_in_usd > request.user.daily_withdrawal_limit:
            messages.error(request, f"لقد تجاوزت حد السحب اليومي المسموح لك ({request.user.daily_withdrawal_limit} USD). المتبقي لك اليوم: {max(0, request.user.daily_withdrawal_limit - request.user.daily_withdrawal_usage):,.2f} USD")
            return redirect("dashboard_withdrawals")

        # B) Method Daily Limit
        method.reset_daily_limits_if_needed()
        if method.daily_withdrawal_usage + amount_in_usd > method.daily_withdrawal_limit:
            messages.error(request, "عذراً، هذه الوسيلة وصلت للحد الأقصى للسحوبات اليومية. يرجى استخدام وسيلة أخرى.")
            return redirect("dashboard_withdrawals")

        # C) Priority/Custom Limits (Check if user has a smaller custom limit for this method)
        if request.user.has_custom_limits:
            user_custom = request.user.custom_payment_limits.get(str(method.id)) or request.user.custom_payment_limits.get(method.id.hex)
            if user_custom and user_custom.get('withdraw'):
                try:
                    custom_limit = Decimal(str(user_custom['withdraw']))
                    if amount_in_usd > custom_limit:
                         messages.error(request, f"عذراً، حد السحب المخصص لك لهذه الوسيلة هو {custom_limit:,.2f} USD")
                         return redirect("dashboard_withdrawals")
                except: pass

        # Create request (unverified)
        with transaction.atomic():
            # Wallet check happens in freeze_funds
            wallet_amount = currency.to_base(amount, "withdraw")
            if request.user.wallet.currency.code != "USD":
                wallet_amount = request.user.wallet.currency.from_base(wallet_amount, "withdraw")

            withdrawal = WithdrawalRequest.objects.create(
                user=request.user,
                payment_method=method,
                currency=currency,
                amount=amount,
                wallet_amount=wallet_amount,
                status=WithdrawalRequest.Status.PENDING,
                is_verified=False
            )
            # Increment Usage
            request.user.add_withdrawal_usage(amount_in_usd)
            method.add_withdrawal_usage(amount_in_usd)

        # Extract payout details
        payout_details = {"dynamic": {}}
        schema = method.withdrawal_form_schema
        for field in schema.get("fields", []) if isinstance(schema, dict) else []:
            field_name = field.get("name") or field.get("id") or field.get("key") or field.get("label")
            
            if field.get("type") == "image":
                val_file = request.FILES.get(f"custom_{field_name}")
                if val_file:
                    # Save file manually to media/withdrawal-proofs/customer/
                    from django.core.files.storage import default_storage
                    path = default_storage.save(f"withdrawal-proofs/customer/{val_file.name}", val_file)
                    val = f"{settings.MEDIA_URL}{path}"
                else:
                    val = ""
            else:
                val = request.POST.get(f"custom_{field_name}")
            
            if field.get("required") and not val:
                messages.error(request, f"الحقل {field.get('label')} مطلوب.")
                return redirect("dashboard_withdrawals")
            payout_details["dynamic"][field_name] = val

        # Create unverified request
        try:
            with transaction.atomic():
                # Correctly convert the withdrawal amount to the wallet's currency (usually USD)
                # wallet_amount is the actual amount to be frozen from the balance
                wallet_amount = currency.to_base(amount, "withdraw")
                
                # Wallet check happens in freeze_funds
                freeze_funds(
                    request.user.wallet.id, 
                    wallet_amount, 
                    reference=f"with_req_{timezone.now().timestamp()}", 
                    reason=f"سحب {amount} {currency.code} عبر {method.name}"
                )
                
                withdrawal = WithdrawalRequest.objects.create(
                    user=request.user,
                    payment_method=method,
                    currency=currency,
                    amount=amount,
                    wallet_amount=wallet_amount, # Important: store the base amount
                    payout_details=payout_details,
                    status=WithdrawalRequest.Status.PENDING,
                    is_verified=False
                )
            
            # Notify user about the request submission
            send_financial_notification(
                user=request.user,
                title="تم استلام طلب السحب",
                body=f"تم استلام طلب السحب الخاص بك رقم {withdrawal.id} بقيمة {withdrawal.amount} {withdrawal.currency.code}. سيتم مراجعته من قبل الإدارة قريباً."
            )
            from apps.notifications.services import notify_staff
            notify_staff(
                title="طلب سحب جديد",
                body=f"قام {request.user.email} بتقديم طلب سحب بقيمة {withdrawal.amount} {withdrawal.currency.code}",
                action_url=f"/control/withdrawals/{withdrawal.id}/",
                category='admin_new_withdrawal'
            )

        except Exception as e:
            messages.error(request, str(e))
            return redirect("dashboard_withdrawals")

        # AFTER creation: Verification
        if v3_init_verification(request, request.user, "withdraw"):
            withdrawal.is_verified = True
            withdrawal.save(update_fields=["is_verified"])
            
            # Notify user about the request submission ONLY if verified
            send_financial_notification(
                user=request.user,
                title="تم استلام طلب السحب",
                body=f"تم استلام طلب السحب الخاص بك رقم {withdrawal.id} بقيمة {withdrawal.amount} {withdrawal.currency.code}. سيتم مراجعته من قبل الإدارة قريباً."
            )
            messages.success(request, "تم تقديم طلب السحب بنجاح.")
            return redirect("dashboard_withdrawals")
        else:
            request.session["v3_pending_action_id"] = str(withdrawal.id)
            messages.info(request, "يرجى التحقق لإكمال طلب السحب.")
            methods = request.session.get("v3_auth_methods", [])
            return v3_redirect_to_verification(request, methods)

    # Task 9: Show all methods, KYC check handled in template/POST
    methods = PaymentMethod.objects.filter(is_active=True, can_withdraw=True)

    return render(request, "site/v3/v3_withdrawals.html", {
        "payment_methods": methods, 
        "requests": WithdrawalRequest.objects.filter(user=request.user, is_verified=True).order_by('-created_at'),
        "daily_limit": request.user.daily_withdrawal_limit,
        "remaining_limit": request.user.remaining_withdrawal_limit,
        "kyc_request": KYCRequest.objects.filter(user=request.user).order_by('-created_at').first(),
    })

@login_required
def kyc_request_view(request):
    existing = KYCRequest.objects.filter(user=request.user).first()
    if existing and existing.status in [KYCRequest.Status.PENDING, KYCRequest.Status.APPROVED]: 
        return render(request, "site/v3/v3_kyc_status.html", {"kyc": existing})

    form = KYCRequestForm(request.POST or None, request.FILES or None, instance=existing, user=request.user)
    if request.method == "POST":
        if form.is_valid():
            kyc = form.save(commit=False)
            kyc.user, kyc.status = request.user, KYCRequest.Status.PENDING
            
            # Save phone if provided in form
            if 'phone' in form.cleaned_data:
                request.user.phone = form.cleaned_data['phone']
                request.user.save(update_fields=['phone'])
                
            # Save password if provided in form (social users)
            if 'password' in form.cleaned_data:
                request.user.set_password(form.cleaned_data['password'])
                request.user.save()
                update_session_auth_hash(request, request.user) # Keep session active after password change

            kyc.save()
            from apps.notifications.services import notify_staff
            notify_staff(
                title="طلب توثيق جديد",
                body=f"مستخدم: {request.user.email}",
                action_url=f"/control/kyc/{kyc.id}/",
                category='admin_new_kyc'
            )
            
            # Send Email Notification to User
            from apps.accounts.services import send_kyc_status_email
            send_kyc_status_email(request.user, 'pending')
            
            messages.success(request, "تم تقديم الطلب."); return redirect("dashboard")
        else:
            messages.error(request, "يرجى تصحيح الأخطاء في النموذج أدناه.")
    
    return render(request, "site/v3/v3_kyc_form.html", {"form": form})

@login_required
def notifications_list(request):
    notifications = Notification.objects.filter(user=request.user).order_by("-created_at")
    Notification.objects.filter(user=request.user, is_read=False).update(is_read=True, read_at=timezone.now())
    return render(request, "site/notifications_list.html", {"notifications": notifications})

@login_required
def notification_settings(request):
    obj, _ = NotificationSetting.objects.get_or_create(user=request.user)
    form = NotificationSettingForm(request.POST or None, instance=obj, is_staff=request.user.is_platform_staff)
    if request.method == "POST" and form.is_valid():
        form.save()
        messages.success(request, "تم حفظ إعدادات الإشعارات بنجاح.")
        return redirect("notification_settings")
    return render(request, "site/v3/v3_notification_settings.html", {"form": form})

@login_required
def v3_change_password_view(request):
    if not v3_init_verification(request, request.user, "settings"):
        return v3_security_redirect(request.session.get("v3_auth_methods", []))

    has_password = request.user.has_usable_password()
    form = ChangePasswordForm(request.POST or None, has_password=has_password)
    if request.method == "POST" and form.is_valid():
        success = False
        if not has_password:
            request.user.set_password(form.cleaned_data["new_password"])
            request.user.save()
            success = True
        elif request.user.check_password(form.cleaned_data["current_password"]):
            request.user.set_password(form.cleaned_data["new_password"])
            request.user.save()
            success = True
        else:
            messages.error(request, "كلمة المرور الحالية غير صحيحة.")

        if success:
            update_session_auth_hash(request, request.user)
            messages.success(request, "تم تعيين/تغيير كلمة المرور بنجاح.")
            return redirect("dashboard")

    return render(request, "site/v3/v3_change_password.html", {
        "form": form,
        "has_password": has_password
    })
@login_required
def v3_change_email_view(request):
    if not v3_init_verification(request, request.user, "settings"):
        return v3_security_redirect(request.session.get("v3_auth_methods", []))
        
    if request.method == "POST":
        new_email = request.POST.get("new_email", "").lower().strip()
        if User.objects.filter(email=new_email).exists(): messages.error(request, "البريد الإلكتروني مستخدم بالفعل.")
        else:
            otp = v3_generate_otp(request.user, "email_change")
            if v3_send_otp_email(request.user, otp):
                request.session["v3_auth_uid"], request.session["v3_auth_purpose"] = str(request.user.id), "email_change"; request.session["v3_new_email"] = new_email
                return redirect("site_verify_otp")
            messages.error(request, "فشل إرسال الرمز.")
    return render(request, "site/v3/v3_change_email.html")

# ==========================================
# --- CATALOG & AJAX ---
# ==========================================

def home(request):
    # Multi-Tenant: TenantManager automatically filters by request.store.
    # When is_tenant=True (store context), all querysets return only that store's data.
    # When is_tenant=False (main site), querysets return main platform data (store=None).
    store = getattr(request, 'store', None)

    # Categories with product count — filtered by TenantManager automatically
    categories = Category.objects.filter(is_active=True).annotate(
        product_count=Count("products", filter=Q(products__is_active=True))
    ).filter(product_count__gt=0).order_by("sort_order", "name")

    featured_products = Product.objects.filter(
        is_active=True, is_featured=True
    ).select_related("category").prefetch_related("variants")[:12]

    sale_products = Product.objects.filter(
        Q(is_sale=True) | Q(variants__is_sale=True),
        is_active=True,
    ).select_related("category").prefetch_related("variants").distinct()[:6]

    ctx = {
        "featured_products": featured_products,
        "sale_products": sale_products,
        "categories": categories,
    }

    if not store:
        # Main Raqamiyat platform: show platform stats and testimonials
        from apps.common.models import PlatformStatistic, Testimonial

        base_stats = {
            "orders": Order.objects.count(),
            "users": User.objects.count(),
            "tickets": ChatRoom.objects.count(),
            "deposits": DepositRequest.objects.filter(status=DepositRequest.Status.COMPLETED).count(),
            "withdrawals": WithdrawalRequest.objects.filter(status=WithdrawalRequest.Status.COMPLETED).count(),
            "products": Product.objects.filter(is_active=True).count()
        }

        custom_stats_qs = PlatformStatistic.objects.filter(is_active=True).order_by('display_order')
        display_stats = []
        for stat in custom_stats_qs:
            val = stat.value_override
            if stat.stat_type == PlatformStatistic.StatType.USERS:
                val += base_stats['users']
            elif stat.stat_type == PlatformStatistic.StatType.ORDERS:
                val += base_stats['orders']
            elif stat.stat_type == PlatformStatistic.StatType.DEPOSITS:
                val += base_stats['deposits']
            elif stat.stat_type == PlatformStatistic.StatType.WITHDRAWALS:
                val += base_stats['withdrawals']
            elif stat.stat_type == PlatformStatistic.StatType.PRODUCTS:
                val += base_stats['products']

            display_stats.append({
                'label': stat.label,
                'value': stat.string_value or f"{val}{stat.value_suffix}",
                'icon_class': stat.icon_class or 'fas fa-star',
                'stat_type': stat.stat_type
            })

        testimonials = Testimonial.objects.filter(is_approved=True).order_by('-created_at')[:6]
        ctx["display_stats"] = display_stats
        ctx["testimonials"] = testimonials

        from apps.stores.models import Store
        platform_stores = Store.objects.filter(is_active=True).order_by('-is_featured', 'display_order', 'name')
        ctx["platform_stores"] = platform_stores

    return render(request, "site/home.html", ctx)

def catalog(request):
    # Multi-Tenant: TenantManager automatically filters by request.store context.
    # No explicit store filtering needed — TenantManager handles it.
    view_type = request.GET.get("view", "products")  # products or categories
    cat_id = request.GET.get("category")
    q = request.GET.get("q", "").strip()
    sort = request.GET.get("sort", "newest")
    cols = request.GET.get("cols", "2")  # Default 2 columns for mobile

    categories = Category.objects.filter(is_active=True).annotate(
        product_count=Count('products', filter=Q(products__is_active=True))
    ).order_by("sort_order", "name")

    products = Product.objects.filter(is_active=True).select_related("category").prefetch_related("variants")

    if cat_id:
        products = products.filter(category_id=cat_id)
        view_type = "products"  # Force product view if category selected

    if q:
        products = products.filter(Q(name__icontains=q) | Q(description__icontains=q))
        view_type = "products"  # Force product view if searching

    if sort == "price_low":
        products = products.order_by("variants__price")
    elif sort == "price_high":
        products = products.order_by("-variants__price")
    else:
        products = products.order_by("sort_order", "name")

    # Calculate total products count for the "All" category tab
    total_products_qs = Product.objects.filter(is_active=True)
    if q:
        total_products_qs = total_products_qs.filter(Q(name__icontains=q) | Q(description__icontains=q))
    total_products_count = total_products_qs.distinct().count()

    # Product Suggestion Form
    suggestion_form = ProductSuggestionForm()

    # Pagination for catalog
    from django.core.paginator import Paginator
    paginator = Paginator(products.distinct(), 24) # 24 products per page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    ctx = {
        "categories": categories,
        "page_obj": page_obj,
        "total_products_count": total_products_count,
        "active_category": cat_id,
        "query": q,
        "sort": sort,
        "view_type": view_type,
        "cols": cols,
        "suggestion_form": suggestion_form,
    }
    return render(request, "site/catalog.html", ctx)

@login_required
def site_submit_testimonial(request):
    form = TestimonialForm(request.POST or None)
    if request.method == "POST" and form.is_valid():
        t = form.save(commit=False)
        t.user = request.user
        t.save()
        messages.success(request, "شكراً لرأيك! سيظهر تعليقك في الموقع بعد مراجعته من الإدارة.")
        return redirect("dashboard")
    return render(request, "site/v3/v3_submit_testimonial.html", {"form": form})

@support_required
def control_testimonials_list(request):
    from apps.common.models import Testimonial
    status_filter = request.GET.get('approved')
    items = Testimonial.objects.all().select_related('user').order_by('-created_at')
    if request.store:
        items = items.filter(user__store=request.store)
    if status_filter:
        items = items.filter(is_approved=(status_filter == '1'))
    return render(request, "site/control_testimonials_list.html", {"testimonials": items, "status_filter": status_filter})

@support_required
def control_testimonial_moderate(request, pk):
    from apps.common.models import Testimonial
    if request.store:
        t = get_object_or_404(Testimonial, pk=pk, user__store=request.store)
    else:
        t = get_object_or_404(Testimonial, pk=pk)
    action = request.POST.get("action")
    
    if action == "approve":
        t.is_approved = True
        t.admin_reply = request.POST.get("admin_reply", t.admin_reply)
        t.save()
        messages.success(request, "تمت الموافقة على التعليق وحفظ الرد.")
    elif action == "unapprove":
        t.is_approved = False
        t.save()
        messages.info(request, "تم إخفاء التعليق.")
    elif action == "delete":
        t.delete()
        messages.warning(request, "تم حذف التعليق.")
    elif action == "update_reply":
        t.admin_reply = request.POST.get("admin_reply")
        t.save()
        messages.success(request, "تم تحديث رد الإدارة.")
        
    return redirect("control_testimonials_list")

@admin_required
def control_stats_list(request):
    from apps.common.models import PlatformStatistic
    return render(request, "site/control_stats_list.html", {"stats": PlatformStatistic.objects.all()})

@admin_required
def control_stat_create(request):
    form = PlatformStatisticForm(request.POST or None)
    if request.method == "POST" and form.is_valid():
        form.save()
        return redirect("control_stats_list")
    return render(request, "site/control_stat_form.html", {"form": form})

@admin_required
def control_stat_edit(request, pk):
    from apps.common.models import PlatformStatistic
    s = get_object_or_404(PlatformStatistic, pk=pk)
    form = PlatformStatisticForm(request.POST or None, instance=s)
    if request.method == "POST" and form.is_valid():
        form.save()
        return redirect("control_stats_list")
    return render(request, "site/control_stat_form.html", {"form": form, "stat": s})

@admin_required
def control_stat_delete(request, pk):
    from apps.common.models import PlatformStatistic
    get_object_or_404(PlatformStatistic, pk=pk).delete()
    return redirect("control_stats_list")

def v3_check_sp_grace_period(request):
    """
    Checks if the user has verified their security password or performed any 
    security verification in the last 5 minutes.
    """
    # Check both specific SP verification and global action verification
    keys = ["v3_sp_verified_at", "v3_action_verified_at"]
    for key in keys:
        last_verified = request.session.get(key)
        if last_verified:
            try:
                from django.utils.dateparse import parse_datetime
                dt = parse_datetime(last_verified)
                if not dt: # Try ISO format if parse_datetime fails
                    from datetime import datetime
                    dt = datetime.fromisoformat(last_verified)
                
                if dt and (timezone.now() - dt).total_seconds() < 300: # 5 minutes
                    return True
            except: pass
    return False

@login_required
def v3_security_triggers_view(request):
    if not v3_init_verification(request, request.user, "settings"):
        return v3_security_redirect(request.session.get("v3_auth_methods", []))
        
    user = request.user
    sp_verified = v3_check_sp_grace_period(request)

    if request.method == "POST":
        action = request.POST.get("action")
        
        # 1. Unlock Settings (Verify SP and start grace period)
        if action == "unlock_settings":
            sp = request.POST.get("security_password")
            from django.contrib.auth.hashers import check_password
            if not user.security_password or not check_password(sp, user.security_password):
                messages.error(request, "كلمة مرور الحماية غير صحيحة.")
            else:
                request.session["v3_sp_verified_at"] = timezone.now().isoformat()
                messages.success(request, "تم فتح قفل الإعدادات لمدة 5 دقائق.")
            return redirect("site_security_triggers")

        # 2. Update Security Password Logic
        if action == "update_security_password":
            current_sp = request.POST.get("current_security_password")
            new_sp = request.POST.get("new_security_password")
            confirm_sp = request.POST.get("confirm_security_password")
            from django.contrib.auth.hashers import make_password, check_password
            
            # If already has one, verify it
            if user.security_password:
                if not check_password(current_sp, user.security_password):
                    messages.error(request, "كلمة مرور الحماية الحالية غير صحيحة.")
                    return redirect("site_security_triggers")
            
            if new_sp and new_sp == confirm_sp:
                if len(new_sp) < 4:
                    messages.error(request, "كلمة المرور يجب أن تكون 4 خانات على الأقل.")
                else:
                    user.security_password = make_password(new_sp)
                    user.security_password_enabled = True
                    user.save()
                    request.session["v3_sp_verified_at"] = timezone.now().isoformat()
                    messages.success(request, "تم تعيين كلمة مرور الحماية بنجاح.")
            else:
                messages.error(request, "كلمات المرور غير متطابقة.")
            return redirect("site_security_triggers")

        # 3. Toggle Protection
        if action == "toggle_protection":
            sp = request.POST.get("security_password")
            from django.contrib.auth.hashers import check_password
            if not user.security_password or not check_password(sp, user.security_password):
                messages.error(request, "كلمة مرور الحماية غير صحيحة.")
            else:
                user.security_password_enabled = not user.security_password_enabled
                user.save()
                messages.success(request, f"تم {'تفعيل' if user.security_password_enabled else 'إيقاف'} حماية الإعدادات.")
            return redirect("site_security_triggers")

        # 4. Update Triggers (Requires SP or Grace Period if enabled)
        if user.security_password_enabled and not sp_verified:
            sp = request.POST.get("security_password_verify")
            from django.contrib.auth.hashers import check_password
            if not sp or not check_password(sp, user.security_password):
                messages.error(request, "يرجى إدخال كلمة مرور الحماية لتغيير هذه الإعدادات.")
                return redirect("site_security_triggers")
            request.session["v3_sp_verified_at"] = timezone.now().isoformat()

        user.security_login_method = request.POST.get("login_method")
        user.security_deposit_method = request.POST.get("deposit_method")
        user.security_purchase_method = request.POST.get("purchase_method")
        user.security_withdraw_method = request.POST.get("withdraw_method")
        user.save(); messages.success(request, "تم تحديث إعدادات الأمان بنجاح."); return redirect("site_security_triggers")
        
    return render(request, "site/v3/v3_security_triggers.html", {"sp_verified": sp_verified})

def product_detail(request, pk):
    product = get_object_or_404(Product.objects.prefetch_related('variants'), pk=pk, is_active=True)
    if request.method == "POST":
        if not request.user.is_authenticated: return redirect("site_login")
        
        variant_id = request.POST.get("variant_id")
        variant = get_object_or_404(ProductVariant, id=variant_id, product=product)
        
        # Collect custom fields
        metadata = {}
        for key in request.POST:
            if key.startswith("custom_"):
                metadata[key.replace("custom_", "")] = request.POST.get(key)
        
        coupon_code = request.POST.get("coupon_code")
        price = variant.get_price_for_user(request.user)
        
        # Check Coupon for balance check
        discount_amount = Decimal("0.00")
        coupon = None
        if coupon_code:
            from apps.orders.services import validate_coupon
            coupon = Coupon.objects.filter(code__iexact=coupon_code).first()
            if coupon:
                try:
                    discount_amount = validate_coupon(coupon, request.user, variant, subtotal=price)
                    price -= discount_amount
                except ValueError as e:
                    messages.error(request, f"خطأ في الكوبون: {str(e)}")
                    coupon = None

        # Check balance
        if request.user.wallet.available_balance < price:
            missing_amount = price - request.user.wallet.available_balance
            wallet = request.user.wallet
            display_missing = missing_amount
            if wallet.currency.code != "USD":
                display_missing = wallet.currency.from_base(missing_amount)
            
            currency_symbol = wallet.currency.symbol
            messages.error(request, f"رصيد غير كافٍ. تحتاج إلى {display_missing:,.2f} {currency_symbol} إضافية لإتمام الطلب.")
            request.session['missing_amount'] = str(display_missing)
            request.session['missing_currency'] = wallet.currency.code
            return redirect("product_detail", pk=pk)

        # Physical product shipping validation
        shipping_name = ""
        shipping_phone = ""
        shipping_address = ""
        if product.product_type == "physical" and not product.form_schema.get("fields"):
            shipping_name = request.POST.get("shipping_name", "").strip()
            shipping_phone = request.POST.get("shipping_phone", "").strip()
            shipping_address = request.POST.get("shipping_address", "").strip()
            if not (shipping_name and shipping_phone and shipping_address):
                messages.error(request, "جميع حقول الشحن والتوصيل مطلوبة للطلب المادي.")
                return redirect("product_detail", pk=pk)

        # Verification check
        if not v3_init_verification(request, request.user, "purchase"):
            # Save pending purchase details in session
            request.session["v3_pending_purchase"] = {
                "variant_id": str(variant.id),
                "coupon_code": coupon_code,
                "metadata": metadata,
                "shipping_name": shipping_name,
                "shipping_phone": shipping_phone,
                "shipping_address": shipping_address,
            }
            last_verified = request.session.get("v3_action_verified_at")
            if not last_verified or (timezone.now() - timezone.datetime.fromisoformat(last_verified)).total_seconds() > 300:
                methods = request.session.get("v3_auth_methods", [])
                return redirect("site_2fa_verify" if methods[0] == "APP" else "site_verify_otp")

        # Purchase Logic using create_order service
        from apps.orders.services import create_order
        try:
            order = create_order(
                customer=request.user,
                variant_id=variant.id,
                quantity=1,
                coupon=coupon,
                metadata=metadata,
                shipping_name=shipping_name,
                shipping_phone=shipping_phone,
                shipping_address=shipping_address,
            )
            messages.success(request, "تم إتمام الطلب بنجاح.")
            if variant.delivery_type == 'keys':
                return redirect(reverse("dashboard_order_detail", kwargs={"pk": order.id}))
            else:
                return redirect("dashboard_orders")
        except ValueError as e:
            messages.error(request, str(e))
            return redirect("product_detail", pk=pk)

    variants = product.variants.filter(is_active=True).order_by('sort_order')
    related_products = Product.objects.filter(category=product.category, is_active=True).exclude(pk=product.pk)[:3]
    
    missing_amount = request.session.pop('missing_amount', None)
    missing_currency = request.session.pop('missing_currency', None)
    
    return render(request, "site/product_detail.html", {
        "product": product, 
        "variants": variants, 
        "related_products": related_products,
        "missing_amount": missing_amount,
        "missing_currency": missing_currency
    })

def ajax_validate_coupon(request):
    try:
        from apps.orders.services import validate_coupon
        variant_id = request.GET.get("variant_id")
        code = request.GET.get("code", "").strip()
        
        if not variant_id or not code:
            return JsonResponse({"valid": False, "error": "بيانات ناقصة"})
            
        variant = ProductVariant.objects.select_related("product").get(id=variant_id)
        coupon = Coupon.objects.filter(code__iexact=code).first()
        
        if not coupon:
            return JsonResponse({"valid": False, "error": "الكوبون غير صحيح"})
            
        price = variant.get_price_for_user(request.user)
        try:
            discount = validate_coupon(coupon, request.user, variant, subtotal=price)
            new_total = price - discount
            return JsonResponse({
                "valid": True,
                "discount_amount": float(discount),
                "new_total": float(new_total),
                "message": f"تم تطبيق الكوبون بنجاح: خصم بقيمة {discount} USD"
            })
        except ValueError as e:
            return JsonResponse({"valid": False, "error": str(e)})

    except Exception as e:
        return JsonResponse({"valid": False, "error": "حدث خطأ أثناء التحقق من الكوبون"})

# ==========================================
# --- ADMINISTRATIVE VIEWS (V4) ---
# ==========================================

@staff_required
def control_dashboard(request):
    from apps.payments.models import DepositRequest, WithdrawalRequest
    store = getattr(request, "store", None)
    
    if store:
        users_qs = User.objects.filter(store=store)
    else:
        users_qs = User.objects.filter(store__isnull=True)

    stats = {
        "users": users_qs.count(),
        "pending_deposits": DepositRequest.objects.filter(status=DepositRequest.Status.PENDING, is_verified=True).count(),
        "pending_withdrawals": WithdrawalRequest.objects.filter(status=WithdrawalRequest.Status.PENDING, is_verified=True).count(),
        "open_tickets": ChatRoom.objects.exclude(status=ChatRoom.Status.CLOSED).count()
    }
    categories_summary = Category.objects.filter(is_active=True).annotate(product_count=Count('products')).order_by('sort_order')[:5]

    return render(request, "site/control_dashboard.html", {
        "stats": stats,
        "categories_summary": categories_summary,
        "recent_orders": Order.objects.select_related('customer').order_by('-created_at')[:5],
        "recent_deposits": DepositRequest.objects.filter(status=DepositRequest.Status.PENDING, is_verified=True).select_related('user', 'payment_method').order_by('-created_at')[:5],
        "recent_users": users_qs.order_by('-date_joined')[:5]
    })

@finance_required
def control_deposits(request):
    deposits = DepositRequest.objects.filter(is_verified=True).select_related('user', 'payment_method', 'currency').order_by('-created_at')
    
    q = request.GET.get("q")
    if q:
        q = q.strip()
        deposits = deposits.filter(
            Q(transaction_id__icontains=q) |
            Q(user__email__icontains=q) |
            Q(user__first_name__icontains=q) |
            Q(user__last_name__icontains=q)
        )
        
    current_status = request.GET.get("status")
    if current_status:
        current_status = current_status.strip()
        deposits = deposits.filter(status=current_status)
        
    latest_deposit = deposits.first()
    
    return render(request, "site/control_deposits.html", {
        "deposits": deposits,
        "latest_deposit": latest_deposit,
        "status_choices": DepositRequest.Status.choices,
        "current_status": current_status,
        "query": q
    })

@finance_required
def control_withdrawals(request):
    import uuid
    withdrawals = WithdrawalRequest.objects.filter(is_verified=True).select_related('user', 'payment_method', 'currency').order_by('-created_at')
    
    q = request.GET.get("q")
    if q:
        q = q.strip()
        is_uuid = False
        try:
            uuid.UUID(q)
            is_uuid = True
        except ValueError:
            pass
            
        if is_uuid:
            withdrawals = withdrawals.filter(id=q)
        else:
            withdrawals = withdrawals.filter(
                Q(user__email__icontains=q) |
                Q(user__first_name__icontains=q) |
                Q(user__last_name__icontains=q)
            )
            
    current_status = request.GET.get("status")
    if current_status:
        current_status = current_status.strip()
        withdrawals = withdrawals.filter(status=current_status)
        
    return render(request, "site/control_withdrawals.html", {
        "withdrawals": withdrawals,
        "status_choices": WithdrawalRequest.Status.choices,
        "current_status": current_status,
        "query": q
    })

@admin_required
def control_transfers(request):
    from apps.wallets.models import BalanceTransfer
    qs = BalanceTransfer.objects.all().select_related("sender", "recipient", "currency").order_by("-created_at")
    if request.store:
        qs = qs.filter(sender__store=request.store)
    
    q = request.GET.get("q")
    if q:
        qs = qs.filter(
            Q(reference__icontains=q) | 
            Q(sender__email__icontains=q) | 
            Q(recipient__email__icontains=q) |
            Q(sender__uid__iexact=q) |
            Q(recipient__uid__iexact=q)
        )
        
    return render(request, "site/control_transfers.html", {
        "page_obj": Paginator(qs, 50).get_page(request.GET.get("page")),
        "query": q
    })

@admin_required
def control_transfer_reverse(request, pk):
    from apps.wallets.models import BalanceTransfer
    from apps.wallets.services import reverse_p2p_transfer, WalletError
    
    if request.method == "POST":
        if request.store:
            transfer = get_object_or_404(BalanceTransfer, pk=pk, sender__store=request.store)
        else:
            transfer = get_object_or_404(BalanceTransfer, pk=pk)
        try:
            reverse_p2p_transfer(transfer, admin_user=request.user)
            messages.success(request, f"تم إلغاء واسترداد التحويل ({transfer.reference}) بنجاح.")
        except WalletError as e:
            messages.error(request, str(e))
        except Exception as e:
            messages.error(request, f"حدث خطأ أثناء الإلغاء: {str(e)}")
            
    return redirect("control_transfers")

@admin_required
def control_transfer_suspend(request, pk):
    from apps.wallets.models import BalanceTransfer
    from apps.wallets.services import suspend_p2p_transfer, WalletError
    
    if request.method == "POST":
        if request.store:
            transfer = get_object_or_404(BalanceTransfer, pk=pk, sender__store=request.store)
        else:
            transfer = get_object_or_404(BalanceTransfer, pk=pk)
        try:
            suspend_p2p_transfer(transfer, admin_user=request.user)
            messages.success(request, f"تم تعليق الحوالة ({transfer.reference}) بنجاح وحجز الرصيد من المستلم.")
        except WalletError as e:
            messages.error(request, str(e))
        except Exception as e:
            messages.error(request, f"حدث خطأ أثناء تعليق الحوالة: {str(e)}")
            
    return redirect("control_transfers")

@admin_required
def control_transfer_unsuspend(request, pk):
    from apps.wallets.models import BalanceTransfer
    from apps.wallets.services import unsuspend_p2p_transfer, WalletError
    
    if request.method == "POST":
        if request.store:
            transfer = get_object_or_404(BalanceTransfer, pk=pk, sender__store=request.store)
        else:
            transfer = get_object_or_404(BalanceTransfer, pk=pk)
        try:
            unsuspend_p2p_transfer(transfer, admin_user=request.user)
            messages.success(request, f"تم إلغاء تعليق الحوالة ({transfer.reference}) بنجاح وإرجاع الرصيد للمستلم.")
        except WalletError as e:
            messages.error(request, str(e))
        except Exception as e:
            messages.error(request, f"حدث خطأ أثناء إلغاء التعليق: {str(e)}")
            
    return redirect("control_transfers")

@admin_required
def control_transfer_edit_amount(request, pk):
    from apps.wallets.models import BalanceTransfer
    from apps.wallets.services import edit_p2p_transfer_amount, WalletError
    
    if request.method == "POST":
        if request.store:
            transfer = get_object_or_404(BalanceTransfer, pk=pk, sender__store=request.store)
        else:
            transfer = get_object_or_404(BalanceTransfer, pk=pk)
        new_amount_str = request.POST.get("new_amount")
        try:
            new_amount = Decimal(new_amount_str)
            edit_p2p_transfer_amount(transfer, new_amount, admin_user=request.user)
            messages.success(request, f"تم تعديل مبلغ الحوالة ({transfer.reference}) إلى {new_amount} {transfer.currency.code} بنجاح.")
        except Exception as e:
            if isinstance(e, WalletError):
                messages.error(request, str(e))
            else:
                messages.error(request, "يرجى إدخال مبلغ صحيح.")
            
    return redirect("control_transfers")

# ==========================================
# --- FINANCIAL NOTIFICATION HELPERS ---
# ==========================================

from urllib.parse import urljoin

def send_financial_notification(user, title, body, action_url="/dashboard/wallet/"):
    # Rely entirely on notify_user which now natively handles in-app, push, and email based on user preferences.
    try:
        notify_user(user=user, title=title, body=body, action_url=action_url, category='financial', priority="high")
    except Exception as e:
        pass

    # Check Email Preferences
    from apps.notifications.models import NotificationSetting
    settings_obj, _ = NotificationSetting.objects.get_or_create(user=user)
    if not settings_obj.email_financial:
        return

    # 2. Email Notification with Modern Design
    subject = f"{title} | Raqamiyat"
    # Ensure no double slashes in URL
    full_url = urljoin(settings.SITE_URL, action_url)

    html_content = f"""
    <!DOCTYPE html>
    <html dir="rtl" lang="ar">
    <head>
        <meta charset="UTF-8">
        <style>
            .container {{
                font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
                max-width: 600px;
                margin: 0 auto;
                padding: 40px 20px;
                background-color: #f8fafc;
            }}
            .card {{
                background-color: #ffffff;
                border-radius: 24px;
                overflow: hidden;
                box-shadow: 0 10px 25px rgba(0,0,0,0.05);
                border: 1px solid #e2e8f0;
            }}
            .header {{
                background: linear-gradient(135deg, #0891b2 0%, #06b6d4 100%);
                padding: 40px 20px;
                text-align: center;
                color: #ffffff;
            }}
            .logo {{
                font-size: 28px;
                font-weight: 900;
                letter-spacing: 1px;
                margin: 0;
                text-shadow: 0 2px 4px rgba(0,0,0,0.1);
            }}
            .content {{
                padding: 40px;
            }}
            .title {{
                color: #0f172a;
                font-size: 22px;
                font-weight: 800;
                margin-top: 0;
                margin-bottom: 20px;
                text-align: center;
            }}
            .message {{
                font-size: 16px;
                line-height: 1.8;
                color: #334155;
                margin-bottom: 30px;
                text-align: right;
            }}
            .cta-container {{
                text-align: center;
                margin-bottom: 20px;
            }}
            .cta-button {{
                display: inline-block;
                padding: 14px 32px;
                background-color: #0891b2;
                color: #ffffff !important;
                text-decoration: none;
                border-radius: 12px;
                font-weight: 700;
                font-size: 16px;
                transition: background-color 0.2s;
            }}
            .security-notice {{
                margin-top: 40px;
                padding: 20px;
                background-color: #fff1f2;
                border-radius: 16px;
                border: 1px solid #fecdd3;
            }}
            .security-title {{
                margin: 0 0 8px 0;
                font-size: 14px;
                color: #be123c;
                font-weight: 800;
                display: flex;
                align-items: center;
            }}
            .security-text {{
                margin: 0;
                font-size: 13px;
                color: #e11d48;
                line-height: 1.6;
            }}
            .footer {{
                margin-top: 30px;
                text-align: center;
                color: #94a3b8;
                font-size: 12px;
            }}
        </style>
    </head>
    <body>
        <div class="container">
            <div class="card">
                <div class="header">
                    <h1 class="logo">رقميات | RAQAMIYAT</h1>
                </div>
                <div class="content">
                    <h2 class="title">{title}</h2>
                    <div class="message">{body}</div>

                    <div class="cta-container">
                        <a href="{full_url}" class="cta-button">عرض في المنصة</a>
                    </div>

                    <div class="security-notice">
                        <p class="security-title">⚠️ تنبيه أمني:</p>
                        <p class="security-text">
                            إذا لم تكن قد قمت بهذه العملية بنفسك، يرجى تغيير كلمة مرور حسابك فوراً والتواصل مع فريق الدعم الفني لحماية حسابك.
                        </p>
                    </div>
                </div>
            </div>
            <div class="footer">
                <p>© 2026 مؤسسة رامي قصاص بن ماهر لخدمات الوساطة الرقمية. جميع الحقوق محفوظة.</p>
                <p>تم إرسال هذا البريد تلقائياً، يرجى عدم الرد عليه.</p>
            </div>
        </div>
    </body>
    </html>
    """
    try:
        from apps.accounts.services import send_brevo_email
        send_brevo_email(to_email=user.email, to_name=user.get_full_name() or user.email, subject=subject, html_content=html_content)
    except: pass

@finance_required
def control_deposit_detail(request, pk):
    from apps.payments.models import DepositRequest
    deposit = get_object_or_404(DepositRequest.objects.select_related('user', 'currency', 'payment_method'), pk=pk)

    if request.method == "POST":
        action = request.POST.get("action")
        admin_note = request.POST.get("admin_note", "")

        try:
            with transaction.atomic():
                if action == "approve":
                    if deposit.status == DepositRequest.Status.COMPLETED:
                        raise ValueError("تم اعتماد هذا الطلب مسبقاً.")

                    original_requested_amount = deposit.amount
                    override_amount = request.POST.get("amount")
                    if override_amount:
                        # Admin is specifying the FINAL amount to credit
                        final_amount = Decimal(str(override_amount))
                        deposit.final_amount = final_amount
                        deposit.amount = final_amount # Treat as gross if they want it exact
                        deposit.fee_amount = Decimal("0.00")
                        deposit._fees_calculated = True
                    else:
                        final_amount = deposit.final_amount

                    wallet = get_or_create_wallet(deposit.user)
                    if deposit.currency.code == wallet.currency.code:
                        wallet_final_amount = final_amount
                    else:
                        base_val = deposit.currency.to_base(final_amount, "deposit")
                        wallet_final_amount = wallet.currency.from_base(base_val, "deposit")

                    if wallet_final_amount <= 0:
                        raise ValueError("يجب أن يكون المبلغ أكبر من صفر.")

                    credit_wallet(
                        wallet_id=wallet.id,
                        amount=wallet_final_amount,
                        reference=f"deposit:{deposit.id}",
                        description=f"إيداع عبر {deposit.payment_method.name}",
                        created_by=request.user,
                        source="deposit",
                        metadata={
                            "from_pending": True,
                            "pending_amount": str(deposit.wallet_amount),
                            "source_amount": str(deposit.amount),
                            "source_currency": deposit.currency.code,
                            "final_amount": str(final_amount),
                            "transaction_id": deposit.transaction_id
                        }
                    )

                    # Ensure record reflects what was actually credited
                    deposit.wallet_amount = wallet_final_amount
                    deposit.status = DepositRequest.Status.COMPLETED
                    deposit.reviewed_by = request.user
                    deposit.reviewed_at = timezone.now()
                    if admin_note:
                        deposit.admin_note = admin_note
                    deposit.save()

                    # Update Global Payment Method Cap (Task 5) - Still needed as it's a lifetime cap
                    try:
                        amount_in_usd = deposit.currency.to_base(deposit.final_amount, "deposit")
                        method = deposit.payment_method
                        method.global_deposit_usage += amount_in_usd
                        if method.global_deposit_cap > 0 and method.global_deposit_usage >= method.global_deposit_cap:
                            method.is_maintenance_mode = True
                            # Notify Admin
                            from apps.notifications.services import notify_staff
                            notify_staff(
                                title="تجاوز الحد الأقصى لطريقة الدفع",
                                body=f"وصلت طريقة الدفع {method.name} للحد الأقصى المسموح به ({method.global_deposit_cap:,.2f} USD). تم تحويلها لوضع الصيانة تلقائياً.",
                                category="system"
                            )
                        method.save()
                    except: pass
                    # Transparent body showing original vs approved if changed
                    amount_text = f"{deposit.final_amount:,.2f} {deposit.currency.code}"
                    if override_amount and Decimal(str(override_amount)) != original_requested_amount:
                        body_msg = f"تمت مراجعة طلب الإيداع رقم {deposit.id}. تم اعتماد مبلغ {amount_text} بدلاً من المبلغ المطلوب {original_requested_amount:,.2f} {deposit.currency.code}. تم إضافة {deposit.wallet_amount:,.2f} {wallet.currency.code} إلى رصيد محفظتك."
                    else:
                        body_msg = f"تم قبول طلب الإيداع رقم {deposit.id} بنجاح. المبلغ المعتمد: {amount_text}. تم إضافة {deposit.wallet_amount:,.2f} {wallet.currency.code} إلى رصيد محفظتك."

                    send_financial_notification(
                        user=deposit.user,
                        title="تم قبول طلب الإيداع",
                        body=body_msg
                    )
                    messages.success(request, "تم قبول طلب الإيداع بنجاح.")

                elif action == "reject":
                    if deposit.status == DepositRequest.Status.COMPLETED:
                        raise ValueError("لا يمكن رفض طلب مكتمل.")
                    
                    if deposit.status != DepositRequest.Status.REJECTED:
                        wallet = get_object_or_404(Wallet, user=deposit.user)
                        from apps.wallets.services import cancel_pending_deposit
                        cancel_pending_deposit(
                            wallet_id=wallet.id,
                            amount=deposit.wallet_amount,
                            reference=f"deposit_reject:{deposit.id}",
                            description=f"إلغاء إيداع معلق مرفوض عبر {deposit.payment_method.name}",
                            created_by=request.user
                        )
                        # Task 2 & 3: Reverse Usage
                        try:
                            amt_usd = deposit.currency.to_base(deposit.amount, "deposit")
                            deposit.user.add_deposit_usage(-amt_usd)
                            deposit.payment_method.add_deposit_usage(-amt_usd)
                        except: pass

                    deposit.status = DepositRequest.Status.REJECTED
                    deposit.admin_note = admin_note
                    deposit.reviewed_by = request.user
                    deposit.reviewed_at = timezone.now()
                    deposit.save()
                    
                    send_financial_notification(
                        user=deposit.user,
                        title="تم رفض طلب الإيداع",
                        body=f"نعتذر، تم رفض طلب الإيداع رقم {deposit.id}. السبب: {admin_note}"
                    )
                    messages.warning(request, "تم رفض طلب الإيداع.")

                elif action == "correct":
                    if deposit.status != DepositRequest.Status.COMPLETED:
                        raise ValueError("يمكن تصحيح الطلبات المكتملة فقط.")

                    new_amount_str = request.POST.get("new_amount")
                    if not new_amount_str:
                        raise ValueError("المبلغ الجديد مطلوب.")
                    
                    new_amount = Decimal(str(new_amount_str))
                    old_amount = deposit.final_amount
                    diff_amount = new_amount - old_amount
                    
                    if diff_amount == 0:
                        raise ValueError("لم يتم تغيير المبلغ.")

                    wallet = get_or_create_wallet(deposit.user)
                    if deposit.currency.code == wallet.currency.code:
                        wallet_diff = diff_amount
                    else:
                        base_diff = deposit.currency.to_base(diff_amount, "deposit")
                        wallet_diff = wallet.currency.from_base(base_diff, "deposit")

                    if wallet_diff > 0:
                        credit_wallet(
                            wallet_id=wallet.id,
                            amount=wallet_diff,
                            reference=f"deposit_adj:{deposit.id}",
                            description=f"تصحيح مبلغ الإيداع (زيادة): {admin_note}",
                            created_by=request.user,
                            source="admin_adjustment",
                            metadata={
                                "source_amount": str(diff_amount),
                                "source_currency": deposit.currency.code,
                                "transaction_id": deposit.transaction_id,
                                "is_adjustment": True
                            }
                        )
                    else:
                        debit_wallet(
                            wallet_id=wallet.id,
                            amount=abs(wallet_diff),
                            reference=f"deposit_adj:{deposit.id}",
                            description=f"تصحيح مبلغ الإيداع (نقص): {admin_note}",
                            created_by=request.user,
                            source="admin_adjustment",
                            metadata={
                                "source_amount": str(diff_amount),
                                "source_currency": deposit.currency.code
                            }
                        )

                    deposit.final_amount = new_amount
                    deposit.wallet_amount += wallet_diff
                    if admin_note:
                        deposit.admin_note = f"{deposit.admin_note or ''}\n[تصحيح {timezone.now().strftime('%Y-%m-%d %H:%M')}]: {admin_note}"
                    deposit.save()
                    
                    send_financial_notification(
                        user=deposit.user,
                        title="تعديل في مبلغ إيداع سابق",
                        body=f"تم تعديل المبلغ المعتمد للإيداع رقم {deposit.id}. القيمة الجديدة: {deposit.final_amount:,.2f} {deposit.currency.code}. تم تعديل رصيد محفظتك بفرق: {wallet_diff:,.2f} {wallet.currency.code}."
                    )
                    
                    messages.success(request, f"تم تصحيح المبلغ بنجاح. الفرق: {wallet_diff:,.2f} {wallet.currency.code}")
            return redirect("control_deposits")
        except Exception as e:
            messages.error(request, f"خطأ: {str(e)}")
            return redirect("control_deposit_detail", pk=pk)
        
    return render(request, "site/control_deposit_detail.html", {"deposit": deposit})

@finance_required
def control_withdrawal_detail(request, pk):
    withdrawal = get_object_or_404(WithdrawalRequest.objects.select_related('user', 'payment_method'), pk=pk)
    if request.method == "POST":
        action = request.POST.get("action")
        admin_note = request.POST.get("admin_note", "")
        proof_image = request.FILES.get("proof_image")
        proof_file = request.FILES.get("proof_file")
        
        try:
            with transaction.atomic():
                if action == "process":
                    withdrawal.status = WithdrawalRequest.Status.PROCESSING
                elif action == "approve":
                    withdrawal.status = WithdrawalRequest.Status.APPROVED
                elif action == "complete":
                    if withdrawal.status != WithdrawalRequest.Status.COMPLETED:
                        finalize_withdrawal(
                            wallet_id=withdrawal.user.wallet.id,
                            amount=withdrawal.wallet_amount,
                            reference=f"with:{withdrawal.id}",
                            description=f"Withdrawal completed via {withdrawal.payment_method.name}",
                            created_by=request.user,
                            metadata={
                                "transaction_id": withdrawal.transaction_id
                            }
                        )
                        withdrawal.status = WithdrawalRequest.Status.COMPLETED
                        withdrawal.reviewed_at = timezone.now()
                elif action == "reject":
                    if withdrawal.status in [WithdrawalRequest.Status.COMPLETED, WithdrawalRequest.Status.CANCELLED, WithdrawalRequest.Status.REJECTED]:
                        raise ValueError("لا يمكن رفض طلب منتهي.")
                    
                    if withdrawal.status != WithdrawalRequest.Status.REJECTED:
                        release_funds(
                            wallet_id=withdrawal.user.wallet.id,
                            amount=withdrawal.wallet_amount,
                            reference=f"with:{withdrawal.id}",
                            description="Withdrawal rejected",
                            created_by=request.user,
                            metadata={
                                "transaction_id": withdrawal.transaction_id
                            }
                        )
                        
                        # Reverse daily usage on rejection
                        try:
                            amount_in_usd = withdrawal.currency.to_base(withdrawal.amount, "withdraw")
                            withdrawal.user.add_withdrawal_usage(-amount_in_usd)
                            withdrawal.payment_method.add_withdrawal_usage(-amount_in_usd)
                        except: pass

                        withdrawal.status = WithdrawalRequest.Status.REJECTED
                        withdrawal.reviewed_at = timezone.now()

                elif action == "correct":
                    if withdrawal.status != WithdrawalRequest.Status.COMPLETED:
                        raise ValueError("يمكن تصحيح الطلبات المكتملة فقط.")

                    new_amount_str = request.POST.get("new_amount")
                    if not new_amount_str:
                        raise ValueError("المبلغ الجديد مطلوب.")

                    new_amount = Decimal(str(new_amount_str))
                    old_amount = withdrawal.amount
                    diff_amount = new_amount - old_amount

                    if diff_amount == 0:
                        raise ValueError("لم يتم تغيير المبلغ.")

                    wallet = get_or_create_wallet(withdrawal.user)
                    # Calculate wallet difference (how much MORE or LESS to debit)
                    wallet_diff = withdrawal.currency.to_base(diff_amount, "withdraw")

                    if wallet_diff > 0:
                        # User withdrew MORE than originally thought -> debit MORE
                        debit_wallet(
                            wallet_id=wallet.id,
                            amount=wallet_diff,
                            reference=f"with_adj:{withdrawal.id}",
                            description=f"تصحيح مبلغ السحب (زيادة): {admin_note}",
                            created_by=request.user,
                            source="admin_adjustment"
                        )
                    else:
                        # User withdrew LESS than originally thought -> credit BACK
                        credit_wallet(
                            wallet_id=wallet.id,
                            amount=abs(wallet_diff),
                            reference=f"with_adj:{withdrawal.id}",
                            description=f"تصحيح مبلغ السحب (نقص): {admin_note}",
                            created_by=request.user,
                            source="admin_adjustment"
                        )

                    withdrawal.amount = new_amount
                    withdrawal.wallet_amount += wallet_diff
                    if admin_note:
                        withdrawal.admin_note = f"{withdrawal.admin_note or ''}\n[تصحيح {timezone.now().strftime('%Y-%m-%d %H:%M')}]: {admin_note}"
                    withdrawal.save()

                    send_financial_notification(
                        user=withdrawal.user,
                        title="تعديل في مبلغ سحب مكتمل",
                        body=f"تم تعديل المبلغ المرسل للسحب رقم {withdrawal.id}. القيمة النهائية: {withdrawal.amount:,.2f} {withdrawal.currency.code}. تم تعديل رصيد محفظتك بفرق: {-wallet_diff:,.2f} {wallet.currency.code}."
                    )
                    messages.success(request, f"تم تصحيح السحب بنجاح. الفرق: {-wallet_diff:,.2f} {wallet.currency.code}")
                
                if admin_note: withdrawal.admin_note = admin_note
                if proof_image: withdrawal.proof_image = proof_image
                if proof_file: withdrawal.proof_file = proof_file
                withdrawal.reviewed_by = request.user
                withdrawal.save()
                messages.success(request, f"تم تحديث حالة الطلب إلى {withdrawal.get_status_display()}")
                
                # Send Notification
                send_financial_notification(
                    user=withdrawal.user,
                    title=f"تحديث طلب السحب #{withdrawal.id}",
                    body=f"تم تغيير حالة طلب السحب الخاص بك إلى: {withdrawal.get_status_display()}. ملاحظة الإدارة: {withdrawal.admin_note or 'لا يوجد'}",
                )
            return redirect("control_withdrawals")
        except Exception as e:
            messages.error(request, f"خطأ: {str(e)}")
            return redirect("control_withdrawal_detail", pk=pk)
        
    return render(request, "site/control_withdrawal_detail.html", {"withdrawal": withdrawal})

@support_required
def control_kycs_list(request):
    store = getattr(request, "store", None)
    q = request.GET.get('q', '')
    status = request.GET.get('status', '') # This will refer to user.is_kyc_verified or kyc_request.status
    
    if store:
        users = User.objects.filter(role=User.Role.CUSTOMER, store=store).select_related('kyc_request').all().order_by('-date_joined')
    else:
        users = User.objects.filter(role=User.Role.CUSTOMER, store__isnull=True).select_related('kyc_request').all().order_by('-date_joined')
    
    if q:
        users = users.filter(
            Q(email__icontains=q) | 
            Q(first_name__icontains=q) | 
            Q(last_name__icontains=q) |
            Q(kyc_request__id_number__icontains=q)
        )
    
    if status == "verified":
        users = users.filter(is_kyc_verified=True)
    elif status == "pending":
        users = users.filter(kyc_request__status=KYCRequest.Status.PENDING)
    elif status == "unverified":
        users = users.filter(is_kyc_verified=False).exclude(kyc_request__status=KYCRequest.Status.PENDING)
    elif status == "rejected":
        users = users.filter(kyc_request__status=KYCRequest.Status.REJECTED)
        
    return render(request, "site/control_kycs_list.html", {
        "users": users, 
        "query": q, 
        "status_filter": status,
    })

@support_required
def control_kyc_detail(request, pk):
    store = getattr(request, "store", None)
    if store:
        kyc = get_object_or_404(KYCRequest.objects.select_related('user'), pk=pk, user__store=store)
    else:
        kyc = get_object_or_404(KYCRequest.objects.select_related('user'), pk=pk, user__store__isnull=True)
    # Pass is_admin=True to allow optional images during updates
    form = KYCRequestForm(request.POST or None, request.FILES or None, instance=kyc, is_admin=True)
    payment_methods = PaymentMethod.objects.filter(is_active=True).order_by("display_order")

    if request.method == "POST":
        action = request.POST.get("action")

        # 1. Update Personal Info (Including Images)
        if action == "update_info" and form.is_valid():
            form.save()
            messages.success(request, "تم تحديث بيانات التوثيق والملفات المرفقة بنجاح.")
            return redirect("control_kyc_detail", pk=pk)

        # 2. Update Limits
        if action == "update_limits":
            u = kyc.user
            u.daily_deposit_limit = Decimal(request.POST.get("global_deposit_limit", u.daily_deposit_limit))
            u.daily_withdrawal_limit = Decimal(request.POST.get("global_withdrawal_limit", u.daily_withdrawal_limit))

            # Custom per-method limits
            limits = u.custom_payment_limits or {}
            for method in payment_methods:
                m_id = str(method.id)
                dep = request.POST.get(f"method_dep_{m_id}")
                withd = request.POST.get(f"method_with_{m_id}")

                if dep or withd:
                    limits[m_id] = {
                        "deposit": dep if dep else None,
                        "withdraw": withd if withd else None
                    }
                elif m_id in limits:
                    del limits[m_id]

            u.custom_payment_limits = limits
            u.has_custom_limits = True if limits else False
            u.save()
            messages.success(request, "تم تحديث الحدود المالية بنجاح.")
            return redirect("control_kyc_detail", pk=pk)

        # 3. Final Decisions
        admin_note = request.POST.get("rejection_reason", "")
        if action == "approve":
            kyc.status = KYCRequest.Status.APPROVED
            kyc.user.is_kyc_verified = True
            
            # Update User Display Name: First Father Last (with spaces)
            kyc.user.first_name = f"{kyc.first_name} {kyc.father_name} {kyc.last_name}"
            kyc.user.last_name = "" # Clear last name to avoid duplication in some templates
            
            # Apply global limits if user doesn't have custom ones
            if not kyc.user.has_custom_limits:
                kyc_settings = KYCSettings.get_settings()
                kyc.user.daily_deposit_limit = kyc_settings.verified_daily_deposit_limit
                kyc.user.daily_withdrawal_limit = kyc_settings.verified_daily_withdrawal_limit
            
            kyc.user.save()
            
            # Send Email Notification
            from apps.accounts.services import send_kyc_status_email
            send_kyc_status_email(kyc.user, 'approved')
            notify_user(
                user=kyc.user,
                title="✅ تم توثيق حسابك",
                body="تهانينا، تمت الموافقة على طلب توثيق هويتك بنجاح. يمكنك الآن الاستمتاع بحدود مالية أعلى.",
                action_url="/dashboard/",
                category='kyc',
                priority=Notification.Priority.HIGH
            )

            messages.success(request, f"تم توثيق حساب {kyc.user.email} بنجاح، وتحديث الاسم المعروض والحدود المالية.")
        elif action == "reject":
            kyc.status = KYCRequest.Status.REJECTED
            kyc.rejection_reason = admin_note
            kyc.user.is_kyc_verified = False

            # Reset to unverified limits if no custom limits
            if not kyc.user.has_custom_limits:
                kyc_settings = KYCSettings.get_settings()
                kyc.user.daily_deposit_limit = kyc_settings.unverified_daily_deposit_limit
                kyc.user.daily_withdrawal_limit = kyc_settings.unverified_daily_withdrawal_limit

            kyc.user.save()

            # Send Email Notification
            from apps.accounts.services import send_kyc_status_email
            send_kyc_status_email(kyc.user, 'rejected', reason=admin_note)
            notify_user(
                user=kyc.user,
                title="❌ تحديث بشأن طلب التوثيق",
                body=f"نعتذر، تم رفض طلب توثيق هويتك. السبب: {admin_note}",
                action_url="/dashboard/verification/",
                category='kyc',
                priority=Notification.Priority.HIGH
            )

            messages.warning(request, f"تم رفض طلب توثيق {kyc.user.email}.")

        elif action == "unverify":
            kyc.status = KYCRequest.Status.REJECTED
            kyc.user.is_kyc_verified = False
            
            # Reset to unverified limits
            if not kyc.user.has_custom_limits:
                kyc_settings = KYCSettings.get_settings()
                kyc.user.daily_deposit_limit = kyc_settings.unverified_daily_deposit_limit
                kyc.user.daily_withdrawal_limit = kyc_settings.unverified_daily_withdrawal_limit
                
            kyc.user.save()
            messages.info(request, "تم إلغاء توثيق الحساب وإعادة الحدود للمستوى الأساسي.")
        elif action == "revert":
            kyc.status = KYCRequest.Status.PENDING
            kyc.save()
            messages.info(request, "تمت إعادة الطلب لحالة قيد المراجعة.")

        kyc.reviewed_by = request.user
        kyc.reviewed_at = timezone.now()
        kyc.save()
        return redirect("control_kycs_list")

    return render(request, "site/control_kyc_detail.html", {
        "kyc": kyc, 
        "form": form,
        "payment_methods": payment_methods
    })

@kyc_required
def control_kyc_settings(request):
    obj = KYCSettings.get_settings()
    form = KYCSettingsForm(request.POST or None, instance=obj)
    
    if request.method == "POST":
        action = request.POST.get("action")
        if action == "unblock_country":
            code = request.POST.get("country_code")
            if code and code in obj.restricted_countries:
                new_list = [c for c in obj.restricted_countries if c != code]
                obj.restricted_countries = new_list
                obj.save()
                messages.success(request, f"تم إلغاء حظر الدولة ({code}) بنجاح.")
            return redirect("control_kyc_settings")
        
        elif action == "block_country":
            code = request.POST.get("country_code")
            if code:
                current_list = list(obj.restricted_countries or [])
                if code not in current_list:
                    current_list.append(code)
                    obj.restricted_countries = current_list
                    obj.save()
                    messages.success(request, "تمت إضافة الدولة لقائمة الحظر.")
            return redirect("control_kyc_settings")
            
        if form.is_valid():
            form.save()
            messages.success(request, "تم حفظ الإعدادات بنجاح.")
            return redirect("control_kyc_settings")
            
    return render(request, "site/control_kyc_settings.html", {"form": form, "settings": obj})

@support_required
def control_orders_list(request):
    orders = Order.objects.select_related('customer').all().order_by('-created_at')
    if request.GET.get('status'): orders = orders.filter(status=request.GET.get('status'))
    if request.GET.get('q'): orders = orders.filter(Q(number__icontains=request.GET.get('q')) | Q(customer__email__icontains=request.GET.get('q')))
    return render(request, "site/control_orders_list.html", {"orders": orders, "order_status_choices": Order.Status.choices})

@support_required
def control_order_detail(request, pk):
    order = get_object_or_404(Order.objects.select_related('customer'), pk=pk)
    if request.method == "POST":
        action = request.POST.get("action")
        if action == "update_status":
            old_status = order.status
            order.status = request.POST.get("status")
            order.save()
            OrderLog.objects.create(order=order, status=order.status, note=request.POST.get("admin_note", ""), created_by=request.user)
            
            if order.status in [Order.Status.REFUNDED, Order.Status.CANCELLED] and old_status not in [Order.Status.REFUNDED, Order.Status.CANCELLED]: 
                credit_wallet(order.customer.wallet.id, order.total_amount, f"refund:{order.id}", f"استرداد مبلغ الطلب رقم #{order.number}", request.user)
            
            messages.success(request, f"تم تحديث حالة الطلب إلى: {order.get_status_display()}")
            
            # Notify user
            try:
                notify_user(
                    user=order.customer,
                    title="تحديث حالة الطلب",
                    body=f"تم تغيير حالة طلبك رقم #{order.number} إلى: {order.get_status_display()}",
                    action_url=f"/dashboard/orders/{order.id}/",
                    category="orders"
                )
            except: pass
        elif action == "update_fulfillment":
            keys = request.POST.getlist("ff_key[]")
            vals = request.POST.getlist("ff_value[]")
            fulfillment_data = {}
            for k, v in zip(keys, vals):
                if k.strip(): fulfillment_data[k.strip()] = v
            order.fulfillment_data = fulfillment_data
            order.save()
            messages.success(request, "تم تحديث بيانات التنفيذ.")
        elif action == "update_shipping":
            order.shipping_carrier = request.POST.get("shipping_carrier", "").strip()
            order.tracking_number = request.POST.get("tracking_number", "").strip()
            order.save()
            OrderLog.objects.create(
                order=order,
                status=order.status,
                note=f"تم تحديث معلومات الشحن (شركة الشحن: {order.shipping_carrier}، رقم التتبع: {order.tracking_number})",
                created_by=request.user
            )
            messages.success(request, "تم تحديث معلومات الشحن.")
            try:
                notify_user(
                    user=order.customer,
                    title="تحديث معلومات الشحن",
                    body=f"تمت إضافة معلومات الشحن لطلبك رقم #{order.number}. شركة الشحن: {order.shipping_carrier}، رقم التتبع: {order.tracking_number}",
                    action_url=f"/dashboard/orders/{order.id}/",
                    category="orders"
                )
            except: pass
        elif action == "update_price":
            val = request.POST.get("total_amount", "").strip()
            if not val:
                new_total = order.total_amount
            else:
                try:
                    new_total = Decimal(val)
                except:
                    messages.error(request, "قيمة السعر غير صالحة.")
                    return redirect("control_order_detail", pk=pk)

            old_total = order.total_amount
            diff = new_total - old_total
            
            if diff != 0:
                wallet = order.customer.wallet
                # Convert diff (USD) to wallet currency
                adj_amount = diff
                if wallet.currency.code != "USD":
                    adj_amount = wallet.currency.from_base(diff)
                
                try:
                    with transaction.atomic():
                        reason = request.POST.get("adjustment_reason", "")
                        if diff > 0:
                            # Price increased, debit user
                            desc = f"تعديل سعر الطلب #{order.number} (زيادة): من {old_total} إلى {new_total}"
                            if reason: desc += f" | السبب: {reason}"
                            debit_wallet(wallet.id, adj_amount, reference=f"order_adj:{order.id}", 
                                         description=desc, 
                                         created_by=request.user)
                        else:
                            # Price decreased, credit user
                            desc = f"تعديل سعر الطلب #{order.number} (تخفيض): من {old_total} إلى {new_total}"
                            if reason: desc += f" | السبب: {reason}"
                            credit_wallet(wallet.id, abs(adj_amount), reference=f"order_adj:{order.id}", 
                                          description=desc, 
                                          created_by=request.user)
                        
                        if not order.original_total:
                            order.original_total = old_total
                        
                        order.total_amount = new_total
                        order.price_adjustment_reason = reason
                        order.save()
                        
                        # Update order items
                        if order.items.count() == 1:
                            item = order.items.first()
                            item.total_price = new_total
                            if item.quantity == 1:
                                item.unit_price = new_total
                            item.save()

                        OrderLog.objects.create(
                            order=order, 
                            status=order.status, 
                            note=f"تم تعديل سعر الطلب بواسطة الإدارة من {old_total} إلى {new_total}. السبب: {order.price_adjustment_reason}", 
                            created_by=request.user
                        )

                        # Notify user about price adjustment
                        try:
                            notify_user(
                                user=order.customer,
                                title="تم تعديل سعر الطلب",
                                body=f"تم تعديل سعر طلبك رقم #{order.number} من قبل الإدارة. السعر الجديد: {new_total} USD. السبب: {order.price_adjustment_reason}",
                                action_url=f"/dashboard/orders/{order.id}/",
                                category="orders"
                            )
                        except: pass
                    messages.success(request, "تم تحديث سعر الطلب وتعديل رصيد المحفظة.")
                except Exception as e:
                    messages.error(request, f"فشل في تعديل الرصيد: {str(e)}")
            else:
                order.price_adjustment_reason = request.POST.get("adjustment_reason", "")
                order.save()
                messages.success(request, "تم تحديث ملاحظات تعديل السعر.")
            
        return redirect("control_order_detail", pk=pk)
    
    ctx = {
        "order": order,
        "mapped_metadata": order.formatted_metadata(),
        "readable_fulfillment": order.fulfillment_data or {}
    }
    return render(request, "site/control_order_detail.html", ctx)


@support_required
def control_order_status_update(request, pk):
    order = get_object_or_404(Order, pk=pk)
    new_status = request.POST.get("status")
    if new_status in Order.Status.values:
        old_status = order.status
        order.status = new_status
        order.save()
        OrderLog.objects.create(order=order, status=order.status, note=request.POST.get("admin_note", ""), created_by=request.user)
        if order.status in [Order.Status.REFUNDED, Order.Status.CANCELLED] and old_status not in [Order.Status.REFUNDED, Order.Status.CANCELLED]:
            credit_wallet(order.customer.wallet.id, order.total_amount, f"refund:{order.id}", f"استرداد مبلغ الطلب رقم #{order.number}", request.user)
        messages.success(request, f"تم تحديث حالة الطلب إلى: {order.get_status_display()}")
        try:
            notify_user(
                user=order.customer,
                title="تحديث حالة الطلب",
                body=f"تم تغيير حالة طلبك رقم #{order.number} إلى: {order.get_status_display()}",
                action_url=f"/dashboard/orders/{order.id}/",
                category="orders"
            )
        except: pass
    if getattr(request, "store", None):
        return redirect("merchant_order_detail", pk=order.pk)
    return redirect("control_order_detail", pk=order.pk)


def export_to_excel(queryset, filename, columns):
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill
        from io import BytesIO
        from django.http import HttpResponse

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Data"
        ws.sheet_view.rightToLeft = True

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="06b6d4", end_color="06b6d4", fill_type="solid")
        center_align = Alignment(horizontal="center", vertical="center")

        # Write Headers
        for col_idx, (header, _) in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align

        # Write Data
        for row_idx, obj in enumerate(queryset, 2):
            for col_idx, (_, getter) in enumerate(columns, 1):
                val = getter(obj)
                ws.cell(row=row_idx, column=col_idx, value=val)

        # Auto-adjust column widths
        from openpyxl.utils import get_column_letter
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except: pass
            ws.column_dimensions[column_letter].width = max_length + 2

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}_{timezone.now().strftime("%Y%m%d")}.xlsx"'
        return response
    except Exception as e:
        return HttpResponse(f"Error generating Excel: {str(e)}", status=500)

@admin_required
def control_users_list(request):
    store = getattr(request, "store", None)
    if request.method == "POST":
        action = request.POST.get("action")
        user_ids = request.POST.getlist("user_ids")
        if not user_ids:
            messages.warning(request, "يرجى اختيار مستخدمين لتنفيذ العملية.")
            return redirect("control_users_list")
        if action == "bulk_update":
            tier = request.POST.get("bulk_tier")
            if tier:
                if store:
                    User.objects.filter(id__in=user_ids, store=store).update(tier=tier)
                else:
                    User.objects.filter(id__in=user_ids, store__isnull=True).update(tier=tier)
                messages.success(request, f"تم تحديث فئة {len(user_ids)} مستخدم بنجاح.")
        return redirect("control_users_list")

    if store:
        users = User.objects.filter(store=store).order_by("-date_joined")
    else:
        users = User.objects.filter(store__isnull=True).order_by("-date_joined")
        
    q = request.GET.get('q', '').strip()
    status = request.GET.get('status', '')

    if q:
        users = users.filter(Q(email__icontains=q) | Q(first_name__icontains=q) | Q(last_name__icontains=q) | Q(phone__icontains=q))
    if status:
        users = users.filter(status=status)

    if request.GET.get("export") == "excel":
        columns = [
            ("الاسم", lambda u: u.get_full_name()),
            ("البريد الإلكتروني", lambda u: u.email),
            ("الهاتف", lambda u: u.phone),
            ("الفئة", lambda u: u.get_tier_display()),
            ("الحالة", lambda u: u.get_status_display()),
            ("الرصيد المتاح", lambda u: u.wallet.available_balance),
            ("الديون", lambda u: u.wallet.debt_balance),
            ("تاريخ الانضمام", lambda u: u.date_joined.strftime("%Y-%m-%d")),
        ]
        return export_to_excel(users, "Users", columns)

    return render(request, "site/control_users_list.html", {"users": users, "query": q, "current_status": status, "tiers": User.Tier.choices})

@support_required
def control_product_toggle_featured(request, pk):
    product = get_object_or_404(Product, pk=pk)
    product.is_featured = not product.is_featured
    product.save(update_fields=['is_featured'])
    return JsonResponse({"status": "success", "is_featured": product.is_featured})

@support_required
def control_product_reorder_ajax(request, pk):
    product = get_object_or_404(Product, pk=pk)
    direction = request.GET.get('direction')
    
    # Normalize sort orders
    all_products = list(Product.objects.all().order_by('sort_order', 'id'))
    for i, p in enumerate(all_products):
        if p.sort_order != i:
            p.sort_order = i
            p.save(update_fields=['sort_order'])
    
    product.refresh_from_db()
    
    if direction == 'up':
        other = Product.objects.filter(sort_order__lt=product.sort_order).order_by('-sort_order').first()
    elif direction == 'down':
        other = Product.objects.filter(sort_order__gt=product.sort_order).order_by('sort_order').first()
    else:
        return JsonResponse({"status": "error", "message": "Invalid direction"})
    
    if other:
        p_order = product.sort_order
        o_order = other.sort_order
        product.sort_order = o_order
        other.sort_order = p_order
        product.save(update_fields=['sort_order'])
        other.save(update_fields=['sort_order'])
        return JsonResponse({"status": "success", "new_order": product.sort_order})
    
    return JsonResponse({"status": "error", "message": "Cannot move further"})

@admin_required
def control_product_import(request):
    if request.method == "POST" and request.FILES.get("file"):
        import openpyxl
        try:
            excel_file = request.FILES["file"]
            wb = openpyxl.load_workbook(excel_file)
            ws = wb.active
            
            created, updated = 0, 0
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row[0]: continue # Skip empty ID rows
                
                try:
                    product_id = row[0]
                    name = row[1]
                    cat_name = row[2]
                    is_active = str(row[4]).strip() == "نعم"
                    is_featured = str(row[5]).strip() == "نعم"
                    
                    store = getattr(request, "store", None)
                    category, _ = Category.objects.get_or_create(
                        name=cat_name,
                        defaults={"store": store}
                    )
                    
                    p, created_now = Product.objects.update_or_create(
                        id=product_id,
                        defaults={
                            "name": name,
                            "category": category,
                            "is_active": is_active,
                            "is_featured": is_featured,
                            "store": store
                        }
                    )
                    if created_now: created += 1
                    else: updated += 1
                except Exception as row_err:
                    logger.warning(f"Error importing row {row}: {row_err}")
            
            messages.success(request, f"تم الاستيراد بنجاح: {created} جديد، {updated} تم تحديثه.")
        except Exception as e:
            messages.error(request, f"فشل الاستيراد: {str(e)}")
            
    return redirect("control_products_list")

@support_required
def control_categories_list(request):
    categories = Category.objects.all().order_by('sort_order', 'name')
    return render(request, "site/control_categories_list.html", {"categories": categories})

@support_required
def control_category_edit(request, pk=None):
    category = get_object_or_404(Category, pk=pk) if pk else None
    form = CategoryForm(request.POST or None, request.FILES or None, instance=category)
    if request.method == "POST" and form.is_valid():
        form.save()
        messages.success(request, "تم حفظ التصنيف بنجاح.")
        return redirect("control_categories_list")
    return render(request, "site/control_category_form.html", {"form": form, "category": category})

@support_required
def control_products_list(request):
    products = Product.objects.select_related('category').prefetch_related('variants').all().order_by('sort_order', 'name')
    cat_id = request.GET.get('category')
    active_category = None
    if cat_id:
        active_category = get_object_or_404(Category, id=cat_id)
        products = products.filter(category_id=cat_id)

    q = request.GET.get('q', '').strip()
    view_mode = request.GET.get('view', 'list')

    if q:
        products = products.filter(
            Q(name__icontains=q) |
            Q(category__name__icontains=q) |
            Q(id__icontains=q) |
            Q(variants__sku__icontains=q) |
            Q(variants__name__icontains=q)
        ).distinct()

    if request.GET.get("export") == "excel":
        columns = [
            ("ID", lambda p: str(p.id)),
            ("اسم المنتج", lambda p: p.name),
            ("التصنيف", lambda p: p.category.name),
            ("عدد الباقات", lambda p: p.variants.count()),
            ("نشط", lambda p: "نعم" if p.is_active else "لا"),
            ("مميز", lambda p: "نعم" if p.is_featured else "لا"),
            ("ترتيب العرض", lambda p: p.sort_order),
        ]
        return export_to_excel(products, "Products", columns)

    paginator = Paginator(products, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    categories = Category.objects.filter(is_active=True).order_by('sort_order', 'name')

    return render(request, "site/control_products_list.html", {
        "products": page_obj,
        "query": q,
        "view_mode": view_mode,
        "active_category": active_category,
        "categories": categories
    })

@support_required
def control_orders_list(request):
    orders = Order.objects.select_related('customer').prefetch_related('items__variant__product').all().order_by('-created_at')
    q = request.GET.get('q', '').strip()
    status = request.GET.get('status', '')

    if q: orders = orders.filter(Q(number__icontains=q) | Q(customer__email__icontains=q))
    if status: orders = orders.filter(status=status)

    if request.GET.get("export") == "excel":
        columns = [
            ("رقم الطلب", lambda o: o.number),
            ("التاريخ", lambda o: o.created_at.strftime("%Y-%m-%d %H:%M")),
            ("العميل", lambda o: o.customer.email),
            ("المنتج", lambda o: o.items.first().variant.product.name if o.items.exists() else ""),
            ("الباقة", lambda o: o.items.first().variant.name if o.items.exists() else ""),
            ("المبلغ", lambda o: o.total_amount),
            ("الحالة", lambda o: o.get_status_display()),
        ]
        return export_to_excel(orders, "Orders", columns)

    return render(request, "site/control_orders_list.html", {"orders": orders, "query": q, "current_status": status, "order_status_choices": Order.Status.choices})

@finance_required
def control_wallets_list(request):
    store = getattr(request, "store", None)
    if store:
        wallets = Wallet.objects.filter(user__store=store).select_related('user', 'currency').all().order_by('-updated_at')
    else:
        wallets = Wallet.objects.filter(user__store__isnull=True).select_related('user', 'currency').all().order_by('-updated_at')
        
    q = request.GET.get('q', '')
    if q: wallets = wallets.filter(Q(user__email__icontains=q) | Q(user__first_name__icontains=q))

    if request.GET.get("export") == "excel":
        columns = [
            ("المستخدم", lambda w: w.user.email),
            ("العملة", lambda w: w.currency.code),
            ("الرصيد المتاح", lambda w: w.available_balance),
            ("الديون", lambda w: w.debt_balance),
            ("المجمد", lambda w: w.frozen_balance),
            ("المحجوز", lambda w: w.held_balance),
            ("المعلق", lambda w: w.pending_balance),
        ]
        return export_to_excel(wallets, "Wallets", columns)

    return render(request, "site/control_wallets_list.html", {"wallets": wallets, "query": q})


@admin_required
def control_user_moderate(request, public_uuid):
    store = getattr(request, "store", None)
    if store:
        user = get_object_or_404(User, public_uuid=public_uuid, store=store)
    else:
        user = get_object_or_404(User, public_uuid=public_uuid, store__isnull=True)
    form = ModerateUserForm(request.POST or None, instance=user)
    if request.method == "POST":
        action = request.POST.get("action")
        if action == "change_email":
            new_email = request.POST.get("new_email", "").strip().lower()
            if new_email and "@" in new_email:
                if User.objects.filter(email=new_email).exclude(id=user.id).exists(): messages.error(request, "هذا البريد مستخدم بالفعل.")
                else: user.email = new_email; user.username = new_email; user.save(); messages.success(request, f"تم تغيير البريد إلى {new_email}")
            return redirect("control_user_moderate", public_uuid=public_uuid)
        elif action == "assign_debt":
            from apps.wallets.services import add_debt, WalletError
            amt_str = request.POST.get("amount", "0")
            reason = request.POST.get("reason", "")
            try:
                amt = Decimal(amt_str)
                add_debt(user.wallet.id, amt, f"admin_debt_{timezone.now().timestamp()}", reason, request.user)
                messages.success(request, f"تم إضافة دين بقيمة {amt} للمستخدم {user.email}")
            except WalletError as e:
                messages.error(request, str(e))
            except Exception as e:
                messages.error(request, f"خطأ أثناء إضافة الدين: {str(e)}")
            return redirect("control_user_moderate", public_uuid=public_uuid)
        elif action == "pay_debt":
            from apps.wallets.services import pay_debt, WalletError
            amt_str = request.POST.get("amount", "0")
            reason = request.POST.get("reason", "")
            try:
                amt = Decimal(amt_str)
                pay_debt(
                    user.wallet.id,
                    amt,
                    f"admin_pay_{timezone.now().timestamp()}",
                    reason or "سداد يدوي للمديونية من الرصيد المتاح",
                    request.user,
                    deduct_from_balance=True,
                    source="admin"
                )
                messages.success(request, f"تم سداد دين بقيمة {amt} للمستخدم {user.email}")
            except WalletError as e:
                messages.error(request, str(e))
            except Exception as e:
                messages.error(request, f"خطأ أثناء سداد الدين: {str(e)}")
            return redirect("control_user_moderate", public_uuid=public_uuid)
        elif action == "cash_deposit":
            from apps.wallets.services import credit_wallet, WalletError
            amt_str = request.POST.get("amount", "0")
            reason = request.POST.get("reason", "")
            try:
                amt = Decimal(amt_str)
                credit_wallet(
                    wallet_id=user.wallet.id,
                    amount=amt,
                    reference=f"cash_dep_{timezone.now().timestamp()}",
                    description=reason or "إيداع نقدي مباشر عبر لوحة التحكم",
                    created_by=request.user,
                    source="admin_cash",
                    reason=reason or "إيداع نقدي مباشر عبر لوحة التحكم"
                )
                messages.success(request, f"تم إيداع مبلغ {amt} نقداً في محفظة العميل {user.email}")
            except WalletError as e:
                messages.error(request, str(e))
            except Exception as e:
                messages.error(request, f"خطأ أثناء الإيداع النقدي: {str(e)}")
            return redirect("control_user_moderate", public_uuid=public_uuid)
        elif action == "reset_otp": user.otp_failed_attempts = 0; user.otp_lockout_until = None; user.otp_resend_count = 0; user.save(); messages.success(request, "تم إعادة ضبط قيود الرمز."); return redirect("control_user_moderate", public_uuid=public_uuid)
        elif action == "reset_2fa": user.totp_enabled = False; user.totp_secret = None; user.save(); messages.success(request, "تم تعطيل 2FA للمستخدم."); return redirect("control_user_moderate", public_uuid=public_uuid)
        elif action == "reset_sp": user.security_password = None; user.security_password_enabled = False; user.save(); messages.success(request, "تم حذف رمز الحماية (SP) للمستخدم بنجاح."); return redirect("control_user_moderate", public_uuid=public_uuid)
        elif action == "reset_password":
            from django.contrib.auth.tokens import default_token_generator
            from apps.accounts.services import send_brevo_email
            
            token = default_token_generator.make_token(user)
            uid = user.id
            reset_url = request.build_absolute_uri(reverse('site_reset_password')) + f"?token={token}&uid={uid}"
            
            # Professional Email Template
            html_content = f"""
            <div dir="rtl" style="font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif; max-width: 500px; margin: 0 auto; padding: 20px; color: #1e293b; background-color: #ffffff; border-radius: 16px; border: 1px solid #e2e8f0;">
                <div style="text-align: center; margin-bottom: 30px;">
                    <h2 style="color: #06b6d4; margin: 0; font-size: 24px; font-weight: 900;">رقميات | RAQAMIYAT</h2>
                </div>
                <div style="background-color: #f8fafc; padding: 30px; border-radius: 12px;">
                    <h3 style="margin-top: 0; color: #0f172a;">مرحباً {user.get_full_name() or user.username}،</h3>
                    <p style="font-size: 16px; color: #64748b;">لقد تلقينا طلباً لإعادة تعيين كلمة المرور لحسابك.</p>
                    <div style="text-align: center; margin: 30px 0;">
                        <a href="{reset_url}" style="padding: 14px 28px; background-color: #06b6d4; color: #ffffff; text-decoration: none; border-radius: 8px; font-weight: bold; font-size: 16px;">إعادة تعيين كلمة المرور</a>
                    </div>
                    <p style="font-size: 14px; color: #94a3b8;">إذا لم تطلب ذلك، يمكنك تجاهل هذا البريد الإلكتروني بأمان. الرابط صالح لفترة محدودة.</p>
                </div>
                <div style="margin-top: 30px; text-align: center; font-size: 12px; color: #94a3b8;">
                    <p>© 2026 مؤسسة رامي قصاص بن ماهر لخدمات الوساطة الرقمية.</p>
                </div>
            </div>
            """
            
            if send_brevo_email(user.email, user.get_full_name() or user.email, "إعادة تعيين كلمة المرور | رقميات", html_content):
                messages.success(request, "تم إرسال رابط إعادة تعيين كلمة المرور بنجاح.")
            else:
                messages.error(request, "فشل إرسال البريد الإلكتروني.")
            return redirect("control_user_moderate", public_uuid=public_uuid)
        elif action == "reset_limits":
            user.daily_deposit_usage = Decimal("0.00")
            user.daily_withdrawal_usage = Decimal("0.00")
            user.last_limit_reset = timezone.now()
            user.save(update_fields=["daily_deposit_usage", "daily_withdrawal_usage", "last_limit_reset"])
            messages.success(request, "تم تصفير جميع حدود الاستخدام اليومي للمستخدم.")
            return redirect("control_user_moderate", public_uuid=public_uuid)

        elif action == "reset_method_limit":
            method_id = request.POST.get("method_id")
            if user.has_custom_limits and method_id in user.custom_payment_limits:
                # Remove custom limit for this method
                user.custom_payment_limits.pop(method_id, None)
                user.save(update_fields=["custom_payment_limits"])
                messages.success(request, f"تم حذف الحد المخصص للوسيلة المحددة للمستخدم.")
            return redirect("control_user_moderate", public_uuid=public_uuid)

        elif form.is_valid():
            form.save()
            messages.success(request, "تم التحديث.")
            return redirect("control_users_list")
    from apps.payments.models import DepositRequest, WithdrawalRequest
    from apps.orders.models import Order
    from apps.accounts.models import ActivityLog

    recent_deposits = DepositRequest.objects.filter(user=user).select_related('payment_method', 'currency').order_by('-created_at')[:20]
    recent_withdrawals = WithdrawalRequest.objects.filter(user=user).select_related('payment_method', 'currency').order_by('-created_at')[:20]
    recent_orders = Order.objects.filter(customer=user).prefetch_related('items__variant__product').order_by('-created_at')[:20]
    recent_activities = ActivityLog.objects.filter(user=user).order_by('-created_at')[:50]

    return render(request, "site/control_user_moderate.html", {
        "form": form, 
        "user_to_moderate": user,
        "recent_deposits": recent_deposits,
        "recent_withdrawals": recent_withdrawals,
        "recent_orders": recent_orders,
        "recent_activities": recent_activities
    })

@admin_required
def currencies_list(request):
    if request.method == "POST":
        for c in Currency.objects.all():
            buy, sell = request.POST.get(f"buy_rate_{c.id}"), request.POST.get(f"sell_rate_{c.id}")
            if buy and sell: c.buy_rate, c.sell_rate = Decimal(buy), Decimal(sell); c.save()
        return redirect("currencies_list")
    return render(request, "site/currencies_list.html", {"currencies": Currency.objects.all().order_by('display_order')})

@admin_required
def currency_create(request):
    form = CurrencyForm(request.POST or None)
    if request.method == "POST" and form.is_valid(): form.save(); return redirect("currencies_list")
    return render(request, "site/currency_form.html", {"form": form})

@admin_required
def currency_edit(request, pk):
    c = get_object_or_404(Currency, pk=pk); form = CurrencyForm(request.POST or None, instance=c)
    if request.method == "POST" and form.is_valid(): form.save(); return redirect("currencies_list")
    return render(request, "site/currency_form.html", {"form": form, "currency": c})

@support_required
def control_categories_list(request):
    categories = Category.objects.all().order_by('sort_order', 'name')
    return render(request, "site/control_categories_list.html", {"categories": categories})

@support_required
def control_category_edit(request, pk=None):
    category = get_object_or_404(Category, pk=pk) if pk else None
    form = CategoryForm(request.POST or None, request.FILES or None, instance=category)
    if request.method == "POST" and form.is_valid():
        form.save()
        messages.success(request, "تم حفظ التصنيف بنجاح.")
        return redirect("control_categories_list")
    return render(request, "site/control_category_form.html", {"form": form, "category": category})

@support_required
def control_category_delete(request, pk):
    get_object_or_404(Category, pk=pk).delete()
    messages.success(request, "تم حذف التصنيف.")
    return redirect("control_categories_list")

@support_required
def control_categories_list(request):
    categories = Category.objects.all().order_by('sort_order', 'name')
    return render(request, "site/control_categories_list.html", {"categories": categories})

@support_required
def control_category_edit(request, pk=None):
    category = get_object_or_404(Category, pk=pk) if pk else None
    form = CategoryForm(request.POST or None, request.FILES or None, instance=category)
    if request.method == "POST" and form.is_valid():
        form.save()
        messages.success(request, "تم حفظ التصنيف بنجاح.")
        return redirect("control_categories_list")
    return render(request, "site/control_category_form.html", {"form": form, "category": category})

@support_required
def control_category_delete(request, pk):
    get_object_or_404(Category, pk=pk).delete()
    messages.success(request, "تم حذف التصنيف.")
    return redirect("control_categories_list")

@support_required
def control_category_create_ajax(request):
    name = request.POST.get("name")
    if name:
        store = getattr(request, "store", None)
        cat = Category.objects.create(name=name, store=store)
        return JsonResponse({"status": "success", "id": str(cat.id), "name": cat.name})
    return JsonResponse({"status": "error"}, status=400)

@support_required
@transaction.atomic
def control_product_create(request):
    form = ProductForm(request.POST or None, request.FILES or None)
    if request.method == "POST" and form.is_valid():
        product = form.save(commit=False)
        product.store = getattr(request, "store", None)
        product.save()
        form.save_m2m()

        # Handle gallery uploads
        gallery_files = request.FILES.getlist("gallery_files")
        for f in gallery_files:
            content_type = f.content_type or ""
            if content_type.startswith("video/"):
                ProductImage.objects.create(product=product, video=f)
            else:
                ProductImage.objects.create(product=product, image=f)

        v_json = request.POST.get("variants_json")
        if v_json:
            v_data = json.loads(v_json)
            def safe_decimal(val, default="0"):
                if not val or str(val).strip() == "": return Decimal(default)
                try: return Decimal(str(val))
                except: return Decimal(default)

            for v in v_data:
                ProductVariant.objects.create(
                    product=product,
                    name=v.get('name'),
                    sku=v.get('sku'),
                    price=safe_decimal(v.get('price')),
                    wholesale_price=safe_decimal(v.get('wholesale_price')),
                    vip_price=safe_decimal(v.get('vip_price')),
                    cost=safe_decimal(v.get('cost')),
                    estimated_delivery_minutes=int(v.get('estimated_delivery_minutes', 0) or 0),
                    sort_order=int(v.get('sort_order', 0) or 0),
                    is_active=v.get('is_active', True),
                    is_sale=v.get('is_sale', False),
                    delivery_type=v.get('delivery_type', 'manual')
                )

        redirect_target = request.POST.get("redirect_to_variant_keys")
        if redirect_target:
            return redirect("control_variant_keys", pk=redirect_target)
        return redirect("control_products_list")
    return render(request, "site/control_product_builder.html", {"form": form, "variants_json_data": [], "title": "إنشاء منتج جديد"})

@support_required
@transaction.atomic
def control_product_edit(request, pk):
    product = get_object_or_404(Product, pk=pk); form = ProductForm(request.POST or None, request.FILES or None, instance=product)
    if request.method == "POST" and form.is_valid():
        product = form.save(commit=False)
        product.store = getattr(request, "store", None)
        product.save()
        form.save_m2m()

        # Handle gallery uploads
        gallery_files = request.FILES.getlist("gallery_files")
        for f in gallery_files:
            content_type = f.content_type or ""
            if content_type.startswith("video/"):
                ProductImage.objects.create(product=product, video=f)
            else:
                ProductImage.objects.create(product=product, image=f)

        v_json = request.POST.get("variants_json")
        if v_json:
            v_data = json.loads(v_json); product.variants.exclude(sku__in=[v.get('sku') for v in v_data if v.get('sku')]).delete()
            for v in v_data:
                def safe_decimal(val, default="0"):
                    if not val or str(val).strip() == "": return Decimal(default)
                    try: return Decimal(str(val))
                    except: return Decimal(default)

                ProductVariant.objects.update_or_create(
                    product=product,
                    sku=v.get('sku'),
                    defaults={
                        "name": v.get('name'),
                        "price": safe_decimal(v.get('price')),
                        "wholesale_price": safe_decimal(v.get('wholesale_price')),
                        "vip_price": safe_decimal(v.get('vip_price')),
                        "cost": safe_decimal(v.get('cost')),
                        "estimated_delivery_minutes": int(v.get('estimated_delivery_minutes', 0) or 0),
                        "sort_order": int(v.get('sort_order', 0) or 0),
                        "is_active": v.get('is_active', True),
                        "is_sale": v.get('is_sale', False),
                        "delivery_type": v.get('delivery_type', 'manual')
                    }
                )

        redirect_target = request.POST.get("redirect_to_variant_keys")
        if redirect_target:
            return redirect("control_variant_keys", pk=redirect_target)
        return redirect("control_products_list")
    v_list = [
        {
            "id": str(v.id), "name": v.name, "sku": v.sku, "price": str(v.price), 
            "wholesale_price": str(v.wholesale_price), "vip_price": str(v.vip_price),
            "cost": str(v.cost), "estimated_delivery_minutes": v.estimated_delivery_minutes,
            "sort_order": v.sort_order, "is_active": v.is_active,
            "is_sale": v.is_sale,
            "is_temporarily_disabled": v.is_temporarily_disabled,
            "delivery_type": v.delivery_type,
            "keys_count": v.keys.filter(is_used=False).count()
        } for v in product.variants.all().order_by('sort_order')
    ]
    return render(
        request, 
        "site/control_product_builder.html", 
        {
            "form": form, 
            "product": product, 
            "variants_json_data": v_list, 
            "gallery": product.gallery.all().order_by('sort_order'),
            "title": f"تعديل: {product.name}"
        }
    )


@support_required
def control_product_delete(request, pk):
    product = get_object_or_404(Product, pk=pk)
    product.delete()
    messages.success(request, "تم حذف المنتج بنجاح.")
    if getattr(request, "store", None):
        return redirect("merchant_products")
    return redirect("control_products_list")


@support_required
def control_gallery_delete_ajax(request, pk):
    item = get_object_or_404(ProductImage, pk=pk)
    store = getattr(request, "store", None)
    if store and item.product.store != store:
        return JsonResponse({"status": "error", "message": "غير مصرح"}, status=403)
    item.delete()
    return JsonResponse({"status": "success"})


@support_required
def control_gallery_reorder_ajax(request, pk):
    item = get_object_or_404(ProductImage, pk=pk)
    store = getattr(request, "store", None)
    if store and item.product.store != store:
        return JsonResponse({"status": "error", "message": "غير مصرح"}, status=403)
        
    direction = request.POST.get("direction", "down")
    
    # Get all gallery items for this product
    items = list(ProductImage.objects.filter(product=item.product).order_by("sort_order", "id"))
    
    # Ensure they have sequential and unique sort_order values before sorting
    has_duplicates_or_zeros = len(set(i.sort_order for i in items)) < len(items) or any(i.sort_order == 0 for i in items)
    if has_duplicates_or_zeros:
        for idx_seq, it in enumerate(items):
            it.sort_order = idx_seq
            it.save(update_fields=["sort_order"])
            
    # Re-fetch sequential lists
    items = list(ProductImage.objects.filter(product=item.product).order_by("sort_order", "id"))
    
    try:
        idx = items.index(item)
    except ValueError:
        return JsonResponse({"status": "error", "message": "العنصر غير موجود"}, status=400)
        
    if direction == "up" and idx > 0:
        prev_item = items[idx - 1]
        temp = prev_item.sort_order
        prev_item.sort_order = item.sort_order
        item.sort_order = temp
        prev_item.save(update_fields=["sort_order"])
        item.save(update_fields=["sort_order"])
    elif direction == "down" and idx < len(items) - 1:
        next_item = items[idx + 1]
        temp = next_item.sort_order
        next_item.sort_order = item.sort_order
        item.sort_order = temp
        next_item.save(update_fields=["sort_order"])
        item.save(update_fields=["sort_order"])
        
    return JsonResponse({"status": "success"})



@support_required
def control_variant_keys(request, pk):
    variant = get_object_or_404(ProductVariant.objects.select_related('product'), pk=pk)
    
    if request.method == "POST":
        action = request.POST.get("action")
        if action == "add_keys":
            keys_text = request.POST.get("keys_text", "").strip()
            if keys_text:
                lines = [line.strip() for line in keys_text.replace(",", "\n").split("\n") if line.strip()]
                added_count = 0
                for code in lines:
                    ProductKey.objects.create(variant=variant, key_code=code)
                    added_count += 1
                messages.success(request, f"تم إضافة {added_count} مفتاح/كود بنجاح.")
            else:
                messages.error(request, "يرجى كتابة كود واحد على الأقل.")
                
        elif action == "delete_key":
            key_id = request.POST.get("key_id")
            key_obj = get_object_or_404(ProductKey, pk=key_id, variant=variant)
            key_obj.delete()
            messages.success(request, "تم حذف المفتاح بنجاح.")
            
        elif action == "delete_all_unused":
            deleted_count = ProductKey.objects.filter(variant=variant, is_used=False).delete()[0]
            messages.success(request, f"تم حذف {deleted_count} مفتاح غير مستخدم بنجاح.")
            
        return redirect("control_variant_keys", pk=variant.id)
        
    keys_list = ProductKey.objects.filter(variant=variant).select_related('used_by', 'order').order_by('-created_at')
    
    paginator = Paginator(keys_list, 50)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    total_keys = keys_list.count()
    used_keys = keys_list.filter(is_used=True).count()
    unused_keys = total_keys - used_keys
    
    context = {
        "variant": variant,
        "product": variant.product,
        "page_obj": page_obj,
        "total_keys": total_keys,
        "used_keys": used_keys,
        "unused_keys": unused_keys,
    }
    return render(request, "site/control_variant_keys.html", context)


@admin_required
def control_announcements(request): return render(request, "site/control_announcements.html", {"announcements": SiteAnnouncement.objects.all().order_by("-created_at")})

@admin_required
def control_announcement_create(request):
    is_merchant = getattr(request, "store", None) is not None
    form = SiteAnnouncementForm(request.POST or None, is_merchant=is_merchant)
    if request.method == "POST" and form.is_valid():
        target_store = request.store if is_merchant else form.cleaned_data.get("store")
        if form.cleaned_data.get("is_active"):
            SiteAnnouncement.objects.filter(store=target_store, is_active=True).update(is_active=False)
        ann = form.save(commit=False)
        ann.store = target_store
        ann.save()
        return redirect("control_announcements")
    return render(request, "site/control_announcement_form.html", {"form": form})

@admin_required
def control_announcement_edit(request, pk):
    ann = get_object_or_404(SiteAnnouncement, pk=pk)
    is_merchant = getattr(request, "store", None) is not None
    form = SiteAnnouncementForm(request.POST or None, instance=ann, is_merchant=is_merchant)
    if request.method == "POST" and form.is_valid():
        target_store = request.store if is_merchant else form.cleaned_data.get("store")
        if form.cleaned_data.get("is_active"):
            SiteAnnouncement.objects.filter(store=target_store, is_active=True).exclude(pk=pk).update(is_active=False)
        ann_obj = form.save(commit=False)
        ann_obj.store = target_store
        ann_obj.save()
        return redirect("control_announcements")
    return render(request, "site/control_announcement_form.html", {"form": form})

@admin_required
def control_announcement_delete(request, pk): get_object_or_404(SiteAnnouncement, pk=pk).delete(); return redirect("control_announcements")

@admin_required
def control_social_media(request):
    from apps.site.forms import SocialMediaLinkForm
    from django.core.exceptions import ValidationError
    if request.method == "POST":
        pk_val = request.POST.get("pk")
        instance = None
        if pk_val:
            try:
                instance = SocialMediaLink.objects.filter(pk=pk_val).first()
            except (ValueError, ValidationError):
                pass
        f = SocialMediaLinkForm(request.POST, request.FILES, instance=instance)
        if f.is_valid(): f.save(); messages.success(request, "تم الحفظ."); return redirect("control_social_media")
    return render(request, "site/control_social_media.html", {"links": SocialMediaLink.objects.all(), "form": SocialMediaLinkForm()})

@admin_required
def control_social_media_delete(request, pk): get_object_or_404(SocialMediaLink, pk=pk).delete(); return redirect("control_social_media")

@support_required
def ajax_user_search(request):
    store = getattr(request, "store", None)
    q = request.GET.get('q', '').strip()
    if len(q) < 2:
        return JsonResponse({"results": []})
    
    if store:
        users = User.objects.filter(store=store)
    else:
        users = User.objects.filter(store__isnull=True)
        
    users = users.filter(
        Q(email__icontains=q) | 
        Q(phone__icontains=q) | 
        Q(first_name__icontains=q) | 
        Q(last_name__icontains=q)
    ).distinct()[:20]
    
    results = []
    for u in users:
        results.append({
            "id": str(u.id),
            "text": f"{u.get_full_name() or 'No Name'} ({u.email}) - {u.phone or 'No Phone'}"
        })
    return JsonResponse({"results": results})

@admin_required
def ajax_country_search(request):
    from apps.common.countries import COUNTRIES
    q = request.GET.get('q', '').lower()
    results = []
    for code, name in COUNTRIES:
        if q in name.lower() or q in code.lower():
            results.append({"id": code, "text": f"{name} ({code})"})
    return JsonResponse({"results": results})

@admin_required
def ajax_product_search(request):
    q = request.GET.get('q', '').strip()
    
    if q:
        products = Product.objects.filter(
            Q(name__icontains=q) |
            Q(variants__sku__icontains=q)
        ).distinct()[:20]
    else:
        # Show recent or active products if no query
        products = Product.objects.filter(is_active=True).order_by('-created_at')[:20]

    results = []
    for p in products:
        results.append({
            "id": str(p.id),
            "text": f"{p.name} (ID: {str(p.id)[:8]})"
        })
    return JsonResponse({"results": results})

@finance_required
def control_debts(request):
    store = getattr(request, "store", None)
    q = request.GET.get('q', '')
    
    if store:
        users = User.objects.filter(store=store)
    else:
        users = User.objects.filter(store__isnull=True)
        
    if q:
        users = users.filter(Q(email__icontains=q) | Q(phone__icontains=q))
    
    if request.method == "POST":
        if store:
            target = get_object_or_404(User, id=request.POST.get("user_id"), store=store)
        else:
            target = get_object_or_404(User, id=request.POST.get("user_id"), store__isnull=True)
            
        amt = Decimal(request.POST.get("amount", "0"))
        action = request.POST.get("action")
        reason = request.POST.get("reason", "")
        
        if action == "add_debt":
            from apps.wallets.services import add_debt
            add_debt(target.wallet.id, amt, f"admin_debt_{timezone.now().timestamp()}", reason, request.user)
            messages.success(request, f"تم إضافة دين بقيمة {amt} للمستخدم {target.email}")
            
        elif action == "pay_debt":
            from apps.wallets.services import pay_debt
            pay_mode = request.POST.get("pay_mode", "balance") # balance or cash
            
            # If cash, we don't deduct from balance, just reduce debt_balance
            deduct = True if pay_mode == "balance" else False
            
            try:
                pay_debt(
                    target.wallet.id, 
                    amt, 
                    f"admin_pay_{timezone.now().timestamp()}", 
                    reason, 
                    request.user, 
                    deduct_from_balance=deduct,
                    source="admin_cash" if pay_mode == "cash" else "admin"
                )
                messages.success(request, f"تم تسجيل سداد بقيمة {amt} للمستخدم {target.email} ({'نقداً' if pay_mode == 'cash' else 'من الرصيد'})")
            except Exception as e:
                messages.error(request, str(e))
                
        elif action == "toggle_debt_withdrawable":
            target.wallet.debt_is_withdrawable = not target.wallet.debt_is_withdrawable
            target.wallet.save(update_fields=["debt_is_withdrawable"])
            messages.success(request, f"تم {'تفعيل' if target.wallet.debt_is_withdrawable else 'تعطيل'} خيار سحب الدين للمستخدم {target.email}")
                
        return redirect(f"{request.path}?q={q}")

    # Stats for the sidebar
    from django.db.models import Sum
    from apps.wallets.models import LedgerEntry
    
    if store:
        total_debt = Wallet.objects.filter(user__store=store).aggregate(total=Sum('debt_balance'))['total'] or 0
        today_payments = LedgerEntry.objects.filter(
            wallet__user__store=store,
            entry_type=LedgerEntry.EntryType.DEBT_PAYMENT,
            created_at__date=timezone.now().date()
        ).aggregate(total=Sum('amount'))['total'] or 0
        recent_logs = LedgerEntry.objects.filter(
            wallet__user__store=store,
            entry_type__in=[LedgerEntry.EntryType.DEBT_ADD, LedgerEntry.EntryType.DEBT_PAYMENT]
        ).select_related('wallet__user').order_by('-created_at')[:10]
    else:
        total_debt = Wallet.objects.filter(user__store__isnull=True).aggregate(total=Sum('debt_balance'))['total'] or 0
        today_payments = LedgerEntry.objects.filter(
            wallet__user__store__isnull=True,
            entry_type=LedgerEntry.EntryType.DEBT_PAYMENT,
            created_at__date=timezone.now().date()
        ).aggregate(total=Sum('amount'))['total'] or 0
        recent_logs = LedgerEntry.objects.filter(
            wallet__user__store__isnull=True,
            entry_type__in=[LedgerEntry.EntryType.DEBT_ADD, LedgerEntry.EntryType.DEBT_PAYMENT]
        ).select_related('wallet__user').order_by('-created_at')[:10]
    
    ctx = {
        "users": users, 
        "query": q,
        "total_debt_balance": total_debt,
        "today_debt_payments": today_payments,
        "recent_debt_logs": recent_logs
    }
    return render(request, "site/control_debts.html", ctx)

@support_required
def control_send_notification(request):
    store = getattr(request, "store", None)
    form = SendNotificationForm(request.POST or None)
    if request.method == "POST" and form.is_valid():
        if store:
            users = User.objects.filter(is_active=True, store=store)
        else:
            users = User.objects.filter(is_active=True, store__isnull=True)
            
        if form.cleaned_data["target"] == "tier":
            users = users.filter(tier=form.cleaned_data["tier"])
        elif form.cleaned_data["target"] == "individual":
            q = form.cleaned_data["user_identifier"]
            users = users.filter(
                Q(email__icontains=q) | 
                Q(first_name__icontains=q) | 
                Q(last_name__icontains=q) | 
                Q(phone__icontains=q)
            )
        
        if users.exists():
            channel = form.cleaned_data["channels"]
            title = form.cleaned_data["title"]
            body = form.cleaned_data["body"]
            action_url = form.cleaned_data["action_url"]
            
            for user in users:
                # 1. In-App Notification
                if channel in ['all', 'in_app']:
                    Notification.objects.create(
                        user=user, title=title, body=body, action_url=action_url,
                        channel=Notification.Channel.IN_APP
                    )
                
                # 2. Push Notification
                if channel in ['all', 'push']:
                    notify_user(user, title, body, action_url=action_url, category='system')
                
                # 3. Email Notification
                if channel in ['all', 'email']:
                    send_brevo_email(user.email, user.get_full_name() or user.email, title, body)
            
            messages.success(request, f"تم الإرسال لـ {users.count()} مستخدم عبر القنوات المحددة.")
            return redirect("control_send_notification")
    return render(request, "site/control_notification_form.html", {"form": form})

@admin_required
def control_support_settings(request):
    obj, _ = SupportSettings.objects.get_or_create(id=1); form = SupportSettingsForm(request.POST or None, instance=obj)
    if request.method == "POST" and form.is_valid(): form.save(); messages.success(request, "تم الحفظ."); return redirect("control_support_settings")
    return render(request, "site/control_support_settings.html", {"form": form})

@support_required
def control_quick_replies(request): return render(request, "site/control_quick_replies.html", {"replies": ChatCannedReply.objects.all().order_by("-created_at")})

@support_required
def control_quick_reply_create(request):
    form = ChatCannedReplyForm(request.POST or None)
    if request.method == "POST" and form.is_valid(): form.save(); return redirect("control_quick_replies")
    return render(request, "site/control_quick_reply_form.html", {"form": form})

@support_required
def control_quick_reply_edit(request, pk):
    reply = get_object_or_404(ChatCannedReply, pk=pk); form = ChatCannedReplyForm(request.POST or None, instance=reply)
    if request.method == "POST" and form.is_valid(): form.save(); return redirect("control_quick_replies")
    return render(request, "site/control_quick_reply_form.html", {"form": form})

@support_required
def control_quick_reply_delete(request, pk): get_object_or_404(ChatCannedReply, pk=pk).delete(); return redirect("control_quick_replies")

@support_required
def control_support_chat_open(request):
    form = AdminChatForm(request.POST or None)
    if request.method == "POST" and form.is_valid():
        q = form.cleaned_data["user_identifier"]
        user = User.objects.filter(
            Q(email__icontains=q) | 
            Q(first_name__icontains=q) | 
            Q(last_name__icontains=q) | 
            Q(phone__icontains=q)
        ).first()
        
        if user:
            with transaction.atomic():
                room = ChatRoom.objects.create(user=user, assigned_agent=request.user, subject=form.cleaned_data["subject"], status=ChatRoom.Status.ASSIGNED)
                ChatMessage.objects.create(room=room, sender=request.user, text=form.cleaned_data["message"], is_staff_reply=True); room.unread_user_count = 1; room.save()
                notify_user(user, title="رسالة من الدعم", body=room.subject, action_url=reverse("dashboard"), category='support')
            messages.success(request, f"تم فتح التذكرة بنجاح مع {user.get_full_name() or user.email}.")
            return redirect("control_dashboard")
        messages.error(request, "المستخدم غير موجود.")
    return render(request, "site/control_support_chat_open.html", {"form": form})

@admin_required
def control_coupons_list(request): return render(request, "site/control_coupons_list.html", {"coupons": Coupon.objects.all().order_by("-created_at")})
@admin_required
def control_coupon_create(request):
    form = CouponForm(request.POST or None)
    if request.method == "POST" and form.is_valid(): form.save(); return redirect("control_coupons_list")
    return render(request, "site/control_coupon_form.html", {"form": form})
@admin_required
def control_coupon_edit(request, pk):
    c = get_object_or_404(Coupon, pk=pk); form = CouponForm(request.POST or None, instance=c)
    if request.method == "POST" and form.is_valid(): form.save(); return redirect("control_coupons_list")
    return render(request, "site/control_coupon_form.html", {"form": form})
@admin_required
def control_coupon_delete(request, pk): get_object_or_404(Coupon, pk=pk).delete(); return redirect("control_coupons_list")

import csv
from django.http import HttpResponse

@admin_required
def export_coupon_usage_csv(request):
    response = HttpResponse(content_type='text/csv; charset=utf-8-sig')
    response['Content-Disposition'] = f'attachment; filename="coupon_usage_{timezone.now().strftime("%Y%m%d")}.csv"'
    
    writer = csv.writer(response)
    writer.writerow(['التاريخ', 'العميل', 'الكوبون', 'المنتج', 'الباقة', 'المبلغ الأصلي', 'المبلغ النهائي', 'رقم الطلب'])
    
    orders = Order.objects.filter(coupon__isnull=False).select_related('customer', 'coupon').prefetch_related('items__variant__product')
    
    for o in orders:
        item = o.items.first()
        writer.writerow([
            o.created_at.strftime("%Y-%m-%d %H:%M"),
            o.customer.email,
            o.coupon.code,
            item.variant.product.name if item else "N/A",
            item.variant.name if item else "N/A",
            o.original_total,
            o.total_amount,
            o.number
        ])
    return response

@admin_required
def control_coupon_usage(request):
    q = request.GET.get('q', '')
    coupon_filter = request.GET.get('coupon', '')
    
    orders = Order.objects.filter(coupon__isnull=False).select_related('customer', 'coupon').prefetch_related('items__variant__product').order_by('-created_at')
    
    if q:
        orders = orders.filter(Q(customer__email__icontains=q) | Q(number__icontains=q))
    if coupon_filter:
        orders = orders.filter(coupon__code__iexact=coupon_filter)
        
    return render(request, "site/control_coupon_usage.html", {
        "orders": orders, 
        "query": q, 
        "coupon_filter": coupon_filter
    })

@admin_required
def control_reports(request):
    from apps.site.analytics_services import FinancialAnalyticsService
    
    # Process GET parameters into filters
    filters = {
        "date_preset": request.GET.get("date_preset", "all"),
        "start_date": request.GET.get("start_date"),
        "end_date": request.GET.get("end_date"),
        "currency_id": request.GET.get("currency_id"),
        "payment_method_id": request.GET.get("payment_method_id"),
        "tier": request.GET.get("tier"),
        "user_id": request.GET.get("user_id"),
        "include_pending": request.GET.get("include_pending") == "on",
        "reporting_currency_code": request.GET.get("reporting_currency", "USD"),
    }
    
    # Clean empty strings
    clean_filters = {k: v for k, v in filters.items() if v}
    
    analytics = FinancialAnalyticsService(clean_filters, store=getattr(request, "store", None))
    
    from apps.wallets.models import RechargeCard
    from apps.catalog.models import ProductKey, Product
    from django.db.models import Sum, F

    recharge_cards_stats = {
        "total": RechargeCard.objects.count(),
        "active": RechargeCard.objects.filter(status='active').count(),
        "redeemed": RechargeCard.objects.filter(status='redeemed').count(),
        "cancelled": RechargeCard.objects.filter(status='cancelled').count(),
    }
    
    product_keys_stats = {
        "total": ProductKey.objects.count(),
        "unused": ProductKey.objects.filter(is_used=False).count(),
        "used": ProductKey.objects.filter(is_used=True).count(),
    }

    # Product Stock/Inventory Statistics
    store = getattr(request, "store", None)
    product_qs = Product.objects.filter(store=store) if store else Product.objects.all()
    low_stock_products = product_qs.filter(is_active=True, track_inventory=True, quantity__lte=F('low_stock_threshold'), quantity__gt=0)
    out_of_stock_products = product_qs.filter(is_active=True, track_inventory=True, quantity__lte=0)
    total_qty = product_qs.filter(track_inventory=True).aggregate(total=Sum('quantity'))['total'] or 0

    product_stock_stats = {
        "low_stock_count": low_stock_products.count(),
        "low_stock_list": low_stock_products[:5],
        "out_of_stock_count": out_of_stock_products.count(),
        "out_of_stock_list": out_of_stock_products[:5],
        "total_quantity": total_qty,
    }

    ctx = {
        "kpis": analytics.get_dashboard_kpis(),
        "pnl": analytics.get_pnl_statement(),
        "treasury": analytics.get_treasury_status(),
        "payment_performance": analytics.get_payment_method_performance(),
        "debt_aging": analytics.get_debt_aging(),
        "trends": analytics.get_trends(),
        "cash_logs": analytics.get_cash_collection_logs(),
        "filters": filters,
        "currencies": Currency.objects.filter(is_active=True),
        "payment_methods": PaymentMethod.objects.filter(is_active=True),
        "tiers": User.Tier.choices,
        "recharge_cards_stats": recharge_cards_stats,
        "product_keys_stats": product_keys_stats,
        "product_stock_stats": product_stock_stats,
    }
    
    export_format = request.GET.get("export")
    if export_format == "excel":
        return export_financial_report_xlsx(ctx, clean_filters)
        
    return render(request, "site/control_reports.html", ctx)

import logging
logger = logging.getLogger(__name__)

@admin_required
def control_db_maintenance(request):
    from apps.accounts.models import User
    from apps.common.tenant_utils import bypass_tenant_filter
    
    if request.method == "POST":
        action = request.POST.get("action")
        if action == "cleanup":
            targets = request.POST.getlist("targets")
            deleted_counts = {}
            
            from django.db.models import ProtectedError
            from decimal import Decimal
            from django.utils import timezone
            from django.contrib.auth.models import Group
            from django.contrib.sites.models import Site
            
            try:
                from rest_framework_simplejwt.token_blacklist.models import OutstandingToken, BlacklistedToken
            except ImportError:
                OutstandingToken, BlacklistedToken = None, None
                
            try:
                with bypass_tenant_filter():
                    with transaction.atomic():
                        # Define safe dependency execution order to prevent ProtectedError cascading
                        ordered_cleanup_keys = [
                            "social_tokens",
                            "social_accounts",
                            "social_apps",
                            "email_addresses",
                            "blacklisted_tokens",
                            "outstanding_tokens",
                            "invoices",
                            "coupons",
                            "orders",
                            "deposits",
                            "withdrawals",
                            "ledger_entries",
                            "wallet_transactions",
                            "balance_transfers",
                            "recharge_cards",
                            "wallets",
                            "store_employees",
                            "store_pages",
                            "store_settings",
                            "subscription_invoices",
                            "store_templates",
                            "saas_audit_logs",
                            "stores",
                            "subscription_plans",
                            "users",
                            "saas_admin_roles",
                            "saas_global_settings",
                            "payment_methods",
                            "payment_methods_reset",
                            "product_variants",
                            "products",
                            "categories",
                            "services",
                            "chat_rooms",
                            "chat_canned_replies",
                            "support_settings",
                            "platform_stats",
                            "currencies",
                            "social_links",
                            "system_audit_logs",
                            "testimonials",
                            "site_announcements",
                            "notifications",
                            "security_events",
                            "user_sessions",
                            "groups",
                            "sites"
                        ]
                        
                        for key in ordered_cleanup_keys:
                            if key in targets:
                                if key == "social_tokens":
                                    from allauth.socialaccount.models import SocialToken
                                    c = SocialToken.objects.all().delete()[0]
                                    deleted_counts["أكواد التطبيقات الاجتماعية"] = c
                                elif key == "social_accounts":
                                    from allauth.socialaccount.models import SocialAccount
                                    c = SocialAccount.objects.all().delete()[0]
                                    deleted_counts["حسابات تواصل اجتماعي"] = c
                                elif key == "social_apps":
                                    from allauth.socialaccount.models import SocialApp
                                    c = SocialApp.objects.all().delete()[0]
                                    deleted_counts["تطبيقات اجتماعية"] = c
                                elif key == "email_addresses":
                                    from allauth.account.models import EmailAddress
                                    c = EmailAddress.objects.all().delete()[0]
                                    deleted_counts["عناوين البريد الإلكتروني"] = c
                                elif key == "blacklisted_tokens":
                                    if BlacklistedToken:
                                        c = BlacklistedToken.objects.all().delete()[0]
                                        deleted_counts["رموز مميزة محظورة"] = c
                                elif key == "outstanding_tokens":
                                    if OutstandingToken:
                                        c = OutstandingToken.objects.all().delete()[0]
                                        deleted_counts["رموز مميزة نشطة"] = c
                                elif key == "coupons":
                                    from apps.orders.models import Coupon
                                    c = Coupon.objects.all().delete()[0]
                                    deleted_counts["الكوبونات"] = c
                                elif key == "invoices":
                                    from apps.orders.models import Invoice
                                    c = Invoice.objects.all().delete()[0]
                                    deleted_counts["الفواتير"] = c
                                elif key == "orders":
                                    from apps.orders.models import Order, OrderItem, OrderLog
                                    c1 = OrderItem.objects.all().delete()[0]
                                    c2 = OrderLog.objects.all().delete()[0]
                                    c3 = Order.objects.all().delete()[0]
                                    deleted_counts["الطلبات والمبيعات"] = c1 + c2 + c3
                                elif key == "deposits":
                                    from apps.payments.models import DepositRequest
                                    c = DepositRequest.objects.all().delete()[0]
                                    deleted_counts["طلبات الإيداع"] = c
                                elif key == "withdrawals":
                                    from apps.payments.models import WithdrawalRequest
                                    c = WithdrawalRequest.objects.all().delete()[0]
                                    deleted_counts["طلبات السحب"] = c
                                elif key == "ledger_entries":
                                    from apps.wallets.models import LedgerEntry
                                    c = LedgerEntry.objects.all().delete()[0]
                                    deleted_counts["سجلات حركة المحفظة"] = c
                                elif key == "wallet_transactions":
                                    from apps.wallets.models import WalletTransaction
                                    c = WalletTransaction.objects.all().delete()[0]
                                    deleted_counts["العمليات المالية للمحافظ"] = c
                                elif key == "balance_transfers":
                                    from apps.wallets.models import BalanceTransfer
                                    c = BalanceTransfer.objects.all().delete()[0]
                                    deleted_counts["تحويلات الأرصدة"] = c
                                elif key == "recharge_cards":
                                    from apps.wallets.models import RechargeCard
                                    c = RechargeCard.objects.all().delete()[0]
                                    deleted_counts["بطاقات الشحن"] = c
                                elif key == "wallets":
                                    from apps.wallets.models import Wallet
                                    c = Wallet.objects.all().delete()[0]
                                    deleted_counts["المحافظ"] = c
                                elif key == "store_employees":
                                    from apps.stores.models import StoreEmployee
                                    c = StoreEmployee.objects.all().delete()[0]
                                    deleted_counts["موظفو المتاجر"] = c
                                elif key == "store_pages":
                                    from apps.stores.models import StorePage
                                    c = StorePage.objects.all().delete()[0]
                                    deleted_counts["صفحات المتاجر"] = c
                                elif key == "store_settings":
                                    from apps.stores.models import StoreSetting
                                    c = StoreSetting.objects.all().delete()[0]
                                    deleted_counts["إعدادات المتاجر"] = c
                                elif key == "subscription_invoices":
                                    from apps.stores.models import SubscriptionInvoice
                                    c = SubscriptionInvoice.objects.all().delete()[0]
                                    deleted_counts["فواتير اشتراكات المتاجر"] = c
                                elif key == "store_templates":
                                    from apps.stores.models import StoreTemplate
                                    c = StoreTemplate.objects.all().delete()[0]
                                    deleted_counts["قوالب المتاجر"] = c
                                elif key == "saas_audit_logs":
                                    from apps.stores.models import SaaSAuditLog
                                    c = SaaSAuditLog.objects.all().delete()[0]
                                    deleted_counts["سجلات تدقيق SaaS"] = c
                                elif key == "stores":
                                    from apps.stores.models import Store
                                    User.objects.all().update(store=None)
                                    c = Store.objects.all().delete()[0]
                                    deleted_counts["المتاجر"] = c
                                elif key == "subscription_plans":
                                    from apps.stores.models import SubscriptionPlan
                                    c = SubscriptionPlan.objects.all().delete()[0]
                                    deleted_counts["خطط اشتراكات SaaS"] = c
                                elif key == "users":
                                    c = User.objects.exclude(is_superuser=True).exclude(is_staff=True).exclude(role__in=[User.Role.SUPER_ADMIN, User.Role.ADMIN]).exclude(id=request.user.id).delete()[0]
                                    deleted_counts["المستخدمين (غير المدراء)"] = c
                                elif key == "saas_admin_roles":
                                    from apps.stores.models import SaaSAdminRole
                                    c = SaaSAdminRole.objects.all().delete()[0]
                                    deleted_counts["أدوار SaaS الإدارية"] = c
                                elif key == "saas_global_settings":
                                    from apps.stores.models import SaaSGlobalSetting
                                    c = SaaSGlobalSetting.objects.all().delete()[0]
                                    deleted_counts["إعدادات عامة SaaS"] = c
                                elif key == "payment_methods":
                                    from apps.payments.models import PaymentMethod
                                    c = PaymentMethod.objects.all().delete()[0]
                                    deleted_counts["وسائل الدفع"] = c
                                elif key == "payment_methods_reset":
                                    from apps.payments.models import PaymentMethod
                                    PaymentMethod.objects.all().update(
                                        daily_deposit_usage=Decimal("0.00"),
                                        daily_withdrawal_usage=Decimal("0.00"),
                                        last_limit_reset=timezone.now()
                                    )
                                    User.objects.all().update(
                                        daily_deposit_usage=Decimal("0.00"),
                                        daily_withdrawal_usage=Decimal("0.00"),
                                        last_limit_reset=timezone.now()
                                    )
                                    deleted_counts["حدود الاستخدام اليومية"] = "تم التصفير"
                                elif key == "product_variants":
                                    from apps.catalog.models import ProductVariant
                                    c = ProductVariant.objects.all().delete()[0]
                                    deleted_counts["باقات المنتجات"] = c
                                elif key == "products":
                                    from apps.catalog.models import Product
                                    c = Product.objects.all().delete()[0]
                                    deleted_counts["المنتجات"] = c
                                elif key == "categories":
                                    from apps.catalog.models import Category
                                    c = Category.objects.all().delete()[0]
                                    deleted_counts["الأقسام والتصنيفات"] = c
                                elif key == "services":
                                    from apps.services.models import Service
                                    c = Service.objects.all().delete()[0]
                                    deleted_counts["الخدمات"] = c
                                elif key == "chat_rooms":
                                    from apps.support.models import ChatRoom
                                    c = ChatRoom.objects.all().delete()[0]
                                    deleted_counts["غرف محادثات الدعم"] = c
                                elif key == "chat_canned_replies":
                                    from apps.support.models import ChatCannedReply
                                    c = ChatCannedReply.objects.all().delete()[0]
                                    deleted_counts["الردود الجاهزة"] = c
                                elif key == "support_settings":
                                    from apps.support.models import SupportSettings
                                    c = SupportSettings.objects.all().delete()[0]
                                    deleted_counts["إعدادات الدعم الفني"] = c
                                elif key == "platform_stats":
                                    from apps.common.models import PlatformStatistic
                                    c = PlatformStatistic.objects.all().delete()[0]
                                    deleted_counts["إحصائيات المنصة"] = c
                                elif key == "currencies":
                                    from apps.common.models import Currency
                                    c = Currency.objects.all().delete()[0]
                                    deleted_counts["العملات"] = c
                                elif key == "social_links":
                                    from apps.common.models import SocialMediaLink
                                    c = SocialMediaLink.objects.all().delete()[0]
                                    deleted_counts["روابط التواصل الاجتماعي"] = c
                                elif key == "system_audit_logs":
                                    from apps.common.models import SystemAuditLog
                                    c = SystemAuditLog.objects.all().delete()[0]
                                    deleted_counts["سجلات تدقيق النظام"] = c
                                elif key == "testimonials":
                                    from apps.common.models import Testimonial
                                    c = Testimonial.objects.all().delete()[0]
                                    deleted_counts["شهادات العملاء"] = c
                                elif key == "site_announcements":
                                    from apps.common.models import SiteAnnouncement
                                    c = SiteAnnouncement.objects.all().delete()[0]
                                    deleted_counts["ملاحظات شريط الموقع"] = c
                                elif key == "notifications":
                                    from apps.notifications.models import Notification
                                    c = Notification.objects.all().delete()[0]
                                    deleted_counts["الإشعارات"] = c
                                elif key == "security_events":
                                    from apps.accounts.models import SecurityEvent
                                    c = SecurityEvent.objects.all().delete()[0]
                                    deleted_counts["سجلات الأمان"] = c
                                elif key == "user_sessions":
                                    from apps.accounts.models import UserSession
                                    c = UserSession.objects.all().delete()[0]
                                    deleted_counts["جلسات النشاط"] = c
                                elif key == "groups":
                                    c = Group.objects.all().delete()[0]
                                    deleted_counts["المجموعات الإدارية"] = c
                                elif key == "sites":
                                    c = Site.objects.all().delete()[0]
                                    from django.conf import settings
                                    Site.objects.create(id=settings.SITE_ID, domain="raqamiyatapp.com", name="Raqamiyat")
                                    deleted_counts["مواقع النظام"] = c

                msg = "تم تصفير البيانات المختارة بنجاح: " + ", ".join([f"{k} ({v})" for k, v in deleted_counts.items()])
                messages.success(request, msg)
                return redirect("control_db_maintenance")
            except ProtectedError as e:
                messages.error(request, f"لا يمكن حذف بعض البيانات لوجود ارتباطات محمية بها. تفاصيل الخطأ: {str(e)}")
                return redirect("control_db_maintenance")
            
    from apps.orders.models import Order, Invoice, Coupon
    from apps.accounts.models import User, SecurityEvent, UserSession
    from apps.payments.models import DepositRequest, WithdrawalRequest, PaymentMethod
    from apps.catalog.models import Product, Category, ProductVariant
    from apps.common.models import PlatformStatistic, Currency, SocialMediaLink, SystemAuditLog, Testimonial, SiteAnnouncement
    from apps.notifications.models import Notification
    from apps.services.models import Service
    from apps.support.models import SupportSettings, ChatCannedReply, ChatRoom
    from apps.wallets.models import Wallet, WalletTransaction, LedgerEntry, BalanceTransfer, RechargeCard
    from apps.stores.models import Store, StoreSetting, SaaSGlobalSetting, SaaSAdminRole, SubscriptionPlan, SaaSAuditLog, StorePage, SubscriptionInvoice, StoreTemplate, StoreEmployee
    from allauth.account.models import EmailAddress
    from django.contrib.auth.models import Group
    from django.contrib.sites.models import Site
    from allauth.socialaccount.models import SocialAccount, SocialApp, SocialToken
    
    try:
        from rest_framework_simplejwt.token_blacklist.models import OutstandingToken, BlacklistedToken
    except ImportError:
        OutstandingToken, BlacklistedToken = None, None

    with bypass_tenant_filter():
        stats = {
            "users": User.objects.count(),
            "security_events": SecurityEvent.objects.count(),
            "user_sessions": UserSession.objects.count(),
            
            "categories": Category.objects.count(),
            "products": Product.objects.count(),
            "product_variants": ProductVariant.objects.count(),
            
            "platform_stats": PlatformStatistic.objects.count(),
            "currencies": Currency.objects.count(),
            "social_links": SocialMediaLink.objects.count(),
            "system_audit_logs": SystemAuditLog.objects.count(),
            "testimonials": Testimonial.objects.count(),
            "site_announcements": SiteAnnouncement.objects.count(),
            
            "notifications": Notification.objects.count(),
            
            "orders": Order.objects.count(),
            "invoices": Invoice.objects.count(),
            "coupons": Coupon.objects.count(),
            
            "deposits": DepositRequest.objects.count(),
            "withdrawals": WithdrawalRequest.objects.count(),
            "payment_methods": PaymentMethod.objects.count(),
            
            "services": Service.objects.count(),
            
            "support_settings": SupportSettings.objects.count(),
            "chat_canned_replies": ChatCannedReply.objects.count(),
            "chat_rooms": ChatRoom.objects.count(),
            
            "outstanding_tokens": OutstandingToken.objects.count() if OutstandingToken else 0,
            "blacklisted_tokens": BlacklistedToken.objects.count() if BlacklistedToken else 0,
            
            "wallets": Wallet.objects.count(),
            "wallet_transactions": WalletTransaction.objects.count(),
            "ledger_entries": LedgerEntry.objects.count(),
            "balance_transfers": BalanceTransfer.objects.count(),
            "recharge_cards": RechargeCard.objects.count(),
            
            "stores": Store.objects.count(),
            "store_settings": StoreSetting.objects.count(),
            "saas_global_settings": SaaSGlobalSetting.objects.count(),
            "saas_admin_roles": SaaSAdminRole.objects.count(),
            "subscription_plans": SubscriptionPlan.objects.count(),
            "saas_audit_logs": SaaSAuditLog.objects.count(),
            "store_pages": StorePage.objects.count(),
            "subscription_invoices": SubscriptionInvoice.objects.count(),
            "store_templates": StoreTemplate.objects.count(),
            "store_employees": StoreEmployee.objects.count(),
            
            "email_addresses": EmailAddress.objects.count(),
            "groups": Group.objects.count(),
            "sites": Site.objects.count(),
            
            "social_accounts": SocialAccount.objects.count(),
            "social_apps": SocialApp.objects.count(),
            "social_tokens": SocialToken.objects.count(),
        }
        
    return render(request, "site/control_db_maintenance.html", {"stats": stats})

@admin_required
def control_geo_stats(request):
    from django.db.models import Count
    
    # 1. Country Stats
    country_stats = User.objects.exclude(last_country='').values('last_country').annotate(count=Count('id')).order_by('-count')
    total_users = User.objects.exclude(last_country='').count() or 1
    
    countries = []
    for c in country_stats:
        countries.append({
            "name": c['last_country'],
            "count": c['count'],
            "percent": round((c['count'] / total_users) * 100, 1)
        })
        
    # 2. City Stats
    city_stats = User.objects.exclude(last_city='').values('last_city', 'last_country').annotate(count=Count('id')).order_by('-count')[:20]
    
    return render(request, "site/control_geo_stats.html", {
        "countries": countries,
        "cities": city_stats,
        "total_users": total_users
    })

def export_financial_report_xlsx(ctx, filters):
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
        from io import BytesIO

        wb = openpyxl.Workbook()
        # ... rest of export logic ...
    except Exception as e:
        logger.exception("Excel export failed")
        return HttpResponse(
            "خدمة تصدير Excel غير متوفرة حالياً (تأكد من تثبيت openpyxl). يرجى التواصل مع المسؤول.",
            status=503,
            content_type="text/plain; charset=utf-8"
        )

    # --- Sheet 1: الملخص المالي ---
    ws = wb.active
    ws.title = "الملخص المالي"
    ws.sheet_view.rightToLeft = True
    
    # Styles
    header_font = Font(name='Arial', bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="06b6d4", end_color="06b6d4", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center")
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    # Branding & Header
    ws.merge_cells('A1:D1')
    ws['A1'] = "تقرير رقميات المالي - Raqamiyat Financial Report"
    ws['A1'].font = Font(bold=True, size=16)
    ws['A1'].alignment = center_align
    
    ws['A2'] = "تاريخ الاستخراج:"
    ws['B2'] = timezone.now().strftime("%Y-%m-%d %H:%M")
    
    ws['A3'] = "عملة التقرير:"
    ws['B3'] = ctx['kpis']['reporting_currency']

    # KPIs Section
    ws['A5'] = "المؤشر"
    ws['B5'] = "القيمة"
    for cell in ['A5', 'B5']:
        ws[cell].font = header_font
        ws[cell].fill = header_fill
        ws[cell].alignment = center_align

    kpis_data = [
        ("إجمالي الإيداعات", ctx['kpis']['total_deposits']),
        ("إجمالي السحوبات", ctx['kpis']['total_withdrawals']),
        ("صافي التدفق النقدي", ctx['kpis']['net_cashflow']),
        ("رسوم العمليات", ctx['kpis']['total_fees_earned']),
        ("أرباح المنتجات الصافية", ctx['kpis']['product_net_profit']),
        ("المقبوضات النقدية (Cash)", ctx['kpis']['total_cash_collections']),
        ("الديون المستحقة", ctx['kpis']['total_outstanding_debt']),
        ("التزامات المحافظ", ctx['kpis']['total_liabilities']),
    ]
    
    row = 6
    for label, val in kpis_data:
        ws.cell(row=row, column=1, value=label).border = border
        ws.cell(row=row, column=2, value=float(val or 0)).border = border
        ws.cell(row=row, column=2).number_format = '#,##0.00'
        row += 1

    # --- Sheet 2: الأداء والسيولة ---
    ws2 = wb.create_sheet("أداء وسائل الدفع")
    ws2.sheet_view.rightToLeft = True
    
    headers = ["الوسيلة", "حجم الإيداعات", "حجم السحوبات", "الصافي", "الرسوم", "الرصيد التقديري"]
    for col, h in enumerate(headers, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        
    row = 2
    for pm in ctx['payment_performance']:
        ws2.cell(row=row, column=1, value=pm.get('name', 'N/A')).border = border
        ws2.cell(row=row, column=2, value=float(pm.get('deposits_volume', 0) or 0)).border = border
        ws2.cell(row=row, column=3, value=float(pm.get('withdrawals_volume', 0) or 0)).border = border
        ws2.cell(row=row, column=4, value=float(pm.get('net_movement', 0) or 0)).border = border
        ws2.cell(row=row, column=5, value=float(pm.get('fees_generated', 0) or 0)).border = border
        ws2.cell(row=row, column=6, value=float(pm.get('real_balance', 0) or 0)).border = border
        row += 1

    # Auto-adjust column widths
    for sheet in wb.worksheets:
        for column in sheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except: pass
            sheet.column_dimensions[column_letter].width = max_length + 2

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="Raqamiyat_Financial_{timezone.now().strftime("%Y%m%d")}.xlsx"'
    return response
@support_required
def control_variant_create(request, product_pk): return redirect("control_product_edit", pk=product_pk)

@support_required
def control_reorder_product(request, pk):
    product = get_object_or_404(Product, pk=pk)
    direction = request.GET.get('direction')
    
    # Ensure all products have unique sort orders for reliable swapping
    # This is a one-time-per-call normalization to handle duplicates/zeros
    all_products = list(Product.objects.all().order_by('sort_order', 'id'))
    for i, p in enumerate(all_products):
        if p.sort_order != i:
            p.sort_order = i
            p.save(update_fields=['sort_order'])
    
    # Re-fetch after normalization
    product.refresh_from_db()
    
    if direction == 'up':
        other = Product.objects.filter(sort_order__lt=product.sort_order).order_by('-sort_order').first()
    elif direction == 'down':
        other = Product.objects.filter(sort_order__gt=product.sort_order).order_by('sort_order').first()
    else:
        return redirect('control_products_list')
    
    if other:
        # Swap
        p_order = product.sort_order
        o_order = other.sort_order
        product.sort_order = o_order
        other.sort_order = p_order
        product.save(update_fields=['sort_order'])
        other.save(update_fields=['sort_order'])
        
    return redirect('control_products_list')
@support_required
def control_variant_edit(request, pk): v = get_object_or_404(ProductVariant, pk=pk); return redirect("control_product_edit", pk=v.product.id)

def terms_of_service(request): return render(request, "site/terms_of_service.html")
def refund_policy(request): return render(request, "site/refund_policy.html")
def contact_page(request): return render(request, "site/contact.html")
def privacy_policy(request): return render(request, "site/privacy_policy.html")
def service_worker(request): return HttpResponse(open("apps/site/static/site/js/sw.js").read() if os.path.exists("apps/site/static/site/js/sw.js") else "", content_type="application/javascript")
def set_currency(request):
    curr = Currency.objects.filter(id=request.GET.get("currency") or request.POST.get("currency"), is_active=True).first()
    if curr:
        request.session["preferred_currency_id"] = str(curr.id)
        if request.user.is_authenticated: request.user.preferred_currency = curr; request.user.save(update_fields=["preferred_currency"])
    return redirect(request.META.get('HTTP_REFERER', 'home'))

@admin_required
def payment_methods_list(request): return render(request, "site/payment_methods_list.html", {"methods": PaymentMethod.objects.all().order_by("display_order")})
@admin_required
def payment_method_create(request):
    form = PaymentMethodForm(request.POST or None, request.FILES or None)
    if request.method == "POST" and form.is_valid(): form.save(); return redirect("payment_methods_list")
    return render(request, "site/payment_method_builder.html", {"form": form})
@admin_required
def payment_method_edit(request, pk):
    method = get_object_or_404(PaymentMethod, pk=pk)
    old_rate = method.capital_exchange_rate
    form = PaymentMethodForm(request.POST or None, request.FILES or None, instance=method)
    if request.method == "POST" and form.is_valid():
        saved_method = form.save()
        if saved_method.capital_exchange_rate != old_rate:
            from apps.payments.models import PaymentMethodExchangeRateLog
            PaymentMethodExchangeRateLog.objects.create(
                payment_method=saved_method,
                old_rate=old_rate,
                new_rate=saved_method.capital_exchange_rate,
                changed_by=request.user,
                reason="Admin Manual Update"
            )
        return redirect("payment_methods_list")
    return render(request, "site/payment_method_builder.html", {"form": form, "method": method})

@login_required
def site_notification_settings(request):
    from apps.site.forms import UserPrivacyForm
    obj, _ = NotificationSetting.objects.get_or_create(user=request.user)
    form = NotificationSettingForm(request.POST or None, instance=obj, is_staff=request.user.is_platform_staff)
    privacy_form = UserPrivacyForm(request.POST or None, instance=request.user)
    
    if request.method == "POST":
        if "notification_settings" in request.POST and form.is_valid():
            form.save()
            messages.success(request, "تم حفظ إعدادات الإشعارات بنجاح.")
            return redirect("notification_settings")
        elif "privacy_settings" in request.POST and privacy_form.is_valid():
            privacy_form.save()
            messages.success(request, "تم تحديث إعدادات الخصوصية بنجاح.")
            return redirect("notification_settings")
            
    return render(request, "site/v3/v3_notification_settings.html", {
        "form": form,
        "privacy_form": privacy_form
    })

@login_required
def site_product_suggestion(request):
    form = ProductSuggestionForm(request.POST or None)
    if request.method == "POST":
        if form.is_valid():
            suggestion = form.save(commit=False)
            suggestion.user = request.user
            suggestion.store = getattr(request, "store", None)
            suggestion.save()
            
            # Support AJAX form submission
            if request.headers.get('x-requested-with') == 'XMLHttpRequest':
                return JsonResponse({"status": "success", "message": "تم إرسال اقتراحك بنجاح."})
                
            messages.success(request, "تم إرسال اقتراحك بنجاح. سنقوم بمراجعته والرد عليك قريباً.")
            return redirect("dashboard")
        else:
            # Support AJAX form validation errors
            if request.headers.get('x-requested-with') == 'XMLHttpRequest':
                return JsonResponse({"status": "error", "message": "يرجى التحقق من الحقول المدخلة وصحتها.", "errors": form.errors.get_json_data()}, status=400)
    
    user_suggestions = ProductSuggestion.objects.filter(user=request.user)
    return render(request, "site/product_suggestion_form.html", {"form": form, "suggestions": user_suggestions})

@support_required
def control_product_suggestions_list(request):
    status_filter = request.GET.get('status')
    suggestions = ProductSuggestion.objects.all().select_related('user')
    if status_filter:
        suggestions = suggestions.filter(status=status_filter)
    
    return render(request, "site/control_product_suggestions_list.html", {
        "suggestions": suggestions,
        "status_filter": status_filter
    })

@support_required
def control_product_suggestion_detail(request, pk):
    suggestion = get_object_or_404(ProductSuggestion, pk=pk)
    if request.method == "POST":
        action = request.POST.get("action")
        notes = request.POST.get("admin_notes", "")
        
        if action == "approve":
            suggestion.status = ProductSuggestion.Status.APPROVED
        elif action == "reject":
            suggestion.status = ProductSuggestion.Status.REJECTED
        elif action == "implement":
            suggestion.status = ProductSuggestion.Status.IMPLEMENTED
            
        suggestion.admin_notes = notes
        suggestion.save()
        messages.success(request, "تم تحديث حالة الاقتراح.")
        return redirect("control_product_suggestions_list")
        
    return render(request, "site/control_product_suggestion_detail.html", {"suggestion": suggestion})


@admin_required
def control_recharge_cards(request):
    from apps.wallets.models import RechargeCard
    from apps.catalog.models import ProductVariant, ProductKey
    from django.core.paginator import Paginator
    from django.db.models import Q
    
    if request.method == "POST":
        action = request.POST.get("action")
        if action == "copy_to_variant":
            codes_list = request.POST.getlist("codes")
            variant_id = request.POST.get("variant_id")
            if not codes_list or not variant_id:
                messages.error(request, "يرجى اختيار الأكواد وباقة المنتج المستهدفة.")
            else:
                variant = get_object_or_404(ProductVariant, id=variant_id)
                added_count = 0
                for code in codes_list:
                    if not ProductKey.objects.filter(variant=variant, key_code=code).exists():
                        ProductKey.objects.create(variant=variant, key_code=code)
                        added_count += 1
                messages.success(request, f"تم نسخ {added_count} كود بنجاح كأكواد تسليم تلقائي للباقة: {variant.product.name} - {variant.name}")
            return redirect("control_recharge_cards")

    qs = RechargeCard.objects.all().select_related("currency", "created_by", "redeemed_by", "order")
    
    q = request.GET.get("q", "").strip()
    if q:
        qs = qs.filter(
            Q(code__icontains=q) |
            Q(created_by__email__icontains=q) |
            Q(redeemed_by__email__icontains=q) |
            Q(created_by__uid__iexact=q) |
            Q(redeemed_by__uid__iexact=q)
        )
        
    status_filter = request.GET.get("status", "").strip()
    if status_filter:
        qs = qs.filter(status=status_filter)
        
    amount_filter = request.GET.get("amount", "").strip()
    if amount_filter:
        try:
            qs = qs.filter(amount=Decimal(amount_filter))
        except:
            pass
            
    date_from = request.GET.get("date_from", "").strip()
    if date_from:
        try:
            qs = qs.filter(created_at__date__gte=date_from)
        except:
            pass
            
    date_to = request.GET.get("date_to", "").strip()
    if date_to:
        try:
            qs = qs.filter(created_at__date__lte=date_to)
        except:
            pass
            
    sort_filter = request.GET.get("sort", "-created_at").strip()
    if sort_filter in ["created_at", "-created_at", "amount", "-amount"]:
        qs = qs.order_by(sort_filter)
    else:
        qs = qs.order_by("-created_at")
        
    paginator = Paginator(qs, 50)
    page_number = request.GET.get("page")
    page_obj = paginator.get_page(page_number)
    
    active_variants = ProductVariant.objects.filter(is_active=True, product__is_active=True, delivery_type='keys').select_related('product').order_by('product__name', 'name')

    return render(request, "site/control_recharge_cards.html", {
        "page_obj": page_obj,
        "query": q,
        "status_filter": status_filter,
        "amount_filter": amount_filter,
        "date_from": date_from,
        "date_to": date_to,
        "sort_filter": sort_filter,
        "status_choices": RechargeCard.Status.choices,
        "active_variants": active_variants
    })


@admin_required
def control_recharge_cards_generate(request):
    from apps.wallets.models import RechargeCard
    from apps.common.models import Currency
    from apps.catalog.models import ProductVariant, ProductKey
    from decimal import Decimal
    import secrets
    
    currencies = Currency.objects.filter(is_active=True)
    active_variants = ProductVariant.objects.filter(is_active=True, product__is_active=True, delivery_type='keys').select_related('product').order_by('product__name', 'name')
    
    if request.method == "POST":
        amount_str = request.POST.get("amount", "0")
        currency_id = request.POST.get("currency")
        count_str = request.POST.get("count", "1")
        variant_id = request.POST.get("variant_id")
        
        try:
            amount = Decimal(amount_str)
            if amount <= 0:
                raise ValueError()
        except:
            messages.error(request, "يرجى إدخل قيمة شحن صحيحة أكبر من صفر.")
            return redirect("control_recharge_cards_generate")
            
        currency = get_object_or_404(Currency, id=currency_id, is_active=True)
        
        try:
            count = int(count_str)
            if count <= 0 or count > 500:
                raise ValueError()
        except:
            messages.error(request, "يرجى إدخال عدد بطاقات صحيح (بين 1 و 500).")
            return redirect("control_recharge_cards_generate")
            
        target_variant = None
        if variant_id:
            target_variant = get_object_or_404(ProductVariant, id=variant_id)
            
        created_count = 0
        keys_added = 0
        for _ in range(count):
            code = f"RC-{secrets.token_hex(4).upper()}-{secrets.token_hex(4).upper()}"
            while RechargeCard.objects.filter(code=code).exists():
                code = f"RC-{secrets.token_hex(4).upper()}-{secrets.token_hex(4).upper()}"
                
            RechargeCard.objects.create(
                code=code,
                amount=amount,
                currency=currency,
                created_by=request.user,
                status=RechargeCard.Status.ACTIVE,
                store=getattr(request, "store", None)
            )
            created_count += 1
            
            if target_variant:
                ProductKey.objects.create(variant=target_variant, key_code=code)
                keys_added += 1
            
        msg = f"تم إنشاء {created_count} بطاقة شحن بنجاح بقيمة {amount} {currency.code} للبطاقة الواحدة."
        if target_variant:
            msg += f" وتم نسخ {keys_added} كود تلقائياً كأكواد تسليم تلقائي للباقة: {target_variant.product.name} - {target_variant.name}."
            
        messages.success(request, msg)
        return redirect("control_recharge_cards")
        
    return render(request, "site/control_recharge_card_generate.html", {
        "currencies": currencies,
        "active_variants": active_variants
    })


@admin_required
def control_recharge_card_cancel(request, pk):
    from apps.wallets.models import RechargeCard
    
    card = get_object_or_404(RechargeCard, pk=pk)
    if request.method == "POST":
        if card.status == RechargeCard.Status.ACTIVE:
            card.status = RechargeCard.Status.CANCELLED
            card.save()
            messages.success(request, f"تم إلغاء بطاقة الشحن ({card.code}) بنجاح.")
        else:
            messages.error(request, "لا يمكن إلغاء هذه البطاقة لأنها ليست نشطة.")
            
    return redirect("control_recharge_cards")


@login_required
def recharge_wallet(request):
    from apps.wallets.models import RechargeCard
    from apps.wallets.services import credit_wallet, WalletError, get_or_create_wallet
    from django.db import transaction
    from django.utils import timezone
    
    if request.method == "POST":
        code_input = request.POST.get("recharge_code", "").strip().upper()
        if not code_input:
            messages.error(request, "يرجى إدخال رمز الشحن.")
            return redirect("dashboard_wallet")
            
        card = RechargeCard.objects.filter(code=code_input).first()
        if not card or card.status != RechargeCard.Status.ACTIVE:
            messages.error(request, "رمز الشحن غير صحيح، أو تم استخدامه، أو ملغى.")
            return redirect("dashboard_wallet")
            
        try:
            with transaction.atomic():
                card = RechargeCard.objects.select_for_update().get(id=card.id)
                if card.status != RechargeCard.Status.ACTIVE:
                    raise WalletError("البطاقة لم تعد صالحة للاستخدام.")
                    
                wallet = get_or_create_wallet(request.user)
                
                credited_amount = card.amount
                if card.currency != wallet.currency:
                    base_val = card.currency.to_base(card.amount, operation="deposit")
                    credited_amount = wallet.currency.from_base(base_val, operation="deposit")
                    
                credit_wallet(
                    wallet_id=wallet.id,
                    amount=credited_amount,
                    reference=f"recharge:{card.id}",
                    description=f"شحن بطاقة رصيد رقم {card.code}",
                    created_by=request.user,
                    metadata={
                        "source_amount": str(card.amount),
                        "source_currency": card.currency.code,
                        "recharge_card_code": card.code
                    },
                    source="recharge_card",
                    reason="شحن بطاقة رصيد"
                )
                
                card.status = RechargeCard.Status.REDEEMED
                card.redeemed_by = request.user
                card.redeemed_at = timezone.now()
                card.save()
                
                if card.currency != wallet.currency:
                    msg = f"تم شحن محفظتك بنجاح بقيمة {credited_amount:.2f} {wallet.currency.code} (ما يعادل {card.amount:.2f} {card.currency.code})."
                else:
                    msg = f"تم شحن محفظتك بنجاح بقيمة {card.amount:.2f} {card.currency.code}."
                messages.success(request, msg)
                
        except Exception as e:
            messages.error(request, f"فشل شحن الرصيد: {str(e)}")
            
    return redirect("dashboard_wallet")


def sso_transfer_view(request):
    active_store = getattr(request, "store", None)
    from django.contrib.auth import login, get_user_model
    from django.core import signing
    from django.http import HttpResponseBadRequest
    from urllib.parse import urlparse, urlunparse, quote
    from django.contrib import messages
    from django.shortcuts import redirect

    User = get_user_model()
    token = request.GET.get("token")
    next_url = request.GET.get("next")

    # 1. If a token is provided, log the user in
    if token:
        try:
            user_id = signing.loads(token, max_age=300)
            sso_user = User.all_objects.get(pk=user_id)

            user_to_login = sso_user

            if active_store:
                # Check if there is already a tenant-specific user record with this email
                tenant_user = User.all_objects.filter(email__iexact=sso_user.email, store=active_store).first()
                if tenant_user:
                    user_to_login = tenant_user
                else:
                    # If user is not the owner and not an admin/staff, create a tenant-specific user record
                    is_owner = (active_store.owner_id == sso_user.pk)
                    is_admin = sso_user.is_superuser or sso_user.is_staff or sso_user.role in ["super_admin", "admin"]
                    if not is_owner and not is_admin:
                        # Create a new isolated user record for this store
                        from django.utils.crypto import get_random_string
                        tenant_user = User.objects.create_user(
                            email=sso_user.email,
                            password=get_random_string(32),
                            first_name=sso_user.first_name,
                            last_name=sso_user.last_name,
                            store=active_store,
                            role=sso_user.role,
                            preferred_language=sso_user.preferred_language,
                            email_verified=True,
                            is_active=True
                        )
                        user_to_login = tenant_user

                # Verify if user_to_login belongs to this store
                from apps.stores.models import StoreEmployee
                is_store_member = False
                
                # Check roles that are allowed everywhere
                if user_to_login.is_superuser or user_to_login.is_staff or getattr(user_to_login, "role", None) == "super_admin":
                    is_store_member = True
                else:
                    is_store_member = (
                        user_to_login.store_id == active_store.pk
                        or active_store.owner_id == user_to_login.pk
                        or StoreEmployee.objects.filter(store=active_store, user=user_to_login).exists()
                    )
                
                if not is_store_member:
                    messages.error(request, "هذا الحساب غير مرتبط بهذا المتجر.")
                    return redirect("site_login")

            user_to_login.backend = "apps.stores.auth_backend.TenantModelBackend"
            login(request, user_to_login)

            messages.success(request, "تم مزامنة تسجيل الدخول بنجاح.")
            return redirect(next_url or "/dashboard/")
        except (signing.SignatureExpired, signing.BadSignature, User.DoesNotExist) as e:
            messages.error(request, "رابط تسجيل الدخول غير صالح أو منتهي الصلاحية. يرجى المحاولة مرة أخرى.")
            return redirect("site_login")

    # 2. If user is authenticated, handle cross-domain redirect with a token
    if request.user.is_authenticated:
        if not next_url:
            return redirect("/dashboard/")

        parsed = urlparse(next_url)
        if parsed.netloc and parsed.netloc != request.get_host():
            token = signing.dumps(request.user.id)
            callback_path = "/auth/sso-callback/"
            target_path = parsed.path
            if parsed.query:
                target_path += f"?{parsed.query}"

            sso_url = urlunparse((
                parsed.scheme or "http",
                parsed.netloc,
                callback_path,
                "",
                f"token={token}&next={target_path or '/dashboard/'}",
                ""
            ))
            return redirect(sso_url)
        else:
            return redirect(next_url)

    # 3. If user is not authenticated, redirect to login page
    if active_store:
        from django.conf import settings
        platform_url = getattr(settings, "SITE_URL", "https://raqamiyatapp.com")
        current_absolute_uri = request.build_absolute_uri(next_url or "/dashboard/")
        sso_login_url = f"{platform_url}/accounts/google/login/?next={platform_url}/auth/sso-callback/%3Fnext%3D{quote(current_absolute_uri)}"
        return redirect(sso_login_url)
    else:
        return redirect("site_login")


@support_required
def control_alkasr_dashboard(request):
    from apps.orders.alkasr_api import get_alkasr_profile, get_alkasr_categories, get_alkasr_products, sync_alkasr_catalog
    from django.conf import settings
    from django.contrib import messages
    from urllib.parse import quote
    from apps.catalog.models import Product, ProductVariant
    
    store = getattr(request, "store", None)
    
    # Check if a POST action was submitted (e.g. to sync)
    if request.method == "POST":
        action = request.POST.get("action")
        if action == "sync":
            category_ids = request.POST.getlist("categories")
            # Convert categories to list of ints
            try:
                category_ids = [int(cid) for cid in category_ids if cid.isdigit()]
            except ValueError:
                category_ids = None
                
            markup_val = request.POST.get("markup_percent", "0.0")
            try:
                markup_percent = float(markup_val)
            except ValueError:
                markup_percent = 0.0
                
            # Perform sync
            res = sync_alkasr_catalog(store, selected_category_ids=category_ids, markup_percent=markup_percent)
            if res and res.get("status") == "success":
                messages.success(request, f"تمت عملية المزامنة بنجاح! تم استيراد {res['created']} منتج جديد، وتحديث {res['updated']} منتج.")
            else:
                error_msg = res.get('message') if res else "Unknown error"
                messages.error(request, f"فشلت عملية المزامنة: {error_msg}")
            return redirect("control_alkasr_dashboard")
            
        elif action == "refresh_cache":
            get_alkasr_profile(store=store, force_refresh=True)
            get_alkasr_categories(store=store, force_refresh=True)
            get_alkasr_products(store=store, force_refresh=True)
            messages.success(request, "تم تحديث التخزين المؤقت للبيانات وسحب كتالوج جديد بنجاح من المزود.")
            return redirect("control_alkasr_dashboard")
            
        elif action == "toggle_active":
            product_id = request.POST.get("product_id")
            try:
                prod = Product.objects.get(id=product_id, store=store)
                prod.is_active = not prod.is_active
                prod.save()
                status_str = "تفعيل" if prod.is_active else "تعطيل"
                messages.success(request, f"تم {status_str} المنتج '{prod.name}' بنجاح.")
            except Product.DoesNotExist:
                messages.error(request, "المنتج غير موجود.")
            return redirect("control_alkasr_dashboard")
            
        elif action == "delete_product":
            product_id = request.POST.get("product_id")
            try:
                prod = Product.objects.get(id=product_id, store=store)
                prod_name = prod.name
                prod.delete()
                messages.success(request, f"تم حذف وإلغاء ربط المنتج '{prod_name}' بنجاح.")
            except Product.DoesNotExist:
                messages.error(request, "المنتج غير موجود.")
            return redirect("control_alkasr_dashboard")
            
    # Fetch Alkasr Profile Info
    profile = get_alkasr_profile(store=store)
    is_connected = profile and profile.get("status") != "error"
    
    # Fetch Alkasr Categories and Products
    categories = []
    products_count = 0
    alkasr_products = []
    local_linked_count = 0
    
    if is_connected:
        categories = get_alkasr_categories(store=store)
        if isinstance(categories, dict) and categories.get("status") == "error":
            categories = []
            
        all_prods = get_alkasr_products(store=store)
        if isinstance(all_prods, list):
            products_count = len(all_prods)
            alkasr_products = all_prods
            
            # Map linked status and details
            if store:
                linked_variants_qs = ProductVariant.objects.filter(product__store=store, api_product_id__isnull=False)
            else:
                linked_variants_qs = ProductVariant.objects.filter(api_product_id__isnull=False)

            linked_variants = {
                v.api_product_id: {
                    "product_id": v.product.id,
                    "price": float(v.price),
                    "cost": float(v.cost),
                    "is_active": v.product.is_active,
                    "api_provider": v.product.api_provider or "alkasr"
                }
                for v in linked_variants_qs.select_related('product')
            }
            local_linked_count = len(linked_variants)
            
            for item in alkasr_products:
                item_id = item.get("id")
                if item_id in linked_variants:
                    item["is_linked"] = True
                    item["local_product_id"] = linked_variants[item_id]["product_id"]
                    item["local_price"] = linked_variants[item_id]["price"]
                    item["local_cost"] = linked_variants[item_id]["cost"]
                    item["local_active"] = linked_variants[item_id]["is_active"]
                    item["api_provider"] = linked_variants[item_id]["api_provider"]
                else:
                    item["is_linked"] = False
        else:
            alkasr_products = []
            
    # Calculate statistics from completed store orders
    from apps.orders.models import Order
    api_orders = Order.objects.filter(
        store=store,
        status__in=[Order.Status.COMPLETED, Order.Status.PROCESSING],
        items__variant__api_product_id__isnull=False
    ).distinct()
    
    total_purchases_usd = 0.0
    total_sales_usd = 0.0
    
    provider_stats = {
        "alkasr": {"name": "Alkasr VIP", "purchases": 0.0, "sales": 0.0, "profit": 0.0, "count": 0},
        "smm": {"name": "SMM Provider", "purchases": 0.0, "sales": 0.0, "profit": 0.0, "count": 0},
        "other": {"name": "Other API", "purchases": 0.0, "sales": 0.0, "profit": 0.0, "count": 0},
    }
    
    for order in api_orders:
        for item in order.items.select_related('variant__product'):
            variant = item.variant
            if not variant or not variant.api_product_id:
                continue
            qty = item.quantity
            item_cost = float(variant.cost or 0.0) * qty
            item_sales = float(item.price or 0.0) * qty
            
            total_purchases_usd += item_cost
            total_sales_usd += item_sales
            
            provider = variant.product.api_provider or "alkasr"
            if provider in provider_stats:
                provider_stats[provider]["purchases"] += item_cost
                provider_stats[provider]["sales"] += item_sales
                provider_stats[provider]["count"] += 1
                
    for p in provider_stats:
        provider_stats[p]["profit"] = provider_stats[p]["sales"] - provider_stats[p]["purchases"]
        
    total_profit_usd = total_sales_usd - total_purchases_usd
            
    # Generate webhook URL
    webhook_url = request.build_absolute_uri('/api/orders/alkasr_webhook/')
    
    # Retrieve active credentials from database APIIntegration model instead of settings.py
    from apps.orders.alkasr_api import get_alkasr_integration
    integration = get_alkasr_integration(store)
    base_url = integration.base_url if integration else ""
    raw_token = integration.api_token if integration else ""
    
    if len(raw_token) > 10:
        obfuscated_token = raw_token[:6] + "..." + raw_token[-6:]
    else:
        obfuscated_token = "غير معين أو قصير"
        
    return render(request, "site/control_alkasr_dashboard.html", {
        "profile": profile if is_connected else None,
        "is_connected": is_connected,
        "categories": categories,
        "products_count": products_count,
        "alkasr_products": alkasr_products[:500], # display up to 500 products for quick client-side filtering
        "webhook_url": webhook_url,
        "base_url": base_url,
        "obfuscated_token": obfuscated_token,
        "local_linked_count": local_linked_count,
        "total_purchases_usd": total_purchases_usd,
        "total_sales_usd": total_sales_usd,
        "total_profit_usd": total_profit_usd,
        "provider_stats": provider_stats,
    })


@support_required
def control_audit_logs(request):
    from apps.common.models import SystemAuditLog
    from django.core.paginator import Paginator
    
    store = getattr(request, "store", None)
    if store:
        logs = SystemAuditLog.objects.filter(actor__store=store).select_related('actor', 'content_type')
    else:
        logs = SystemAuditLog.objects.filter(actor__store__isnull=True).select_related('actor', 'content_type')
        
    # Search and filtering
    q = request.GET.get("q", "").strip()
    if q:
        logs = logs.filter(
            Q(action_type__icontains=q) |
            Q(description__icontains=q) |
            Q(actor__email__icontains=q) |
            Q(reason__icontains=q)
        )
        
    action_filter = request.GET.get("action_type", "").strip()
    if action_filter:
        logs = logs.filter(action_type=action_filter)
        
    # Distinct actions for dropdown
    # We do a quick list scan from the last 1000 logs to prevent heavy queries on large database
    distinct_actions = set(logs[:1000].values_list('action_type', flat=True))
    
    # Paginate logs to 40 per page
    paginator = Paginator(logs, 40)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    return render(request, "site/control_audit_logs.html", {
        "page_obj": page_obj,
        "distinct_actions": sorted(list(distinct_actions)),
        "query": q,
        "action_filter": action_filter
    })


@support_required
def control_api_integrations_list(request):
    from apps.catalog.models import APIIntegration
    store = getattr(request, "store", None)
    
    # List private integrations AND global ones that have allow_sub_stores=True
    if store:
        integrations = APIIntegration.objects.filter(
            Q(store=store) | Q(store__isnull=True, allow_sub_stores=True)
        )
    else:
        integrations = APIIntegration.objects.all()
        
    return render(request, "site/control_api_integrations_list.html", {
        "integrations": integrations,
        "is_tenant": bool(store),
    })


@support_required
def control_api_integration_create(request):
    from apps.site.forms import APIIntegrationForm
    store = getattr(request, "store", None)
    
    form = APIIntegrationForm(request.POST or None)
    if request.method == "POST" and form.is_valid():
        integration = form.save(commit=False)
        integration.store = store
        integration.save()
        messages.success(request, "تمت إضافة إعدادات ربط الـ API الجديد بنجاح.")
        return redirect("control_api_integrations_list")
        
    return render(request, "site/control_api_integration_form.html", {
        "form": form,
        "title": "إضافة بوابة ربط API جديدة",
    })


@support_required
def control_api_integration_edit(request, pk):
    from apps.catalog.models import APIIntegration
    from apps.site.forms import APIIntegrationForm
    store = getattr(request, "store", None)
    
    try:
        if store:
            integration = APIIntegration.objects.get(pk=pk, store=store)
        else:
            integration = APIIntegration.objects.get(pk=pk)
    except APIIntegration.DoesNotExist:
        messages.error(request, "عذراً، لا تملك الصلاحية لتعديل هذه البوابة.")
        return redirect("control_api_integrations_list")
        
    form = APIIntegrationForm(request.POST or None, instance=integration)
    if request.method == "POST" and form.is_valid():
        form.save()
        messages.success(request, "تم تحديث إعدادات الربط بنجاح.")
        return redirect("control_api_integrations_list")
        
    return render(request, "site/control_api_integration_form.html", {
        "form": form,
        "title": "تعديل بوابة ربط API",
        "integration": integration,
    })


@support_required
def control_api_integration_delete(request, pk):
    from apps.catalog.models import APIIntegration
    store = getattr(request, "store", None)
    
    try:
        if store:
            integration = APIIntegration.objects.get(pk=pk, store=store)
        else:
            integration = APIIntegration.objects.get(pk=pk)
        
        name = integration.name
        integration.delete()
        messages.success(request, f"تم حذف بوابة الربط '{name}' بنجاح.")
    except APIIntegration.DoesNotExist:
        messages.error(request, "بوابة الربط غير موجودة أو لا تملك صلاحية حذفها.")
        
    return redirect("control_api_integrations_list")


